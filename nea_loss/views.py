"""
NEA Loss Analysis System - Views
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views import View
from django.views.generic import ListView, DetailView
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Sum, Avg, Count, Q, Exists, OuterRef
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
import json
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import datetime
import decimal

from .models import (
    NEAUser, LossReport, MonthlyLossData, MeterPoint, MeterReading,
    ConsumerCategory, EnergyUtilisation, ConsumerCount, FiscalYear,
    DistributionCenter, ProvincialOffice, Province, Notification, AuditLog,
    ProvincialReport, MonthlyMeterPointStatus, DCYearlyTarget, DCMonthlyTarget, Message,
    DCReportOverride, FeederRequest, DCSDetail, DCSOfficial, DCSFeeder, DCSConsumerType,
    DCSDetailEditRequest, DCHistoryEntry, FeederFile,
)


def home_redirect(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')


def logout_view(request):
    logout(request)
    return redirect('login')


# ─────────────────────────── AUTH VIEWS ───────────────────────────

class LoginView(View):
    template_name = 'nea_loss/login.html'

    def get(self, request):
        if request.user.is_authenticated:
            return redirect('dashboard')
        return render(request, self.template_name)

    def post(self, request):
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            user.last_login_ip = request.META.get('REMOTE_ADDR')
            user.save(update_fields=['last_login_ip'])
            AuditLog.objects.create(
                user=user, action='LOGIN', model_name='NEAUser',
                object_id=user.pk, description=f"User {user.username} logged in",
                ip_address=request.META.get('REMOTE_ADDR')
            )
            return redirect('dashboard')
        messages.error(request, 'Invalid username or password.')
        return render(request, self.template_name)


class ProfileView(LoginRequiredMixin, View):
    template_name = 'nea_loss/profile.html'

    def get(self, request):
        return render(request, self.template_name, {'user': request.user})

    def post(self, request):
        user = request.user
        user.full_name = request.POST.get('full_name', user.full_name)
        user.phone = request.POST.get('phone', user.phone)
        user.designation = request.POST.get('designation', user.designation)
        user.email = request.POST.get('email', user.email)

        new_password = request.POST.get('new_password')
        if new_password:
            old_password = request.POST.get('old_password')
            if user.check_password(old_password):
                user.set_password(new_password)
                messages.success(request, 'Password updated successfully.')
            else:
                messages.error(request, 'Old password is incorrect.')
                return render(request, self.template_name, {'user': user})

        user.save()
        messages.success(request, 'Profile updated successfully.')
        return redirect('profile')


# ─────────────────────────── DASHBOARD ───────────────────────────

class DashboardView(LoginRequiredMixin, View):
    template_name = 'nea_loss/dashboard.html'

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        context = {'active_fy': active_fy}

        if getattr(user, 'is_system_admin', False):
            context.update(self._get_admin_context(active_fy))
        elif user.is_top_management:
            context.update(self._get_top_management_context(active_fy))
        elif user.is_provincial:
            context.update(self._get_provincial_context(user, active_fy))
        else:
            context.update(self._get_dc_context(user, active_fy))

        context['notifications'] = Notification.objects.filter(
            recipient=user, is_read=False
        ).order_by('-created_at')[:5]

        return render(request, self.template_name, context)

    def _get_admin_context(self, active_fy):
        # Reuse the top-management dashboard numbers, then overlay admin-specific controls.
        base = self._get_top_management_context(active_fy)

        reports = LossReport.objects.filter(fiscal_year=active_fy) if active_fy else LossReport.objects.none()
        total_received = reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
        total_utilised = reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
        total_loss = total_received - total_utilised
        overall_loss_pct = (total_loss / total_received * 100) if total_received > 0 else 0

        # Override management data for admin
        pending_overrides = DCReportOverride.objects.filter(status='PENDING').order_by('-created_at')
        recent_overrides = DCReportOverride.objects.all().order_by('-created_at')[:10]

        # Ensure all required context variables are properly initialized
        admin_context = {
            'admin_total_users': NEAUser.objects.count(),
            'admin_reports_total': reports.count(),
            'admin_recent_audits': AuditLog.objects.select_related('user').order_by('-timestamp')[:12],
            'admin_quicklink_active_fy': active_fy.year_bs if active_fy else '',
            'admin_overall_loss_pct': round(overall_loss_pct, 2),
            'admin_pending_overrides': pending_overrides,
            'admin_recent_overrides': recent_overrides,
            'mgmtProvData': base.get('mgmtProvData', []),  # Use empty list for admin
            'prov_monthly_detail': base.get('prov_monthly_detail', {}),
            'monthly_trend': base.get('monthly_trend', []),
            'top_5_loss': base.get('top_5_loss', []),
            'bottom_5_loss': base.get('bottom_5_loss', []),
        }
        
        base.update(admin_context)
        return base

    def _get_top_management_context(self, active_fy):
        """MD/DMD/Director: interactive read-only dashboard — approved reports only."""
        MONTH_NAMES = {
            1:'Shrawan',2:'Bhadra',3:'Ashwin',4:'Kartik',
            5:'Mangsir',6:'Poush',7:'Magh',8:'Falgun',
            9:'Chaitra',10:'Baisakh',11:'Jestha',12:'Ashadh'
        }
        reports = LossReport.objects.filter(
            fiscal_year=active_fy, status='APPROVED'
        ).select_related('distribution_center','distribution_center__provincial_office')          if active_fy else LossReport.objects.none()

        total_received = reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
        total_utilised = reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
        total_loss = total_received - total_utilised
        overall_loss_pct = round(float(total_loss) / float(total_received) * 100, 4) if total_received else 0

        # Province-wise breakdown
        prov_data = []
        if active_fy and reports.exists():
            for po in ProvincialOffice.objects.prefetch_related('distribution_centers').all():
                po_reports = reports.filter(distribution_center__provincial_office=po)
                po_recv = float(po_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0)
                po_util = float(po_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0)
                po_loss = po_recv - po_util
                po_pct  = round(po_loss / po_recv * 100, 4) if po_recv else 0
                dc_count = po_reports.values('distribution_center').distinct().count()
                prov_data.append({
                    'name': po.name,
                    'loss_pct': po_pct,
                    'received': po_recv,
                    'loss_kwh': po_loss,
                    'dc_count': dc_count,
                    'approved_count': po_reports.count(),
                })

        # DC-level details for interactive table (all approved DCs)
        dc_table = []
        if active_fy and reports.exists():
            for r in reports.order_by('distribution_center__provincial_office__name',
                                       'distribution_center__name'):
                dc_table.append({
                    'dc_name':   r.distribution_center.name,
                    'dc_code':   r.distribution_center.code,
                    'province':  r.distribution_center.provincial_office.name,
                    'month':     r.get_month_display(),
                    'month_num': r.month,
                    'received':  float(r.total_received_kwh),
                    'utilised':  float(r.total_utilised_kwh),
                    'loss_kwh':  float(r.total_loss_kwh),
                    'loss_pct':  round(float(r.cumulative_loss_percent) * 100, 4),
                    'status':    r.status,
                    'report_pk': r.pk,
                })

        # Month-wise aggregated trend (across all DCs)
        monthly_trend = {}
        if active_fy and reports.exists():
            all_monthly = MonthlyLossData.objects.filter(
                report__in=reports
            ).values('month','month_name').annotate(
                tot_recv=Sum('net_energy_received'),
                tot_util=Sum('total_energy_utilised'),
                tot_loss=Sum('loss_unit'),
            ).order_by('month')
            for row in all_monthly:
                recv = float(row['tot_recv'] or 0)
                loss = float(row['tot_loss'] or 0)
                monthly_trend[row['month_name']] = {
                    'received': recv,
                    'utilised': float(row['tot_util'] or 0),
                    'loss': loss,
                    'loss_pct': round(loss/recv*100,4) if recv else 0,
                }

        # Top/bottom 5 DCs by loss %
        sorted_dc = sorted(dc_table, key=lambda x: x['loss_pct'], reverse=True)
        top_5_loss    = sorted_dc[:5]
        bottom_5_loss = sorted_dc[-5:] if len(sorted_dc) >= 5 else sorted_dc

        # Province list for filter dropdown
        all_provinces = list(ProvincialOffice.objects.values_list('name', flat=True))

        # Build per-DC monthly detail for sidebar browsing
        # Each DC: list of months with received, utilised, loss_unit, monthly_loss_pct, cumul_loss_pct
        dc_monthly_detail = {}  # dc_name -> list of month dicts
        all_monthly_data = (
            MonthlyLossData.objects
            .filter(report__in=reports)
            .select_related('report__distribution_center','report__distribution_center__provincial_office')
            .order_by('report__distribution_center__name', 'month')
        )
        
        # Group by DC and calculate cumulative correctly
        from collections import defaultdict
        dc_data = defaultdict(list)  # dc_name -> list of monthly data
        for md in all_monthly_data:
            dc_name = md.report.distribution_center.name
            dc_data[dc_name].append(md)
        
        # Calculate cumulative loss for each DC
        for dc_name, monthly_data_list in dc_data.items():
            dc_code = monthly_data_list[0].report.distribution_center.code
            province = monthly_data_list[0].report.distribution_center.provincial_office.name
            report_pk = monthly_data_list[0].report.pk
            
            cumulative_received = 0
            cumulative_utilised = 0
            
            dc_monthly_detail[dc_name] = {
                'dc_name': dc_name,
                'dc_code': dc_code,
                'province': province,
                'report_pk': report_pk,
                'months': [],
            }
            
            # Sort by month and calculate progressive cumulative
            monthly_data_list.sort(key=lambda x: x.month)
            for md in monthly_data_list:
                recv = float(md.net_energy_received)
                utilised = float(md.total_energy_utilised)
                loss = float(md.loss_unit)
                mpct = round(loss / recv * 100, 4) if recv else 0
                
                # Add to cumulative
                cumulative_received += recv
                cumulative_utilised += utilised
                
                # Calculate cumulative loss %
                cumulative_loss = cumulative_received - cumulative_utilised
                cpct = round(cumulative_loss / cumulative_received * 100, 4) if cumulative_received else 0
                
                dc_monthly_detail[dc_name]['months'].append({
                    'month_name': md.month_name,
                    'month': md.month,
                    'received': recv,
                    'utilised': utilised,
                    'loss_unit': loss,
                    'monthly_loss_pct': mpct,
                    'cumul_loss_pct': cpct,  # Now using correct progressive calculation
                })

        # Build per-province monthly detail for sidebar browsing
        prov_monthly_detail = {}
        for po in ProvincialOffice.objects.all():
            po_monthly_qs = MonthlyLossData.objects.filter(
                report__in=reports,
                report__distribution_center__provincial_office=po,
            ).values('month','month_name').annotate(
                tot_recv=Sum('net_energy_received'),
                tot_util=Sum('total_energy_utilised'),
                tot_loss=Sum('loss_unit'),
            ).order_by('month')
            months_list = []
            cum_recv = 0.0
            cum_loss = 0.0
            for row in po_monthly_qs:
                recv = float(row['tot_recv'] or 0)
                loss = float(row['tot_loss'] or 0)
                cum_recv += recv
                cum_loss += loss
                months_list.append({
                    'month_name': row['month_name'],
                    'month': row['month'],
                    'received': recv,
                    'utilised': float(row['tot_util'] or 0),
                    'loss_unit': loss,
                    'monthly_loss_pct': round(loss/recv*100,4) if recv else 0,
                    'cumul_loss_pct': round(cum_loss/cum_recv*100,4) if cum_recv else 0,
                })
            if months_list:
                prov_monthly_detail[po.name] = months_list

        # Build dc_report_table for Report Explorer (for top management) with correct cumulative loss calculation
        from nea_loss.models import DCYearlyTarget
        from collections import defaultdict
        
        # Load all monthly data for approved reports
        monthly_qs = MonthlyLossData.objects.filter(
            report__in=reports
        ).select_related('report__distribution_center').order_by('report__distribution_center__name', 'month')

        # Group monthly data by DC
        dc_monthly = defaultdict(dict)  # dc_id -> {month -> MonthlyLossData}
        for md in monthly_qs:
            dc_monthly[md.report.distribution_center_id][md.month] = md

        # Load provincial yearly targets
        targets = {}
        if active_fy:
            for t in DCYearlyTarget.objects.filter(fiscal_year=active_fy):
                targets[t.distribution_center_id] = float(t.target_loss_percent)

        # Build table rows
        dc_report_table = []
        # Only show months that have approved reports
        approved_months = sorted(set(
            md.month for md in monthly_qs
        ))
        
        # If no approved months, show empty
        if not approved_months:
            approved_months = []

        for dc in DistributionCenter.objects.filter(is_active=True).order_by('name'):
            dc_report = reports.filter(distribution_center=dc).first()
            month_rows = []
            
            # Calculate cumulative loss progressively for each approved month
            cumulative_received = 0
            cumulative_utilised = 0
            
            # Get all months in order and calculate cumulative progressively
            for m in approved_months:
                md = dc_monthly.get(dc.pk, {}).get(m)
                target = targets.get(dc.pk)  # Use yearly target for all months
                
                # Always calculate cumulative, but only add if DC has data for this month
                if md and md.net_energy_received:
                    cumulative_received += float(md.net_energy_received)
                if md and md.total_energy_utilised:
                    cumulative_utilised += float(md.total_energy_utilised)
                
                # Calculate cumulative loss % up to this month
                cumulative_loss = cumulative_received - cumulative_utilised
                cumulative_loss_pct = round(cumulative_loss / cumulative_received * 100, 4) if cumulative_received else 0
                
                # Only include month row if this DC has approved data for this month
                if md:  # Only show months where DC has data
                    month_rows.append({
                        'month': m,
                        'month_name': MONTH_NAMES.get(m, ''),
                        'received': float(md.net_energy_received) if md else None,
                        'sold': float(md.total_energy_utilised) if md else None,
                        'loss_unit': float(md.loss_unit) if md else None,
                        'monthly_loss_pct': round(float(md.monthly_loss_percent) * 100, 4) if md else None,
                        'cumulative_loss_pct': cumulative_loss_pct,  # This will show proper cumulative
                        'target': target,
                        'status': 'APPROVED',  # All shown data is approved
                    })
            dc_report_table.append({
                'dc': dc,
                'report': dc_report,
                'month_rows': month_rows,
                'has_data': any(r['received'] is not None for r in month_rows),
            })

        return {
            'total_received_kwh':   float(total_received),
            'total_utilised_kwh':   float(total_utilised),
            'total_loss_kwh':       float(total_loss),
            'overall_loss_pct':     overall_loss_pct,
            'total_dc_count':       DistributionCenter.objects.filter(is_active=True).count(),
            'reports_approved':     reports.count(),
            'reports_submitted':    LossReport.objects.filter(
                                        fiscal_year=active_fy,
                                        status__in=['SUBMITTED','PROVINCIAL_REVIEWED']
                                    ).count() if active_fy else 0,
            'reports_pending':      LossReport.objects.filter(
                                        fiscal_year=active_fy, status='DRAFT'
                                    ).count() if active_fy else 0,
            'dc_table':             dc_table,
            'provincial_data':      prov_data,
            'monthly_trend':        monthly_trend,
            'top_5_loss':           top_5_loss,
            'bottom_5_loss':        bottom_5_loss,
            'nea_target_pct':       float(active_fy.loss_target_percent) if active_fy else 3.35,
            'target_loss_pct':      float(active_fy.loss_target_percent) if active_fy else 3.35,
            'all_provinces':        all_provinces,
            'month_names_list':     list(MONTH_NAMES.values()),
            'dc_monthly_detail':    dc_monthly_detail,      # for sidebar DC browser
            'prov_monthly_detail':  prov_monthly_detail,    # for sidebar province browser
            'dc_report_table':      dc_report_table if 'dc_report_table' in locals() else [],  # Safe: only if exists
        }

    def _get_provincial_context(self, user, active_fy):
        po = user.provincial_office
        MONTH_NAMES = {
            1:'Shrawan',2:'Bhadra',3:'Ashwin',4:'Kartik',
            5:'Mangsir',6:'Poush',7:'Magh',8:'Falgun',
            9:'Chaitra',10:'Baisakh',11:'Jestha',12:'Ashadh'
        }
        all_reports = LossReport.objects.filter(
            fiscal_year=active_fy,
            distribution_center__provincial_office=po
        ).select_related('distribution_center') if active_fy else LossReport.objects.none()

        approved_reports = all_reports.filter(status='APPROVED')

        # Build month-wise summary table per DC (Excel format):
        # For each DC: one row per month with received, sold, loss units, monthly loss %, cumulative loss %
        # Plus provincial target per month if set
        dcs = DistributionCenter.objects.filter(provincial_office=po).order_by('name')

        # Load all monthly data for approved reports under this province
        from nea_loss.models import MonthlyLossData, DCYearlyTarget
        monthly_qs = MonthlyLossData.objects.filter(
            report__in=approved_reports
        ).select_related('report__distribution_center').order_by('report__distribution_center__name', 'month')

        # Group monthly data by DC
        from collections import defaultdict
        dc_monthly = defaultdict(dict)  # dc_id -> {month -> MonthlyLossData}
        for md in monthly_qs:
            dc_monthly[md.report.distribution_center_id][md.month] = md

        # Load provincial yearly targets
        targets = {}
        if active_fy:
            for t in DCYearlyTarget.objects.filter(
                fiscal_year=active_fy,
                distribution_center__provincial_office=po
            ):
                targets[t.distribution_center_id] = float(t.target_loss_percent)

        # Build table rows
        dc_report_table = []
        # Only show months that have approved reports
        # Since monthly_qs already filters approved_reports, we just need to get unique months
        approved_months = sorted(set(
            md.month for md in monthly_qs
        ))
        
        # If no approved months, show empty
        if not approved_months:
            approved_months = []

        for dc in dcs:
            dc_report = all_reports.filter(distribution_center=dc).first()
            month_rows = []
            
            # Calculate cumulative loss progressively for each approved month
            cumulative_received = 0
            cumulative_utilised = 0
            
            # Get all months in order and calculate cumulative progressively
            for m in approved_months:
                md = dc_monthly.get(dc.pk, {}).get(m)
                target = targets.get(dc.pk)  # Use yearly target for all months
                
                # Always calculate cumulative, but only add if DC has data for this month
                if md and md.net_energy_received:
                    cumulative_received += float(md.net_energy_received)
                if md and md.total_energy_utilised:
                    cumulative_utilised += float(md.total_energy_utilised)
                
                # Calculate cumulative loss % up to this month
                cumulative_loss = cumulative_received - cumulative_utilised
                cumulative_loss_pct = round(cumulative_loss / cumulative_received * 100, 4) if cumulative_received else 0
                
                # Only include month row if this DC has approved data for this month
                if md:  # Only show months where DC has data
                    month_rows.append({
                        'month': m,
                        'month_name': MONTH_NAMES.get(m, ''),
                        'received': float(md.net_energy_received) if md else None,
                        'sold': float(md.total_energy_utilised) if md else None,
                        'loss_unit': float(md.loss_unit) if md else None,
                        'monthly_loss_pct': round(float(md.monthly_loss_percent) * 100, 4) if md else None,
                        'cumulative_loss_pct': cumulative_loss_pct,  # This will show proper cumulative
                        'target': target,
                        'status': 'APPROVED',  # All shown data is approved
                    })
            dc_report_table.append({
                'dc': dc,
                'report': dc_report,
                'month_rows': month_rows,
                'has_data': any(r['received'] is not None for r in month_rows),
            })

        # Totals across all approved reports
        total_received = approved_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
        total_utilised = approved_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
        total_loss = float(total_received) - float(total_utilised)
        overall_loss_pct = round(total_loss / float(total_received) * 100, 4) if total_received else 0

        return {
            'provincial_office': po,
            'dc_count': dcs.count(),
            'reports_submitted': all_reports.filter(status__in=['SUBMITTED','PROVINCIAL_REVIEWED']).count(),
            'reports_approved': all_reports.filter(status='APPROVED').count(),
            'reports_rejected': all_reports.filter(status='REJECTED').count(),
            'pending_review': all_reports.filter(status__in=['SUBMITTED','PROVINCIAL_REVIEWED']).select_related('distribution_center').order_by('-updated_at')[:10],
            'total_received_kwh': float(total_received),
            'total_loss_kwh': total_loss,
            'overall_loss_pct': overall_loss_pct,
            'target_loss_pct': float(active_fy.loss_target_percent) if active_fy else 3.35,
            'dc_report_table': dc_report_table,
            'all_months': [MONTH_NAMES[m] for m in approved_months],
            'all_month_nums': approved_months,
        }

    def _get_dc_context(self, user, active_fy):
        dc = user.distribution_center
        report = LossReport.objects.filter(
            distribution_center=dc, fiscal_year=active_fy
        ).first() if (dc and active_fy) else None

        MONTH_NAMES = {
            1:'Shrawan', 2:'Bhadra', 3:'Ashwin', 4:'Kartik',
            5:'Mangsir', 6:'Poush', 7:'Magh', 8:'Falgun',
            9:'Chaitra', 10:'Baisakh', 11:'Jestha', 12:'Ashadh'
        }
        ALL_MONTHS = list(range(1, 13))

        # Provincial yearly targets for this DC
        prov_targets = {}
        if active_fy and dc:
            yearly_target = DCYearlyTarget.objects.filter(
                distribution_center=dc, fiscal_year=active_fy
            ).first()
            prov_targets = float(yearly_target.target_loss_percent) if yearly_target else None

        # Initialize variables
        dc_monthly_cols = []
        approved_months = []
        ytd_received = 0
        ytd_loss = 0
        ytd_sold = 0
        ytd_loss_pct = 0

        # Build columns from ALL approved reports for current DC
        approved_reports = LossReport.objects.filter(
            distribution_center=dc,
            fiscal_year=active_fy,
            status='APPROVED'
        ).order_by('created_at')
        
        if approved_reports:
            # Collect all monthly data from ALL approved reports
            all_monthly_data = []
            for report in approved_reports:
                all_monthly_data.extend(report.monthly_data.all())
            
            # Sort by month number to ensure proper order
            all_monthly_data.sort(key=lambda x: x.month)
            
            cumulative_received = decimal.Decimal('0')
            cumulative_loss = decimal.Decimal('0')
            
            # Build feeder readings for all approved reports
            month_pks = [md.pk for md in all_monthly_data]
            deleted_pairs = set(
                MonthlyMeterPointStatus.objects.filter(
                    monthly_data_id__in=month_pks, is_active=False
                ).values_list('monthly_data_id', 'meter_point_id')
            )
            all_readings = (
                MeterReading.objects
                .filter(monthly_data__in=all_monthly_data)
                .select_related('meter_point')
                .order_by('meter_point__source_type', 'meter_point__name')
            )
            reading_map = {(r.monthly_data_id, r.meter_point_id): r for r in all_readings}

            import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT'}
            export_types = {'EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT'}
            
            # Get all active feeders first
            all_feeders = list(
                MeterPoint.objects.filter(
                    distribution_center=dc,
                    is_active=True,
                    source_type__in=list(import_types | export_types)
                ).order_by('source_type', 'name')
            )

            # Build columns for ALL approved months from ALL reports
            for md in all_monthly_data:
                received   = float(md.net_energy_received)
                sold       = float(md.total_energy_utilised)
                import_kwh = float(md.total_energy_import)
                export_kwh = float(md.total_energy_export)
                loss_unit  = float(md.loss_unit)

                monthly_loss_pct = round(loss_unit / received * 100, 4) if received else 0
                cumulative_received += md.net_energy_received
                cumulative_loss     += md.loss_unit
                cum_loss_pct = round(
                    float(cumulative_loss) / float(cumulative_received) * 100, 4
                ) if cumulative_received else 0

                prov_target = prov_targets

                # Get meter points that have readings in this month or prior months of this fiscal year
                # This ensures feeders only appear from the month they were first used onwards
                if active_fy:
                    meter_points_with_readings = set(
                        MeterReading.objects.filter(
                            meter_point__distribution_center=dc,
                            monthly_data__report__fiscal_year=active_fy,
                            monthly_data__month__lte=md.month
                        ).values_list('meter_point_id', flat=True).distinct()
                    )
                    # Filter all_feeders to only include those with readings in this month or prior months
                    month_feeders = [mp for mp in all_feeders if mp.pk in meter_points_with_readings]
                else:
                    # Fallback if no fiscal year: show all active feeders
                    month_feeders = all_feeders

                # Build feeder list for this month
                feeders = []
                for mp in month_feeders:
                    if (md.pk, mp.pk) in deleted_pairs:
                        continue
                    r = reading_map.get((md.pk, mp.pk))
                    feeders.append({
                        'name': mp.name,
                        'type': mp.get_source_type_display(),
                        'is_export': mp.source_type in export_types,
                        'prev': float(r.previous_reading) if r else None,
                        'pres': float(r.present_reading)  if r else None,
                        'mf':   float(r.multiplying_factor) if r else 1,
                        'kwh':  float(r.unit_kwh) if r else None,
                    })

                dc_monthly_cols.append({
                    'month':            md.month,
                    'month_name':       MONTH_NAMES[md.month],
                    'has_data':         True,
                    'import_kwh':       import_kwh,
                    'export_kwh':       export_kwh,
                    'received':         received,
                    'sold':             sold,
                    'loss_unit':        loss_unit,
                    'monthly_loss_pct': monthly_loss_pct,
                    'cumulative_loss_pct': cum_loss_pct,
                    'prov_target':      prov_target,
                    'status':           'APPROVED',
                    'report_pk':        md.report.pk,
                    'feeders':          feeders,
                })

            # YTD totals from all approved months
            ytd_received = float(cumulative_received)
            ytd_loss     = float(cumulative_loss)
            ytd_sold     = sum(c['sold'] for c in dc_monthly_cols if c['sold'] is not None)
            ytd_loss_pct = round(ytd_loss / ytd_received * 100, 4) if ytd_received else 0
            
            # Approved months are the columns we just built
            approved_months = dc_monthly_cols

        return {
            'distribution_center':    dc,
            'current_report':         report,
            'dc_monthly_cols':        dc_monthly_cols,
            'approved_months':        approved_months,
            'nea_target_pct':         float(active_fy.loss_target_percent) if active_fy else None,
            'ytd_received':           ytd_received,
            'ytd_sold':               ytd_sold,
            'ytd_loss':               ytd_loss,
            'ytd_loss_pct':           ytd_loss_pct,
            'prov_targets':           prov_targets,
            'can_create_loss_report': _can_create_loss_report(user),
        }


# ─────────────────────────── REPORT VIEWS ───────────────────────────

class ReportListView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/list.html'

    def get(self, request):
        user = request.user
        reports = LossReport.objects.select_related(
            'distribution_center', 'distribution_center__provincial_office',
            'fiscal_year', 'created_by'
        ).order_by('-fiscal_year__year_ad_start', 'distribution_center__name')

        if user.is_dc_level and user.distribution_center:
            # DC users see all their own reports (including drafts)
            reports = reports.filter(distribution_center=user.distribution_center)
        elif user.is_provincial:
            if user.provincial_office:
                # Provincial sees only SUBMITTED and above (not DRAFT) from their DCs
                reports = reports.filter(
                    distribution_center__provincial_office=user.provincial_office,
                    status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED', 'APPROVED', 'REJECTED']
                )
            else:
                reports = reports.none()
        elif user.is_top_management:
            # MD/DMD/Director sees only APPROVED reports
            reports = reports.filter(status='APPROVED')
        elif getattr(user, 'is_system_admin', False):
            # System admin sees everything
            pass

        # Filters
        fy_id = request.GET.get('fiscal_year')
        status = request.GET.get('status')
        dc_id = request.GET.get('dc')
        search = request.GET.get('search')
        month_filter = request.GET.get('month', '')

        if fy_id:
            reports = reports.filter(fiscal_year_id=fy_id)
        if status:
            reports = reports.filter(status=status)
        if dc_id and not user.is_dc_level:
            reports = reports.filter(distribution_center_id=dc_id)
        if search:
            reports = reports.filter(distribution_center__name__icontains=search)
        if month_filter:
            try:
                reports = reports.filter(month=int(month_filter))
            except (ValueError, TypeError):
                pass

        paginator = Paginator(reports, 20)
        page = paginator.get_page(request.GET.get('page', 1))

        # UI helper: show Edit action only when the report is editable.
        for r in page:
            setattr(r, 'can_edit', _can_edit_report(request.user, r))

        # DC-level users cannot filter by other DCs — hide the dropdown entirely
        if user.is_dc_level:
            visible_dcs = []  # No DC filter shown to DC users
        elif user.is_provincial and user.provincial_office:
            visible_dcs = DistributionCenter.objects.filter(
                provincial_office=user.provincial_office, is_active=True
            )
        else:
            visible_dcs = DistributionCenter.objects.filter(is_active=True)

        MONTH_CHOICES = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]

        return render(request, self.template_name, {
            'reports': page,
            'fiscal_years': FiscalYear.objects.all().order_by('-year_ad_start'),
            'distribution_centers': visible_dcs,
            'status_choices': LossReport.STATUS_CHOICES,
            'month_choices': MONTH_CHOICES,
            'selected_fy': fy_id,
            'selected_status': status,
            'selected_month': month_filter,
        })


class ReportCreateView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/create.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            # Check if user is DC level and needs admin override
            if request.user.is_dc_level and not _can_create_loss_report(request.user):
                # Check if admin override should be used
                if _can_admin_override_create_report(request.user):
                    # Use admin override function
                    return super().dispatch(request, *args, **kwargs)
                else:
                    # Use normal DC restrictions
                    messages.error(
                        request,
                        'Only distribution center and provincial office users can create loss reports. '
                        'Head office roles can view reports, analytics, and approvals from the menu.',
                    )
                    return redirect('report_list')
            else:
                # Non-DC users use normal restrictions
                if not _can_create_loss_report(request.user):
                    messages.error(
                        request,
                        'Only distribution center and provincial office users can create loss reports. '
                        'Head office roles can view reports, analytics, and approvals from the menu.',
                    )
                    return redirect('report_list')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        
        dcs = DistributionCenter.objects.all()
        if user.is_dc_level and user.distribution_center:
            dcs = dcs.filter(pk=user.distribution_center.pk)
        elif user.is_provincial and user.provincial_office:
            # Provincial users should not create reports - disable DC selection
            dcs = DistributionCenter.objects.none()
            messages.info(
                request,
                'Provincial office reports are generated automatically from DC reports under your province. '
                'You can review and approve/reject reports from the dashboard.'
            )
            return redirect('report_list')

        # Show all months from start month onwards with status indicators
        all_months = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]
        
        # Build month list with status information
        months_with_status = []
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        
        for month_num, month_name in all_months:
            month_info = {
                'month_num': month_num,
                'month_name': month_name,
                'can_create': False,
                'exists': False,
                'needs_override': False,
                'has_override': False,
                'status': '',
                'message': ''
            }
            
            # Skip months before the DC's start month
            if active_fy and dcs.exists():
                dc_start_month = dcs.first().report_start_month
                if month_num < dc_start_month:
                    month_info['status'] = 'before_start'
                    month_info['message'] = f'Before start month ({dc_start_month})'
                    months_with_status.append(month_info)
                    continue
            
            # Check if report already exists for this month
            if active_fy:
                existing_report = LossReport.objects.filter(
                    distribution_center__in=dcs,
                    fiscal_year=active_fy,
                    month=month_num
                ).first()
                
                if existing_report:
                    month_info['exists'] = True
                    month_info['status'] = 'exists'
                    month_info['message'] = f'Report exists ({existing_report.get_status_display()})'
                    months_with_status.append(month_info)
                    continue
            
            # Check if this month needs an override or can be created normally
            if month_num > 1:
                # Check if there's an approved and active override for this month
                has_override = False
                for dc in dcs:
                    override = DCReportOverride.objects.filter(
                        distribution_center=dc,
                        fiscal_year=active_fy,
                        status='APPROVED',
                        resume_month=month_num,
                        is_active=True
                    ).first()
                    if override:
                        has_override = True
                        break
                
                if has_override:
                    month_info['can_create'] = True
                    month_info['has_override'] = True
                    month_info['status'] = 'override_approved'
                    month_info['message'] = 'Available (override approved)'
                else:
                    # Check if previous month is approved
                    previous_month = month_num - 1
                    previous_report = LossReport.objects.filter(
                        distribution_center__in=dcs,
                        fiscal_year=active_fy,
                        month=previous_month,
                        status='APPROVED'
                    ).first()
                    
                    if previous_report or month_num == dc_start_month:
                        # Previous month is approved or this is start month - can create
                        month_info['can_create'] = True
                        month_info['status'] = 'available'
                        month_info['message'] = 'Available'
                    else:
                        # Previous month not approved - needs override
                        month_info['needs_override'] = True
                        month_info['status'] = 'needs_override'
                        month_info['message'] = 'Needs override (previous month not approved)'
            else:
                # Shrawan (month 1) - always available from start month
                month_info['can_create'] = True
                month_info['status'] = 'available'
                month_info['message'] = 'Available (first month)'
            
            months_with_status.append(month_info)
        
        months_list = [(m['month_num'], m['month_name']) for m in months_with_status]
        
        return render(request, self.template_name, {
            'fiscal_years': FiscalYear.objects.all(),
            'distribution_centers': dcs,
            'months_list': months_list,
            'months_with_status': months_with_status,
        })

    def post(self, request):
        user = request.user
        fy_id = request.POST.get('fiscal_year')
        dc_id = request.POST.get('distribution_center')
        month_id = request.POST.get('month')

        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }

        if not fy_id or not dc_id or not month_id:
            messages.error(request, 'Please select distribution center, fiscal year, and month.')
            return self.get(request)

        try:
            fy = FiscalYear.objects.get(pk=fy_id)
            dc = DistributionCenter.objects.get(pk=dc_id)
            month = int(month_id)
        except (FiscalYear.DoesNotExist, DistributionCenter.DoesNotExist):
            messages.error(request, 'Selected fiscal year or distribution center is invalid.')
            return self.get(request)
        except (ValueError, TypeError):
            messages.error(request, 'Selected month is invalid.')
            return self.get(request)

        if month < 1 or month > 12:
            messages.error(request, 'Selected month is invalid.')
            return self.get(request)

        # Check if previous month is approved (except for Shrawan and months before start month)
        if month > 1:  # Not Shrawan
            # If this month is >= the DC's start month, check previous month approval
            if month >= dc.report_start_month:
                previous_month = month - 1
                
                # Check if there's an approved and active override for this situation
                approved_override = DCReportOverride.objects.filter(
                    distribution_center=dc,
                    fiscal_year=fy,
                    status='APPROVED',
                    resume_month=month,
                    is_active=True
                ).first()
                
                if approved_override:
                    # Override exists - allow creation with a notification
                    messages.info(
                        request,
                        f'Using approved override to create {month_names.get(month, "")} report. '
                        f'Previous months may be skipped due to: {approved_override.reason[:100]}...'
                    )
                else:
                    # No override - check previous month approval normally
                    try:
                        previous_report = LossReport.objects.get(
                            distribution_center=dc,
                            fiscal_year=fy,
                            month=previous_month
                        )
                        if previous_report.status != 'APPROVED':
                            messages.error(
                                request,
                                f'Previous month report ({month_names.get(previous_month, "")}) must be approved before creating {month_names.get(month, "")} report. '
                                f'If you experienced technical issues, please contact your system administrator for an override.'
                            )
                            return self.get(request)
                    except LossReport.DoesNotExist:
                        messages.error(
                            request,
                            f'Previous month report ({month_names.get(previous_month, "")}) doesn\'t exist. Please create that first. '
                            f'If you experienced technical issues preventing previous submissions, contact your system administrator.'
                        )
                        return self.get(request)
            # If month < start month, skip previous month check completely

        allowed = DistributionCenter.objects.all()
        if user.is_dc_level and user.distribution_center:
            allowed = allowed.filter(pk=user.distribution_center.pk)
        elif user.is_provincial and user.provincial_office:
            allowed = allowed.filter(provincial_office=user.provincial_office)
        else:
            allowed = allowed.none()
        if not allowed.filter(pk=dc.pk).exists():
            messages.error(request, 'You cannot create a report for this distribution center.')
            return redirect('report_list')

        # If this month report already exists, open the data-entry directly.
        existing_report = LossReport.objects.filter(
            fiscal_year=fy,
            distribution_center=dc,
            month=month,
        ).first()
        if existing_report:
            month_name = month_names.get(month, '')
            messages.info(request, f'Report for {month_name} already exists. Opening data entry...')
            return redirect('monthly_data', existing_report.pk, month)

        # Create new monthly report
        report = LossReport.objects.create(
            fiscal_year=fy,
            distribution_center=dc,
            month=month,
            created_by=request.user,
            status='DRAFT'
        )
        AuditLog.objects.create(
            user=request.user,
            action='CREATE',
            model_name='LossReport',
            object_id=report.pk,
            description=f"Created monthly loss report for {dc.name} - {fy.year_bs} - {month_names.get(month, '')}",
        )
        messages.success(request, f'Monthly report created for {dc.name}.')
        return redirect('monthly_data', report.pk, month)


class ReportDetailView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/report_detail.html'

    def get(self, request, pk):
        report = get_object_or_404(LossReport, pk=pk)
        if not _can_view_report(request.user, report):
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('report_list')
        # Only get monthly data if it exists (don't auto-create)
        try:
            monthly_data = MonthlyLossData.objects.filter(report=report).order_by('month')
        except:
            monthly_data = MonthlyLossData.objects.none()
        fy = report.fiscal_year
        months_entered = monthly_data.count()

        # Monthly chart data
        chart_data = {
            'months': [m.month_name for m in monthly_data],
            'received': [float(m.net_energy_received) for m in monthly_data],
            'utilised': [float(m.total_energy_utilised) for m in monthly_data],
            'loss_pct': [round(float(m.monthly_loss_percent) * 100, 2) for m in monthly_data],
            'cumulative_pct': [round(float(m.cumulative_loss_percent) * 100, 2) for m in monthly_data],
        }

        dc = request.user.distribution_center if request.user.is_dc_level else None
        start_month = dc.report_start_month if dc else 1
        all_months = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]
        months_list = [(n, name) for n, name in all_months if n >= start_month]
        entered_months = [m.month for m in monthly_data]
        return render(request, self.template_name, {
            'report': report,
            'monthly_data': monthly_data,
            'months_entered': months_entered,
            'months_list': months_list,
            'entered_months': entered_months,
            'chart_data_json': json.dumps(chart_data),
            'target_pct': float(fy.loss_target_percent),
            'can_edit': _can_edit_report(request.user, report),
            'can_approve': _can_approve_report(request.user),
        })


class ReportReviewView(LoginRequiredMixin, View):
    """Minimal data-only view for reviewing reports without extra UI elements"""
    template_name = 'nea_loss/reports/report_review.html'

    def get(self, request, pk):
        report = get_object_or_404(LossReport, pk=pk)
        if not _can_view_report(request.user, report):
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('report_list')
        
        # Get monthly data
        monthly_data = MonthlyLossData.objects.filter(report=report).order_by('month')
        
        # Get DC yearly target for this fiscal year
        dc_yearly_target = None
        if report.fiscal_year:
            try:
                target = DCYearlyTarget.objects.get(
                    distribution_center=report.distribution_center,
                    fiscal_year=report.fiscal_year
                )
                dc_yearly_target = float(target.target_loss_percent)
            except DCYearlyTarget.DoesNotExist:
                dc_yearly_target = None
            except Exception:
                dc_yearly_target = None
        
        return render(request, self.template_name, {
            'report': report,
            'monthly_data': monthly_data,
            'can_approve': _can_approve_report(request.user),
            'dc_yearly_target': dc_yearly_target,
        })


class ReportEditView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/edit.html'

    def get(self, request, pk):
        report = get_object_or_404(LossReport, pk=pk)
        if not _can_edit_report(request.user, report):
            messages.error(request, 'You do not have permission to edit this report.')
            return redirect('report_detail', pk=pk)
        return render(request, self.template_name, {'report': report})


class MonthlyDataView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/monthly_data.html'

    def get(self, request, report_pk, month):
        report = get_object_or_404(LossReport, pk=report_pk)
        
        # Check if report is submitted/approved - normally view-only.
        # System admin and top management can edit APPROVED reports to make corrections.
        if _can_edit_report(request.user, report):
            can_edit = True
        elif _can_view_report(request.user, report):
            can_edit = False
        else:
            messages.error(request, 'You do not have permission to view this report.')
            return redirect('report_list')
            
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }
        
        # Get previous month's present readings for each meter point
        previous_readings_dict = {}
        
        if month > 1:  # For months after Shrawan
            previous_month = month - 1
            
            # Check if there's an approved and active override for this month that allows skipping previous months
            approved_override = DCReportOverride.objects.filter(
                distribution_center=report.distribution_center,
                fiscal_year=report.fiscal_year,
                resume_month=month,
                status='APPROVED',
                is_active=True
            ).first()
            
            try:
                # First check if previous month report exists and is approved
                previous_loss_report = LossReport.objects.get(
                    distribution_center=report.distribution_center,
                    fiscal_year=report.fiscal_year,
                    month=previous_month
                )
                
                if previous_loss_report.status != 'APPROVED' and not approved_override:
                    messages.error(
                        request,
                        f'Previous month report ({month_names.get(previous_month, "")}) must be approved before creating {month_names.get(month, "")} report.'
                    )
                    return redirect('report_create')
                
                # Then get monthly data and create previous readings dictionary
                try:
                    previous_month_report = MonthlyLossData.objects.get(
                        report=previous_loss_report,
                        month=previous_month
                    )
                    # Create dictionary of previous readings for each meter point
                    # Only carry forward for non-single-reading types
                    for prev_reading in previous_month_report.meter_readings.select_related('meter_point').all():
                        if not prev_reading.meter_point.is_single_reading:
                            previous_readings_dict[prev_reading.meter_point_id] = prev_reading.present_reading or 0
                    
                    # Debug: Print what we found
                                
                except MonthlyLossData.DoesNotExist:
                    # Previous month report exists but no monthly data yet
                    # This is OK - just use 0 for all meter points
                    pass
                    
            except LossReport.DoesNotExist:
                # Previous month report doesn't exist
                # Check if there's an approved override that allows this
                if approved_override:
                    # Override approved - use 0 for all previous readings
                    pass  # previous_readings_dict will remain empty, so all readings start from 0
                elif request.user.is_dc_level:  # Only show error to DC users
                    messages.error(
                        request,
                        f'Previous month report ({month_names.get(previous_month, "")}) hasn\'t been created. Please create that first.'
                    )
                    return redirect('report_create')
        else:
            # For Shrawan (month 1), previous reading is 0
            if month == 1:
                previous_month_present_reading = 0
        
        # Editable flows must have a MonthlyLossData row to persist AJAX saves.
        if can_edit:
            monthly, _ = MonthlyLossData.objects.get_or_create(
                report=report,
                month=month,
                defaults={'month_name': month_names.get(month, '')},
            )
        else:
            # View-only: do not create empty rows.
            monthly = MonthlyLossData.objects.filter(report=report, month=month).first()

        existing_readings = {r.meter_point_id: r for r in monthly.meter_readings.all()} if monthly else {}

        # IDs explicitly marked inactive for this month (deleted/disabled for this month only)
        inactive_for_month = set()
        if monthly:
            inactive_for_month = set(
                MonthlyMeterPointStatus.objects.filter(
                    monthly_data=monthly,
                    is_active=False
                ).values_list('meter_point_id', flat=True)
            )

        # Determine which meter points to show based on report status
        if report.status in ['APPROVED']:
            # Approved report: show feeders that had readings, BUT still respect
            # per-month soft-deletes — if a feeder was deleted for THIS specific month,
            # hide it from this month's view (other months are unaffected).
            # Also, only show feeders that existed at the time of this report's month
            # (feeders added in later months should not appear in earlier months' reports)
            
            # Get meter points that have readings in this month or prior months of this fiscal year
            # This ensures feeders only appear from the month they were first used onwards
            if report.fiscal_year:
                # Meter points that have readings in this month or any prior month of this fiscal year
                meter_points_with_readings = set(
                    MeterReading.objects.filter(
                        meter_point__distribution_center=report.distribution_center,
                        monthly_data__report__fiscal_year=report.fiscal_year,
                        monthly_data__month__lte=month
                    ).values_list('meter_point_id', flat=True).distinct()
                )
                
                meter_points = MeterPoint.objects.filter(
                    distribution_center=report.distribution_center,
                    is_active=True,
                    pk__in=meter_points_with_readings
                ).exclude(
                    pk__in=inactive_for_month
                ).order_by('source_type', 'name')
            else:
                # Fallback if no fiscal year: show all active meter points
                meter_points = MeterPoint.objects.filter(
                    distribution_center=report.distribution_center,
                    is_active=True
                ).exclude(
                    pk__in=inactive_for_month
                ).order_by('source_type', 'name')
        elif not monthly:
            # No monthly data yet — show all active meter points
            meter_points = MeterPoint.objects.filter(
                distribution_center=report.distribution_center,
                is_active=True
            ).order_by('source_type', 'name')
        else:
            # Draft/Rejected report: show all active feeders EXCEPT those
            # explicitly removed for this month via MonthlyMeterPointStatus
            meter_points = MeterPoint.objects.filter(
                distribution_center=report.distribution_center,
                is_active=True
            ).exclude(
                pk__in=inactive_for_month
            ).order_by('source_type', 'name')

        existing_readings = {r.meter_point_id: r for r in monthly.meter_readings.select_related('meter_point').all()} if monthly else {}

        # Identify new meter points (those created in the current month)
        # A meter point is "new for this month" ONLY if it has NO reading in any
        # prior month for this DC (approved or same-report draft).
        # If it has prior readings it MUST auto-fill previous reading from last month.
        all_meter_point_ids = set(meter_points.values_list('pk', flat=True))

        if month > 1:
            # IDs that already have at least one reading in a prior approved report
            prior_approved_ids = set(
                MeterReading.objects.filter(
                    meter_point__distribution_center=report.distribution_center,
                    monthly_data__report__status='APPROVED',
                    monthly_data__report__fiscal_year=report.fiscal_year,
                    monthly_data__month__lt=month,
                ).values_list('meter_point_id', flat=True).distinct()
            )
            # IDs that have a reading in a prior month of the same (possibly draft) report
            prior_same_report_ids = set(
                MeterReading.objects.filter(
                    monthly_data__report=report,
                    monthly_data__month__lt=month,
                ).values_list('meter_point_id', flat=True).distinct()
            )
            has_prior = prior_approved_ids | prior_same_report_ids
            # "New" = never appeared in any prior month → allow manual previous-reading entry
            new_meter_point_ids = all_meter_point_ids - has_prior
        else:
            # Shrawan (month 1): all feeders are first-time, previous reading is editable
            new_meter_point_ids = all_meter_point_ids

        import_points = meter_points.filter(
            source_type__in=['SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT']
        )
        export_points = meter_points.filter(source_type__in=['EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT'])

        # IDs of single-reading meter points (ENERGY_IMPORT / ENERGY_EXPORT)
        single_reading_ids = set(
            meter_points.filter(source_type__in=['ENERGY_IMPORT', 'ENERGY_EXPORT']).values_list('pk', flat=True)
        )

        consumer_categories = ConsumerCategory.objects.filter(is_active=True).filter(
            Q(distribution_center__isnull=True)
            | Q(distribution_center_id=report.distribution_center_id)
        ).order_by('display_order', 'name')

        # Load DC yearly target for this fiscal year (set by provincial office)
        dc_yearly_target = None
        if report.fiscal_year:
            t = DCYearlyTarget.objects.filter(
                distribution_center=report.distribution_center,
                fiscal_year=report.fiscal_year,
            ).first()
            dc_yearly_target = float(t.target_loss_percent) if t else None

        existing_utilisations = {e.consumer_category_id: e for e in monthly.energy_utilisations.all()} if monthly else {}
        existing_counts = {c.consumer_category_id: c for c in monthly.consumer_counts.all()} if monthly else {}

        import_type_choices = [
            (k, v) for k, v in MeterPoint.SOURCE_TYPE_CHOICES
            if k in ['SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT']
        ]
        export_type_choices = [
            (k, v) for k, v in MeterPoint.SOURCE_TYPE_CHOICES
            if k in ['EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT']
        ]

        months_nav = list(month_names.items())
        
        # Get distribution centers for cross-DC energy transfer dropdown
        # Only show DCs whose previous month report is approved (or if it's Shrawan, show all)
        if month > 1:
            previous_month = month - 1
            # Get DC IDs that have an approved report for the previous month in the same fiscal year
            eligible_dc_ids = set(DistributionCenter.objects.filter(
                is_active=True,
                loss_reports__fiscal_year=report.fiscal_year,
                loss_reports__month=previous_month,
                loss_reports__status='APPROVED'
            ).values_list('pk', flat=True))
            # Also include the current DC itself
            eligible_dc_ids.add(report.distribution_center.pk)
            # Query with the combined IDs
            all_dcs = DistributionCenter.objects.filter(
                pk__in=eligible_dc_ids,
                is_active=True
            ).order_by('name')
        else:
            # For Shrawan (first month), show all active DCs
            all_dcs = DistributionCenter.objects.filter(is_active=True).order_by('name')
        
        return render(request, self.template_name, {
            'report': report,
            'monthly': monthly,
            'month': month,  # Use URL parameter month instead of report.month
            'month_name': month_names.get(month, ''),
            'previous_readings_dict': previous_readings_dict,
            'months_nav': months_nav,
            'import_points': import_points,
            'export_points': export_points,
            'existing_readings': existing_readings,
            'consumer_categories': consumer_categories,
            'existing_utilisations': existing_utilisations,
            'existing_counts': existing_counts,
            'can_edit': can_edit,
            'import_type_choices': import_type_choices,
            'export_type_choices': export_type_choices,
            'new_meter_point_ids': list(new_meter_point_ids),
            'single_reading_ids': list(single_reading_ids),  # ENERGY_IMPORT / ENERGY_EXPORT
            'dc_yearly_target': dc_yearly_target,
            'approved_override': approved_override if month > 1 else None,
            'all_dcs': all_dcs,  # For cross-DC energy transfer dropdown
        })


class ReportPrintView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/print.html'

    def get(self, request, pk):
        report = get_object_or_404(LossReport, pk=pk)
        monthly_data = report.monthly_data.order_by('month').prefetch_related(
            'meter_readings__meter_point',
            'energy_utilisations__consumer_category',
            'consumer_counts__consumer_category'
        )
        return render(request, self.template_name, {
            'report': report,
            'monthly_data': monthly_data,
        })


# ─────────────────────────── FEEDER FILE UPLOADS ───────────────────────────

class FeederFileUploadView(LoginRequiredMixin, View):
    """DCS users can upload feeder files for their monthly reports"""
    template_name = 'nea_loss/reports/feeder_file_upload.html'

    def get(self, request, report_pk):
        report = get_object_or_404(LossReport, pk=report_pk)
        # Check if user can edit this report
        if not _can_edit_report(request.user, report):
            messages.error(request, 'You do not have permission to upload files for this report.')
            return redirect('report_detail', pk=report_pk)
        
        existing_files = report.feeder_files.all()
        return render(request, self.template_name, {
            'report': report,
            'existing_files': existing_files,
        })

    def post(self, request, report_pk):
        report = get_object_or_404(LossReport, pk=report_pk)
        if not _can_edit_report(request.user, report):
            messages.error(request, 'You do not have permission to upload files for this report.')
            return redirect('report_detail', pk=report_pk)
        
        feeder_name = request.POST.get('feeder_name', '').strip()
        file_type = request.POST.get('file_type', 'OTHER')
        description = request.POST.get('description', '').strip()
        uploaded_file = request.FILES.get('file')
        
        if not feeder_name or not uploaded_file:
            messages.error(request, 'Feeder name and file are required.')
            return redirect('feeder_file_upload', report_pk=report_pk)
        
        # Auto-detect file type if not specified
        if file_type == 'OTHER':
            file_ext = uploaded_file.name.split('.')[-1].lower()
            if file_ext == 'pdf':
                file_type = 'PDF'
            elif file_ext in ['doc', 'docx']:
                file_type = 'WORD'
            elif file_ext in ['xls', 'xlsx']:
                file_type = 'EXCEL'
            elif file_ext in ['jpg', 'jpeg', 'png', 'gif']:
                file_type = 'IMAGE'
        
        FeederFile.objects.create(
            report=report,
            feeder_name=feeder_name,
            file=uploaded_file,
            file_type=file_type,
            description=description,
            uploaded_by=request.user
        )
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            action='UPLOAD',
            model_name='FeederFile',
            object_id=None,
            description=f'Uploaded feeder file "{feeder_name}" for report {report}',
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        messages.success(request, f'File "{feeder_name}" uploaded successfully.')
        return redirect('feeder_file_upload', report_pk=report_pk)


class FeederFileDeleteView(LoginRequiredMixin, View):
    """Delete a feeder file"""
    
    def post(self, request, file_pk):
        file_obj = get_object_or_404(FeederFile, pk=file_pk)
        report = file_obj.report
        
        # Check if user can edit this report
        if not _can_edit_report(request.user, report):
            messages.error(request, 'You do not have permission to delete this file.')
            return redirect('report_detail', pk=report.pk)
        
        feeder_name = file_obj.feeder_name
        file_obj.delete()
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            model_name='FeederFile',
            object_id=str(file_pk),
            description=f'Deleted feeder file "{feeder_name}" from report {report}',
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        messages.success(request, f'File "{feeder_name}" deleted successfully.')
        return redirect('feeder_file_upload', report_pk=report.pk)


class FeederFileView(LoginRequiredMixin, View):
    """View a feeder file - accessible to higher authorities"""
    
    def get(self, request, pk):
        file_obj = get_object_or_404(FeederFile, pk=pk)
        
        # Check if user can view this file (DCS user who uploaded it or higher authority)
        can_view = (
            request.user == file_obj.uploaded_by or
            request.user.is_provincial or
            request.user.is_top_management or
            getattr(request.user, 'is_system_admin', False)
        )
        
        if not can_view:
            messages.error(request, 'You do not have permission to view this file.')
            return redirect('dashboard')
        
        # Serve the file
        from django.http import FileResponse
        import os
        
        if file_obj.file and os.path.exists(file_obj.file.path):
            return FileResponse(open(file_obj.file.path, 'rb'), as_attachment=True, filename=file_obj.get_filename())
        else:
            messages.error(request, 'File not found.')
            return redirect('report_detail', pk=file_obj.report.pk)


# ─────────────────────────── FEEDER FILE API ───────────────────────────

@login_required
def api_feeder_file_upload(request):
    """API endpoint for uploading feeder files via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Only POST method allowed'})
    
    report_pk = request.POST.get('report_pk')
    feeder_name = request.POST.get('feeder_name', '').strip()
    file_type = request.POST.get('file_type', 'OTHER')
    description = request.POST.get('description', '').strip()
    uploaded_file = request.FILES.get('file')
    
    if not report_pk or not feeder_name or not uploaded_file:
        return JsonResponse({'success': False, 'error': 'Missing required fields'})
    
    try:
        report = get_object_or_404(LossReport, pk=report_pk)
        
        # Check if user can edit this report
        if not _can_edit_report(request.user, report):
            return JsonResponse({'success': False, 'error': 'Permission denied'})
        
        # Auto-detect file type if not specified
        if file_type == 'OTHER':
            file_ext = uploaded_file.name.split('.')[-1].lower()
            if file_ext == 'pdf':
                file_type = 'PDF'
            elif file_ext in ['doc', 'docx']:
                file_type = 'WORD'
            elif file_ext in ['xls', 'xlsx']:
                file_type = 'EXCEL'
            elif file_ext in ['jpg', 'jpeg', 'png', 'gif']:
                file_type = 'IMAGE'
        
        file_obj = FeederFile.objects.create(
            report=report,
            feeder_name=feeder_name,
            file=uploaded_file,
            file_type=file_type,
            description=description,
            uploaded_by=request.user
        )
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            action='UPLOAD',
            model_name='FeederFile',
            object_id=None,
            description=f'Uploaded feeder file "{feeder_name}" for report {report}',
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return JsonResponse({
            'success': True,
            'file': {
                'id': file_obj.pk,
                'feeder_name': file_obj.feeder_name,
                'filename': file_obj.get_filename(),
                'file_type': file_obj.file_type
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def api_pending_reports(request):
    """API endpoint to get pending reports for provincial review"""
    if not request.user.is_provincial:
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    try:
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        if not active_fy:
            return JsonResponse({'success': False, 'error': 'No active fiscal year'})
        
        po = request.user.provincial_office
        if not po:
            return JsonResponse({'success': False, 'error': 'No provincial office assigned'})
        
        pending_reports = LossReport.objects.filter(
            distribution_center__provincial_office=po,
            fiscal_year=active_fy,
            status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED']
        ).select_related('distribution_center').order_by('-updated_at')
        
        reports_data = []
        for report in pending_reports:
            reports_data.append({
                'id': report.pk,
                'dc_name': report.distribution_center.name,
                'month': report.get_month_display(),
                'status': report.get_status_display(),
                'updated_at': report.updated_at.strftime('%Y-%m-%d %H:%M')
            })
        
        return JsonResponse({
            'success': True,
            'reports': reports_data,
            'count': len(reports_data)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ─────────────────────────── REPORT ACTIONS ───────────────────────────

@login_required
def report_submit(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    if request.method == 'POST' and _can_edit_report(request.user, report):
        report.status = 'SUBMITTED'
        report.submitted_by = request.user
        report.submission_date = timezone.now()
        dc_remarks = (request.POST.get('remarks') or '').strip()
        if dc_remarks:
            report.remarks = dc_remarks
        report.save()

        # Notify provincial manager
        if report.distribution_center.provincial_office:
            prov_managers = NEAUser.objects.filter(
                role='PROVINCIAL_MANAGER',
                provincial_office=report.distribution_center.provincial_office
            )
            for mgr in prov_managers:
                Notification.objects.create(
                    recipient=mgr,
                    notification_type='REPORT_SUBMITTED',
                    title=f'New Report: {report.distribution_center.name}',
                    message=f'Loss report for {report.fiscal_year.year_bs} has been submitted by {report.distribution_center.name}.',
                    related_report=report
                )
        # No flash message: user feedback is shown in the UI flow.
    return redirect('report_detail', pk=pk)


@login_required
def report_approve(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    if request.method == 'POST' and _can_approve_report(request.user):
        report.status = 'APPROVED'
        report.approved_by = request.user
        report.approval_date = timezone.now()
        approval_remarks = (request.POST.get('remarks') or '').strip()
        if approval_remarks:
            report.remarks = approval_remarks
        report.save()
        record_dcs_history_snapshot(report)
        messages.success(request, 'Report approved successfully.')
    return redirect('report_detail', pk=pk)


@login_required
def report_reject(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    if request.method == 'POST' and _can_approve_report(request.user):
        report.status = 'REJECTED'
        report.remarks = request.POST.get('remarks', '')
        report.save()
        messages.warning(request, 'Report has been rejected.')
    return redirect('report_detail', pk=pk)


@login_required
@require_POST
def report_delete(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    if not _can_edit_report(request.user, report):
        messages.error(request, 'You do not have permission to delete this report.')
        return redirect('report_detail', pk=pk)
    if report.status != 'DRAFT':
        messages.error(request, 'Only DRAFT reports can be deleted.')
        return redirect('report_detail', pk=pk)
    report.delete()
    messages.success(request, 'Report deleted successfully.')
    return redirect('report_list')


@login_required
@require_POST
def monthly_data_delete(request, report_pk, month):
    report = get_object_or_404(LossReport, pk=report_pk)
    if not _can_edit_report(request.user, report):
        messages.error(request, 'You do not have permission to delete this month data.')
        return redirect('report_detail', pk=report.pk)

    # Since each report is now monthly, delete the entire report
    month_names = {
        1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
        5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
        9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
    }
    month_name = month_names.get(month, '')
    
    # Delete the entire monthly report
    report.delete()
    messages.success(request, f'Deleted monthly report for {month_name}.')
    return redirect('report_list')


@login_required
@require_POST
def api_delete_meter_reading_for_month(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    monthly_id = data.get('monthly_id')
    meter_point_id = data.get('meter_point_id')
    if not monthly_id or not meter_point_id:
        return JsonResponse({'error': 'monthly_id and meter_point_id are required'}, status=400)

    monthly = get_object_or_404(MonthlyLossData, pk=monthly_id)
    if not _can_edit_report(request.user, monthly.report):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    MeterReading.objects.filter(monthly_data=monthly, meter_point_id=meter_point_id).delete()

    # Recalculate totals from remaining readings
    import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP'}
    export_types = {'EXPORT_DC', 'EXPORT_IPP'}
    total_import = decimal.Decimal('0')
    total_export = decimal.Decimal('0')
    for mr in monthly.meter_readings.select_related('meter_point').all():
        if mr.meter_point.source_type in import_types:
            total_import += mr.unit_kwh
        elif mr.meter_point.source_type in export_types:
            total_export += mr.unit_kwh

    monthly.total_energy_import = total_import
    monthly.total_energy_export = total_export
    monthly.net_energy_received = total_import - total_export
    monthly.loss_unit = monthly.net_energy_received - monthly.total_energy_utilised
    if monthly.net_energy_received > 0:
        monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
    else:
        monthly.monthly_loss_percent = 0
    monthly.save()
    monthly.report.calculate_summary()

    return JsonResponse({
        'success': True,
        'total_import': float(total_import),
        'total_export': float(total_export),
        'net_received': float(monthly.net_energy_received),
    })


@login_required
@require_POST
def api_disable_meter_point_for_month(request):
    """Disable a meter point for a specific month while preserving historical data"""
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    monthly_id = data.get('monthly_id')
    meter_point_id = data.get('meter_point_id')
    if not monthly_id or not meter_point_id:
        return JsonResponse({'error': 'monthly_id and meter_point_id are required'}, status=400)

    monthly = get_object_or_404(MonthlyLossData, pk=monthly_id)
    if not _can_edit_report(request.user, monthly.report):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    meter_point = get_object_or_404(MeterPoint, pk=meter_point_id)
    if meter_point.distribution_center_id != monthly.report.distribution_center_id:
        return JsonResponse({'error': 'Invalid meter point for this distribution center'}, status=400)

    # Create or update the monthly status to mark this meter point as inactive for this month
    status, created = MonthlyMeterPointStatus.objects.update_or_create(
        monthly_data=monthly,
        meter_point=meter_point,
        defaults={'is_active': False}
    )

    # Delete the meter reading for this month (if it exists)
    MeterReading.objects.filter(monthly_data=monthly, meter_point=meter_point).delete()

    # Recalculate totals
    import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP'}
    export_types = {'EXPORT_DC', 'EXPORT_IPP'}
    total_import = decimal.Decimal('0')
    total_export = decimal.Decimal('0')
    for mr in monthly.meter_readings.select_related('meter_point').all():
        if mr.meter_point.source_type in import_types:
            total_import += mr.unit_kwh
        elif mr.meter_point.source_type in export_types:
            total_export += mr.unit_kwh

    monthly.total_energy_import = total_import
    monthly.total_energy_export = total_export
    monthly.net_energy_received = total_import - total_export
    monthly.loss_unit = monthly.net_energy_received - monthly.total_energy_utilised
    if monthly.net_energy_received > 0:
        monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
    else:
        monthly.monthly_loss_percent = 0
    monthly.save()
    monthly.report.calculate_summary()

    return JsonResponse({
        'success': True,
        'message': f'Meter point "{meter_point.name}" disabled for {monthly.month_name}. Historical data preserved.',
        'total_import': float(total_import),
        'total_export': float(total_export),
        'net_received': float(monthly.net_energy_received),
    })


@login_required
def report_export_excel(request, pk):
    report = get_object_or_404(LossReport, pk=pk)
    wb = _generate_excel_report(report)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="LossReport_{report.distribution_center.code}_{report.fiscal_year.year_bs.replace("/","_")}.xlsx"'
    wb.save(response)
    return response


# ─────────────────────────── ORGANIZATION VIEWS ───────────────────────────

class OrgOverviewView(LoginRequiredMixin, View):
    template_name = 'nea_loss/organizations/overview.html'

    def get(self, request):
        provinces = Province.objects.prefetch_related(
            'offices__distribution_centers'
        ).all()
        return render(request, self.template_name, {'provinces': provinces})


class DCDetailView(LoginRequiredMixin, View):
    template_name = 'nea_loss/organizations/dc_detail.html'

    def get(self, request, pk):
        dc = get_object_or_404(DistributionCenter, pk=pk)
        reports = LossReport.objects.filter(distribution_center=dc).order_by('-fiscal_year__year_ad_start')
        meter_points = dc.meter_points.all()
        return render(request, self.template_name, {
            'dc': dc,
            'reports': reports,
            'meter_points': meter_points,
        })


# ─────────────────────────── ANALYTICS ───────────────────────────

class AnalyticsView(LoginRequiredMixin, View):
    template_name = 'nea_loss/analytics/overview.html'

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()

        # Base queryset — only APPROVED reports feed analytics for all roles above DC
        if getattr(user, 'is_system_admin', False):
            # Sysadmin sees approved reports across all
            all_reports = LossReport.objects.filter(
                fiscal_year=active_fy, status='APPROVED'
            ) if active_fy else LossReport.objects.none()
        elif user.is_top_management:
            all_reports = LossReport.objects.filter(
                fiscal_year=active_fy, status='APPROVED'
            ) if active_fy else LossReport.objects.none()
        elif user.is_provincial and user.provincial_office:
            all_reports = LossReport.objects.filter(
                fiscal_year=active_fy,
                status='APPROVED',
                distribution_center__provincial_office=user.provincial_office,
            ) if active_fy else LossReport.objects.none()
        elif user.is_dc_level and user.distribution_center:
            # DC users see their own submitted+ reports in analytics
            all_reports = LossReport.objects.filter(
                fiscal_year=active_fy,
                distribution_center=user.distribution_center,
                status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED', 'APPROVED'],
            ) if active_fy else LossReport.objects.none()
        else:
            all_reports = LossReport.objects.none()

        view_mode = request.GET.get('view', 'dc')  # 'dc' or 'province'

        # Loss by DC
        dc_data = all_reports.values(
            'distribution_center__name'
        ).annotate(
            total_received=Sum('total_received_kwh'),
            total_loss=Sum('total_loss_kwh')
        ).order_by('-total_loss')[:15]

        # Monthly trend
        monthly_trend = {}
        for report in all_reports:
            for md in report.monthly_data.all():
                m = md.month_name
                if m not in monthly_trend:
                    monthly_trend[m] = {'received': 0, 'loss': 0}
                monthly_trend[m]['received'] += float(md.net_energy_received)
                monthly_trend[m]['loss'] += float(md.loss_unit)

        # Province-wise aggregation
        prov_data = []
        for po in ProvincialOffice.objects.all():
            po_reports = all_reports.filter(distribution_center__provincial_office=po)
            po_received = po_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
            po_utilised = po_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
            po_loss = float(po_received) - float(po_utilised)
            po_loss_pct = (po_loss / float(po_received) * 100) if po_received else 0
            prov_data.append({
                'name': po.name,
                'loss_pct': round(po_loss_pct, 2),
                'total_received': float(po_received),
                'total_loss': po_loss,
            })

        return render(request, self.template_name, {
            'active_fy': active_fy,
            'dc_data': list(dc_data),
            'monthly_trend': monthly_trend,
            'prov_data': prov_data,
            'view_mode': view_mode,
            'total_received': all_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0,
            'total_loss': all_reports.aggregate(s=Sum('total_loss_kwh'))['s'] or 0,
            'target_pct': float(active_fy.loss_target_percent) if active_fy else 3.35,
        })


class ComparisonView(LoginRequiredMixin, View):
    template_name = 'nea_loss/analytics/comparison.html'

    def get(self, request):
        user = request.user
        fiscal_years = FiscalYear.objects.all().order_by('-year_ad_start')

        MONTH_NAMES = [
            'Shrawan','Bhadra','Ashwin','Kartik','Mangsir','Poush',
            'Magh','Falgun','Chaitra','Baisakh','Jestha','Ashadh'
        ]

        # Build datasets: one per fiscal year — monthly average loss % from APPROVED reports
        datasets = []
        for fy in fiscal_years:
            if user.is_dc_level and user.distribution_center:
                fy_reports = LossReport.objects.filter(
                    fiscal_year=fy,
                    distribution_center=user.distribution_center,
                    status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED', 'APPROVED'],
                )
            elif user.is_provincial and user.provincial_office:
                fy_reports = LossReport.objects.filter(
                    fiscal_year=fy,
                    status='APPROVED',
                    distribution_center__provincial_office=user.provincial_office,
                )
            else:
                fy_reports = LossReport.objects.filter(fiscal_year=fy, status='APPROVED')

            # Aggregate per month: sum of received and loss across all DCs in this FY
            monthly_data = {}
            for report in fy_reports:
                for md in report.monthly_data.all():
                    m = md.month
                    if m not in monthly_data:
                        monthly_data[m] = {'received': 0.0, 'loss': 0.0}
                    monthly_data[m]['received'] += float(md.net_energy_received)
                    monthly_data[m]['loss'] += float(md.loss_unit)

            # Build 12-slot array of loss %, None for months with no data
            loss_pct_by_month = []
            has_data = False
            for m_num in range(1, 13):
                if m_num in monthly_data and monthly_data[m_num]['received'] > 0:
                    pct = round(monthly_data[m_num]['loss'] / monthly_data[m_num]['received'] * 100, 4)
                    loss_pct_by_month.append(pct)
                    has_data = True
                else:
                    loss_pct_by_month.append(None)

            if has_data:
                datasets.append({
                    'fy': fy.year_bs,
                    'data': loss_pct_by_month,
                })

        return render(request, self.template_name, {
            'fiscal_years': fiscal_years,
            'datasets': datasets,
            'month_names': MONTH_NAMES,
            'has_data': bool(datasets),
        })


class ManagerialAnalyticsView(LoginRequiredMixin, View):
    """DMD/MD Managerial Analytics Dashboard with visualizations and DC comparisons"""
    template_name = 'nea_loss/analytics/managerial.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_top_management):
                messages.error(request, 'Only top management (MD/DMD/Director) can access managerial analytics.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        selected_dc_id = request.GET.get('dc')
        selected_province_id = request.GET.get('province')
        
        MONTH_NAMES = {
            1:'Shrawan',2:'Bhadra',3:'Ashwin',4:'Kartik',
            5:'Mangsir',6:'Poush',7:'Magh',8:'Falgun',
            9:'Chaitra',10:'Baisakh',11:'Jestha',12:'Ashadh'
        }

        # Get all approved reports for active fiscal year
        all_reports = LossReport.objects.filter(
            fiscal_year=active_fy, status='APPROVED'
        ).select_related('distribution_center', 'distribution_center__provincial_office') if active_fy else LossReport.objects.none()
        
        # Filter by province if selected
        if selected_province_id:
            all_reports = all_reports.filter(distribution_center__provincial_office_id=selected_province_id)

        # Overall statistics
        total_received = all_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
        total_utilised = all_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
        total_loss = float(total_received) - float(total_utilised)
        overall_loss_pct = round(total_loss / float(total_received) * 100, 2) if total_received else 0

        # DC-wise analytics - filter by province if selected
        dc_queryset = DistributionCenter.objects.filter(is_active=True)
        if selected_province_id:
            dc_queryset = dc_queryset.filter(provincial_office_id=selected_province_id)
        
        dc_analytics = []
        for dc in dc_queryset.order_by('name'):
            dc_reports = all_reports.filter(distribution_center=dc)
            if dc_reports.exists():
                dc_received = dc_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
                dc_utilised = dc_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
                dc_loss = float(dc_received) - float(dc_utilised)
                dc_loss_pct = round(dc_loss / float(dc_received) * 100, 2) if dc_received else 0
                
                # Get consumer counts for this DC
                total_consumers = 0
                for report in dc_reports:
                    for md in report.monthly_data.all():
                        for cc in md.consumer_counts.all():
                            total_consumers += cc.count
                
                dc_analytics.append({
                    'id': dc.id,
                    'name': dc.name,
                    'code': dc.code,
                    'province': dc.provincial_office.name if dc.provincial_office else 'N/A',
                    'received': float(dc_received),
                    'utilised': float(dc_utilised),
                    'loss': dc_loss,
                    'loss_pct': dc_loss_pct,
                    'consumers': total_consumers,
                })

        # Province-wise analytics for pie chart
        province_analytics = []
        for po in ProvincialOffice.objects.all():
            po_reports = all_reports.filter(distribution_center__provincial_office=po)
            if po_reports.exists():
                po_received = po_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
                po_utilised = po_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
                po_loss = float(po_received) - float(po_utilised)
                po_loss_pct = round(po_loss / float(po_received) * 100, 2) if po_received else 0
                province_analytics.append({
                    'name': po.name,
                    'received': float(po_received),
                    'loss': po_loss,
                    'loss_pct': po_loss_pct,
                })

        # Monthly trend data
        monthly_trend = {}
        for report in all_reports:
            for md in report.monthly_data.all():
                m = md.month
                if m not in monthly_trend:
                    monthly_trend[m] = {'received': 0, 'utilised': 0, 'loss': 0}
                monthly_trend[m]['received'] += float(md.net_energy_received)
                monthly_trend[m]['utilised'] += float(md.total_energy_utilised)
                monthly_trend[m]['loss'] += float(md.loss_unit)

        # Sort monthly trend by month number
        monthly_trend_sorted = []
        for m in sorted(monthly_trend.keys()):
            monthly_trend_sorted.append({
                'month': MONTH_NAMES.get(m, ''),
                'month_num': m,
                'received': monthly_trend[m]['received'],
                'utilised': monthly_trend[m]['utilised'],
                'loss': monthly_trend[m]['loss'],
                'loss_pct': round(monthly_trend[m]['loss'] / monthly_trend[m]['received'] * 100, 2) if monthly_trend[m]['received'] else 0,
            })

        # DC comparison data (if a DC is selected)
        selected_dc_data = None
        dc_comparison = []
        if selected_dc_id:
            selected_dc = DistributionCenter.objects.filter(id=selected_dc_id).first()
            if selected_dc:
                selected_dc_reports = all_reports.filter(distribution_center=selected_dc)
                if selected_dc_reports.exists():
                    # Get selected DC's data
                    sel_received = selected_dc_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
                    sel_utilised = selected_dc_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
                    sel_loss = float(sel_received) - float(sel_utilised)
                    sel_loss_pct = round(sel_loss / float(sel_received) * 100, 2) if sel_received else 0
                    
                    # Get selected DC's consumer counts
                    sel_consumers = 0
                    for report in selected_dc_reports:
                        for md in report.monthly_data.all():
                            for cc in md.consumer_counts.all():
                                sel_consumers += cc.count
                    
                    selected_dc_data = {
                        'id': selected_dc.id,
                        'name': selected_dc.name,
                        'code': selected_dc.code,
                        'province': selected_dc.provincial_office.name if selected_dc.provincial_office else 'N/A',
                        'received': float(sel_received),
                        'utilised': float(sel_utilised),
                        'loss': sel_loss,
                        'loss_pct': sel_loss_pct,
                        'consumers': sel_consumers,
                    }
                    
                    # Compare with other DCs
                    for dc_data in dc_analytics:
                        if dc_data['id'] != selected_dc.id:
                            dc_comparison.append({
                                'name': dc_data['name'],
                                'loss_pct': dc_data['loss_pct'],
                                'loss_diff': round(dc_data['loss_pct'] - sel_loss_pct, 2),
                                'received': dc_data['received'],
                                'consumers': dc_data['consumers'],
                            })
                    
                    # Sort by loss difference
                    dc_comparison.sort(key=lambda x: x['loss_diff'])

        # Top and bottom performers
        sorted_by_loss = sorted(dc_analytics, key=lambda x: x['loss_pct'])
        top_performers = sorted_by_loss[:5]
        bottom_performers = sorted_by_loss[-5:]

        # Get all provinces for filter dropdown
        all_provinces = ProvincialOffice.objects.all().order_by('name')
        
        # Add DC history data (year-over-year trends)
        dc_history = []
        if selected_dc_id:
            selected_dc = DistributionCenter.objects.filter(id=selected_dc_id).first()
            if selected_dc:
                # Get all fiscal years for this DC
                all_fiscal_years = FiscalYear.objects.all().order_by('-year_bs')
                for fy in all_fiscal_years:
                    fy_reports = LossReport.objects.filter(
                        distribution_center=selected_dc,
                        fiscal_year=fy,
                        status='APPROVED'
                    )
                    if fy_reports.exists():
                        fy_received = fy_reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
                        fy_utilised = fy_reports.aggregate(s=Sum('total_utilised_kwh'))['s'] or 0
                        fy_loss = float(fy_received) - float(fy_utilised)
                        fy_loss_pct = round(fy_loss / float(fy_received) * 100, 2) if fy_received else 0
                        
                        # Monthly breakdown for this fiscal year
                        monthly_breakdown = []
                        for report in fy_reports.order_by('month'):
                            # Get monthly data for this report
                            monthly_data = report.monthly_data.filter(month=report.month).first()
                            monthly_loss_pct = 0
                            if monthly_data and monthly_data.net_energy_received > 0:
                                monthly_loss_pct = float(monthly_data.loss_unit / monthly_data.net_energy_received * 100)
                            
                            monthly_breakdown.append({
                                'month': MONTH_NAMES.get(report.month, ''),
                                'month_num': report.month,
                                'loss_pct': monthly_loss_pct,
                                'received': float(report.total_received_kwh),
                                'loss': float(report.total_loss_kwh),
                            })
                        
                        dc_history.append({
                            'fiscal_year': fy.year_bs,
                            'received': float(fy_received),
                            'utilised': float(fy_utilised),
                            'loss': fy_loss,
                            'loss_pct': fy_loss_pct,
                            'monthly_breakdown': monthly_breakdown,
                        })
        
        return render(request, self.template_name, {
            'active_fy': active_fy,
            'total_received': float(total_received),
            'total_utilised': float(total_utilised),
            'total_loss': total_loss,
            'overall_loss_pct': overall_loss_pct,
            'dc_analytics': dc_analytics,
            'province_analytics': province_analytics,
            'monthly_trend': monthly_trend_sorted,
            'selected_dc_data': selected_dc_data,
            'dc_comparison': dc_comparison,
            'top_performers': top_performers,
            'bottom_performers': bottom_performers,
            'selected_dc_id': selected_dc_id,
            'selected_province_id': selected_province_id,
            'all_provinces': all_provinces,
            'dc_history': dc_history,
            'target_loss_pct': float(active_fy.loss_target_percent) if active_fy else 3.35,
        })


# ─────────────────────────── DC YEARLY TARGETS (PROVINCIAL) ───────────────────────────

class DCYearlyTargetView(LoginRequiredMixin, View):
    """Provincial manager sets yearly loss % targets for each DC under their office."""
    template_name = 'nea_loss/reports/dc_yearly_targets.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_provincial or user.is_top_management):
                messages.error(request, 'Only provincial managers can set DC yearly targets.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()

        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            dcs = DistributionCenter.objects.select_related('provincial_office').all().order_by('provincial_office__name', 'name')
        else:
            po = user.provincial_office
            dcs = DistributionCenter.objects.filter(provincial_office=po).order_by('name') if po else DistributionCenter.objects.none()

        # Build current targets map: {dc_id: target_loss_percent}
        existing_targets = {}
        if active_fy:
            for t in DCYearlyTarget.objects.filter(fiscal_year=active_fy, distribution_center__in=dcs):
                existing_targets[t.distribution_center_id] = float(t.target_loss_percent)

        dc_rows = []
        for dc in dcs:
            # Determine if this DC's target can be edited
            target_exists = dc.pk in existing_targets
            
            # Provincial users can only edit if target doesn't exist yet
            # Admin users can always edit
            can_edit = False
            if getattr(user, 'is_system_admin', False) or user.is_top_management:
                can_edit = True  # Admin can always edit
            elif user.is_provincial:
                can_edit = not target_exists  # Provincial can edit only if no target exists yet
            
            dc_rows.append({
                'dc': dc,
                'target': existing_targets.get(dc.pk, ''),
                'can_edit': can_edit,
                'target_exists': target_exists,
            })

        return render(request, self.template_name, {
            'active_fy': active_fy,
            'dc_rows': dc_rows,
        })

    def post(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        if not active_fy:
            messages.error(request, 'No active fiscal year found.')
            return redirect('dc_yearly_targets')

        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            allowed_dcs = set(DistributionCenter.objects.values_list('pk', flat=True))
        else:
            po = user.provincial_office
            allowed_dcs = set(DistributionCenter.objects.filter(provincial_office=po).values_list('pk', flat=True)) if po else set()

        saved = 0
        for key, val in request.POST.items():
            # key format: target_<dc_id>
            if not key.startswith('target_'):
                continue
            parts = key.split('_')
            if len(parts) != 2:
                continue
            try:
                dc_id = int(parts[1])
                val = val.strip()
                if not val:
                    # Delete existing target if blank submitted (admin only)
                    if getattr(user, 'is_system_admin', False) or user.is_top_management:
                        DCYearlyTarget.objects.filter(
                            distribution_center_id=dc_id, fiscal_year=active_fy
                        ).delete()
                    continue
                target_pct = float(val)
            except (ValueError, TypeError):
                continue

            if dc_id not in allowed_dcs:
                continue

            # Check if target already exists
            existing_target = DCYearlyTarget.objects.filter(
                distribution_center_id=dc_id, fiscal_year=active_fy
            ).first()
            
            # Provincial users cannot edit existing targets
            if existing_target and user.is_provincial:
                continue  # Skip this DC, provincial user cannot edit existing target
            
            # Admin users can always edit, provincial users can only create new targets
            DCYearlyTarget.objects.update_or_create(
                distribution_center_id=dc_id,
                fiscal_year=active_fy,
                defaults={'target_loss_percent': target_pct, 'set_by': user},
            )
            saved += 1

        messages.success(request, f'Saved {saved} yearly target(s) successfully.')
        return redirect('dc_yearly_targets')


# ─────────────────────────── DC MONTHLY TARGETS (DEPRECATED) ───────────────────────────

class DCMonthlyTargetView(LoginRequiredMixin, View):
    """Provincial manager sets monthly loss % targets for each DC under their office."""
    template_name = 'nea_loss/reports/dc_monthly_targets.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_provincial or user.is_top_management):
                messages.error(request, 'Only provincial managers can set DC monthly targets.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()

        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            dcs = DistributionCenter.objects.select_related('provincial_office').all().order_by('provincial_office__name', 'name')
        else:
            po = user.provincial_office
            dcs = DistributionCenter.objects.filter(provincial_office=po).order_by('name') if po else DistributionCenter.objects.none()

        MONTHS = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]

        # Build current targets map: {(dc_id, month): target_loss_percent}
        existing = {}
        if active_fy:
            for t in DCMonthlyTarget.objects.filter(fiscal_year=active_fy, distribution_center__in=dcs):
                existing[(t.distribution_center_id, t.month)] = float(t.target_loss_percent)

        dc_rows = []
        for dc in dcs:
            month_targets = []
            for m_num, m_name in MONTHS:
                month_targets.append({
                    'month': m_num,
                    'month_name': m_name,
                    'target': existing.get((dc.pk, m_num), ''),
                })
            dc_rows.append({'dc': dc, 'month_targets': month_targets})

        return render(request, self.template_name, {
            'active_fy': active_fy,
            'dc_rows': dc_rows,
            'months': MONTHS,
        })

    def post(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        if not active_fy:
            messages.error(request, 'No active fiscal year found.')
            return redirect('dc_monthly_targets')

        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            allowed_dcs = set(DistributionCenter.objects.values_list('pk', flat=True))
        else:
            po = user.provincial_office
            allowed_dcs = set(DistributionCenter.objects.filter(provincial_office=po).values_list('pk', flat=True)) if po else set()

        saved = 0
        for key, val in request.POST.items():
            # key format: target_<dc_id>_<month>
            if not key.startswith('target_'):
                continue
            parts = key.split('_')
            if len(parts) != 3:
                continue
            try:
                dc_id = int(parts[1])
                month = int(parts[2])
                val = val.strip()
                if not val:
                    # Delete existing target if blank submitted
                    DCMonthlyTarget.objects.filter(
                        distribution_center_id=dc_id, fiscal_year=active_fy, month=month
                    ).delete()
                    continue
                target_pct = float(val)
            except (ValueError, TypeError):
                continue

            if dc_id not in allowed_dcs:
                continue
            if not (1 <= month <= 12):
                continue

            DCMonthlyTarget.objects.update_or_create(
                distribution_center_id=dc_id,
                fiscal_year=active_fy,
                month=month,
                defaults={'target_loss_percent': target_pct, 'set_by': user},
            )
            saved += 1

        messages.success(request, f'Saved {saved} monthly target(s) successfully.')
        return redirect('dc_monthly_targets')


# ─────────────────────────── USER MANAGEMENT ───────────────────────────

class UserListView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/list.html'

    def get(self, request):
        if not request.user.is_system_admin and not request.user.is_top_management and not request.user.is_staff:
            return redirect('dashboard')
        users = NEAUser.objects.select_related('provincial_office', 'distribution_center').all()
        return render(request, self.template_name, {'users': users})


class UserCreateView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/create.html'

    def get(self, request):
        if not request.user.is_system_admin and not request.user.is_top_management and not request.user.is_staff:
            return redirect('dashboard')
        return render(request, self.template_name, {
            'role_choices': NEAUser.ROLE_CHOICES,
            'provincial_offices': ProvincialOffice.objects.all(),
            'distribution_centers': DistributionCenter.objects.all(),
        })

    def post(self, request):
        username = request.POST.get('username')
        email = request.POST.get('email')
        full_name = request.POST.get('full_name')
        role = request.POST.get('role')
        password = request.POST.get('password')
        po_id = request.POST.get('provincial_office') or None
        dc_id = request.POST.get('distribution_center') or None

        if NEAUser.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return redirect('user_create')

        user = NEAUser.objects.create_user(
            username=username, email=email, full_name=full_name,
            role=role, password=password,
            provincial_office_id=po_id, distribution_center_id=dc_id,
            employee_id=request.POST.get('employee_id', ''),
            phone=request.POST.get('phone', ''),
            designation=request.POST.get('designation', ''),
        )
        messages.success(request, f'User {username} created successfully.')
        return redirect('user_list')


class UserEditView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/edit.html'

    def get(self, request, pk):
        if not request.user.is_system_admin and not request.user.is_top_management and not request.user.is_staff:
            return redirect('dashboard')
        user = get_object_or_404(NEAUser, pk=pk)
        return render(request, self.template_name, {
            'edit_user': user,
            'role_choices': NEAUser.ROLE_CHOICES,
            'provincial_offices': ProvincialOffice.objects.all(),
            'distribution_centers': DistributionCenter.objects.all(),
        })

    def post(self, request, pk):
        user = get_object_or_404(NEAUser, pk=pk)
        user.full_name = request.POST.get('full_name', user.full_name)
        user.role = request.POST.get('role', user.role)
        user.email = request.POST.get('email', user.email)
        user.phone = request.POST.get('phone', user.phone)
        user.designation = request.POST.get('designation', user.designation)
        po_id = request.POST.get('provincial_office') or None
        dc_id = request.POST.get('distribution_center') or None
        user.provincial_office_id = po_id
        user.distribution_center_id = dc_id
        user.is_active = request.POST.get('is_active') == 'on'
        user.save()
        messages.success(request, 'User updated successfully.')
        return redirect('user_list')


# ─────────────────────────── API VIEWS ───────────────────────────

@login_required
def api_dashboard_chart(request):
    active_fy = FiscalYear.objects.filter(is_active=True).first()
    if not active_fy:
        return JsonResponse({'months': [], 'received': [], 'utilised': [], 'loss': []})

    user = request.user
    # Only APPROVED reports feed the dashboard chart for management views
    if user.is_dc_level and user.distribution_center:
        reports = LossReport.objects.filter(
            fiscal_year=active_fy,
            distribution_center=user.distribution_center,
            status__in=['SUBMITTED', 'PROVINCIAL_REVIEWED', 'APPROVED'],
        )
    elif user.is_provincial and user.provincial_office:
        reports = LossReport.objects.filter(
            fiscal_year=active_fy,
            status='APPROVED',
            distribution_center__provincial_office=user.provincial_office,
        )
    else:
        reports = LossReport.objects.filter(fiscal_year=active_fy, status='APPROVED')

    data = {}
    for report in reports:
        for md in report.monthly_data.order_by('month'):
            m = md.month_name
            if m not in data:
                data[m] = {'received': 0, 'utilised': 0, 'loss': 0}
            data[m]['received'] += float(md.net_energy_received)
            data[m]['utilised'] += float(md.total_energy_utilised)
            data[m]['loss'] += float(md.loss_unit)

    if not data:
        return JsonResponse({'months': [], 'received': [], 'utilised': [], 'loss': []})

    months = list(data.keys())
    return JsonResponse({
        'months': months,
        'received': [data[m]['received'] for m in months],
        'utilised': [data[m]['utilised'] for m in months],
        'loss': [data[m]['loss'] for m in months],
    })


@login_required
def api_loss_summary(request):
    active_fy = FiscalYear.objects.filter(is_active=True).first()
    reports = LossReport.objects.filter(fiscal_year=active_fy) if active_fy else LossReport.objects.none()
    total_received = reports.aggregate(s=Sum('total_received_kwh'))['s'] or 0
    total_loss = reports.aggregate(s=Sum('total_loss_kwh'))['s'] or 0
    loss_pct = float(total_loss / total_received * 100) if total_received > 0 else 0
    return JsonResponse({
        'total_received_mwh': round(float(total_received) / 1000, 2),
        'total_loss_mwh': round(float(total_loss) / 1000, 2),
        'loss_pct': round(loss_pct, 2),
        'target_pct': float(active_fy.loss_target_percent) if active_fy else 3.35,
    })


@login_required
def api_mark_notifications_read(request):
    if request.method == 'POST':
        Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
        return JsonResponse({'success': True})
    return JsonResponse({'error': 'POST required'})


@login_required
def api_create_monthly_data(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'})
    try:
        data = json.loads(request.body)
        report_id = data.get('report_id')
        month = data.get('month')
        
        if not report_id or not month:
            return JsonResponse({'error': 'report_id and month are required'}, status=400)
            
        report = get_object_or_404(LossReport, pk=report_id)
        
        # Check if user can edit this report
        if not _can_edit_report(request.user, report):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }

        # Create monthly data
        monthly, created = MonthlyLossData.objects.get_or_create(
            report=report,
            month=month,
            defaults={
                'month_name': month_names.get(int(month), ''),
            }
        )

        return JsonResponse({
            'success': True,
            'monthly_id': monthly.pk,
            'created': created
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_save_meter_readings(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'})
    try:
        data = json.loads(request.body)
        monthly_id = data.get('monthly_id')
        readings = data.get('readings', [])
        
        if not monthly_id:
            return JsonResponse({'error': 'monthly_id is required'}, status=400)
            
        # Get the monthly data, return error if doesn't exist
        try:
            monthly = MonthlyLossData.objects.get(pk=monthly_id)
        except MonthlyLossData.DoesNotExist:
            return JsonResponse({'error': 'Monthly data not found. Please enter some data first.'}, status=400)
            
        if not _can_edit_report(request.user, monthly.report):
            return JsonResponse({'error': 'Permission denied'}, status=403)

        import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT'}
        export_types = {'EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT'}

        # Cross-DC validation for energy import/export
        for r in readings:
            mp = get_object_or_404(MeterPoint, pk=r['meter_point_id'])
            if mp.distribution_center_id != monthly.report.distribution_center_id:
                return JsonResponse({'error': 'Invalid meter point'}, status=400)

            # Validate cross-DC energy transfer
            if mp.source_type == 'ENERGY_IMPORT' and (mp.linked_distribution_center or mp.linked_distribution_center_name):
                # Determine linked DC - use foreign key if available, otherwise search by name
                linked_dc = mp.linked_distribution_center
                linked_dc_name = mp.linked_distribution_center_name
                
                if not linked_dc and linked_dc_name:
                    # Try to find DC by name (case-insensitive)
                    linked_dc = DistributionCenter.objects.filter(name__iexact=linked_dc_name).first()
                
                if linked_dc:
                    # Check if the linked DC has a corresponding ENERGY_EXPORT with the same name
                    linked_mp = MeterPoint.objects.filter(
                        distribution_center=linked_dc,
                        name=mp.name,
                        source_type__in=['ENERGY_EXPORT', 'EXPORT_DC'],
                        is_active=True
                    ).first()
                    
                    if not linked_mp:
                        return JsonResponse({
                            'error': f'Cross-DC validation failed: {linked_dc.name} must have an ENERGY_EXPORT/EXPORT_DC meter point named "{mp.name}" to match this ENERGY_IMPORT.'
                        }, status=400)
                    
                    # Check if the linked DC has entered a reading for the same month
                    linked_monthly = MonthlyLossData.objects.filter(
                        report__distribution_center=linked_dc,
                        report__fiscal_year=monthly.report.fiscal_year,
                        month=monthly.month
                    ).first()
                    
                    if linked_monthly:
                        linked_reading = MeterReading.objects.filter(
                            monthly_data=linked_monthly,
                            meter_point=linked_mp
                        ).first()
                        
                        if linked_reading:
                            current_reading_value = decimal.Decimal(str(r['present_reading'])) * decimal.Decimal(str(r.get('multiplying_factor', mp.multiplying_factor)))
                            linked_reading_value = linked_reading.unit_kwh
                            
                            # Allow small tolerance for floating point differences
                            if abs(current_reading_value - linked_reading_value) > decimal.Decimal('0.01'):
                                return JsonResponse({
                                    'error': f'Cross-DC validation failed: Reading for "{mp.name}" does not match. {mp.distribution_center.name} entered {current_reading_value} kWh, but {linked_dc.name} entered {linked_reading_value} kWh. Both must be the same.'
                                }, status=400)
            
            elif mp.source_type in ['ENERGY_EXPORT', 'EXPORT_DC'] and (mp.linked_distribution_center or mp.linked_distribution_center_name):
                # Determine linked DC - use foreign key if available, otherwise search by name
                linked_dc = mp.linked_distribution_center
                linked_dc_name = mp.linked_distribution_center_name
                
                if not linked_dc and linked_dc_name:
                    # Try to find DC by name (case-insensitive)
                    linked_dc = DistributionCenter.objects.filter(name__iexact=linked_dc_name).first()
                
                if linked_dc:
                    # Check if the linked DC has a corresponding ENERGY_IMPORT with the same name
                    linked_mp = MeterPoint.objects.filter(
                        distribution_center=linked_dc,
                        name=mp.name,
                        source_type='ENERGY_IMPORT',
                        is_active=True
                    ).first()
                    
                    if not linked_mp:
                        return JsonResponse({
                            'error': f'Cross-DC validation failed: {linked_dc.name} must have an ENERGY_IMPORT meter point named "{mp.name}" to match this ENERGY_EXPORT.'
                        }, status=400)
                    
                    # Check if the linked DC has entered a reading for the same month
                    linked_monthly = MonthlyLossData.objects.filter(
                        report__distribution_center=linked_dc,
                        report__fiscal_year=monthly.report.fiscal_year,
                        month=monthly.month
                    ).first()
                    
                    if linked_monthly:
                        linked_reading = MeterReading.objects.filter(
                            monthly_data=linked_monthly,
                            meter_point=linked_mp
                        ).first()
                        
                        if linked_reading:
                            current_reading_value = decimal.Decimal(str(r['present_reading'])) * decimal.Decimal(str(r.get('multiplying_factor', mp.multiplying_factor)))
                            linked_reading_value = linked_reading.unit_kwh
                            
                            # Allow small tolerance for floating point differences
                            if abs(current_reading_value - linked_reading_value) > decimal.Decimal('0.01'):
                                return JsonResponse({
                                    'error': f'Cross-DC validation failed: Reading for "{mp.name}" does not match. {mp.distribution_center.name} entered {current_reading_value} kWh, but {linked_dc.name} entered {linked_reading_value} kWh. Both must be the same.'
                                }, status=400)

            if mp.is_single_reading:
                # ENERGY_IMPORT / ENERGY_EXPORT: only present_reading matters; previous=0
                mr, created = MeterReading.objects.update_or_create(
                    monthly_data=monthly, meter_point=mp,
                    defaults={
                        'present_reading': decimal.Decimal(str(r['present_reading'])),
                        'previous_reading': decimal.Decimal('0'),
                        'multiplying_factor': decimal.Decimal(str(r.get('multiplying_factor', mp.multiplying_factor))),
                    }
                )
                if not created:
                    # update_or_create uses queryset.update() on existing rows,
                    # bypassing the custom save() that recalculates unit_kwh.
                    # Call save() explicitly so unit_kwh is always recalculated.
                    mr.save()
            else:
                # For regular feeders: if previous_reading is 0 / not provided and this is
                # not Shrawan, auto-fill from the last approved month's present reading.
                provided_prev = decimal.Decimal(str(r.get('previous_reading', 0) or 0))
                report_month = monthly.report.month

                # Check if there's an approved and active override for this month
                approved_override = DCReportOverride.objects.filter(
                    distribution_center=monthly.report.distribution_center,
                    fiscal_year=monthly.report.fiscal_year,
                    resume_month=report_month,
                    status='APPROVED',
                    is_active=True
                ).first()
                
                if provided_prev == 0 and report_month > 1 and not approved_override:
                    # Look up last month's approved present reading for this meter point
                    # Only auto-fill if there's NO override
                    prev_month_num = report_month - 1
                    auto_prev = MeterReading.objects.filter(
                        meter_point=mp,
                        monthly_data__report__distribution_center=monthly.report.distribution_center,
                        monthly_data__report__fiscal_year=monthly.report.fiscal_year,
                        monthly_data__report__status='APPROVED',
                        monthly_data__month=prev_month_num,
                    ).values_list('present_reading', flat=True).first()
                    if auto_prev is not None:
                        provided_prev = auto_prev

                mr, created = MeterReading.objects.update_or_create(
                    monthly_data=monthly, meter_point=mp,
                    defaults={
                        'present_reading': decimal.Decimal(str(r['present_reading'])),
                        'previous_reading': provided_prev,
                        'multiplying_factor': decimal.Decimal(str(r.get('multiplying_factor', mp.multiplying_factor))),
                    }
                )
                if not created:
                    # update_or_create uses queryset.update() on existing rows,
                    # bypassing the custom save() that recalculates unit_kwh.
                    # Call save() explicitly so unit_kwh is always recalculated.
                    mr.save()
        total_import = decimal.Decimal('0')
        total_export = decimal.Decimal('0')
        for mr in monthly.meter_readings.select_related('meter_point').all():
            if mr.meter_point.source_type in import_types:
                total_import += mr.unit_kwh
            elif mr.meter_point.source_type in export_types:
                total_export += mr.unit_kwh

        monthly.total_energy_import = total_import
        monthly.total_energy_export = total_export
        monthly.net_energy_received = total_import - total_export
        monthly.loss_unit = monthly.net_energy_received - monthly.total_energy_utilised
        if monthly.net_energy_received > 0:
            monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
        else:
            monthly.monthly_loss_percent = decimal.Decimal('0')
        monthly.save()
        monthly.report.calculate_summary()

        return JsonResponse({
            'success': True,
            'total_import': float(total_import),
            'total_export': float(total_export),
            'net_received': float(monthly.net_energy_received),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def api_save_consumer_data(request):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'})
    try:
        data = json.loads(request.body)
        monthly_id = data.get('monthly_id')
        utilisations = data.get('utilisations', [])
        counts = data.get('counts', [])
        
        if not monthly_id:
            return JsonResponse({'error': 'monthly_id is required'}, status=400)
            
        # Get the monthly data, return error if doesn't exist
        try:
            monthly = MonthlyLossData.objects.get(pk=monthly_id)
        except MonthlyLossData.DoesNotExist:
            return JsonResponse({'error': 'Monthly data not found. Please enter some data first.'}, status=400)
            
        if not _can_edit_report(request.user, monthly.report):
            return JsonResponse({'error': 'Permission denied'}, status=403)
        dc_id = monthly.report.distribution_center_id

        for u in utilisations:
            cat = get_object_or_404(ConsumerCategory, pk=u['category_id'])
            if cat.distribution_center_id not in (None, dc_id):
                return JsonResponse({'error': 'Invalid consumer category'}, status=400)
            eu, _ = EnergyUtilisation.objects.update_or_create(
                monthly_data=monthly, consumer_category=cat,
                defaults={
                    'energy_kwh': decimal.Decimal(str(u['energy_kwh'])),
                    'remarks': (u.get('remarks') or '')[:200],
                }
            )

        for c in counts:
            cat = get_object_or_404(ConsumerCategory, pk=c['category_id'])
            if cat.distribution_center_id not in (None, dc_id):
                return JsonResponse({'error': 'Invalid consumer category'}, status=400)
            ConsumerCount.objects.update_or_create(
                monthly_data=monthly, consumer_category=cat,
                defaults={
                    'count': int(c['count']),
                    'remarks': (c.get('remarks') or '')[:200],
                }
            )

        total_utilised = monthly.energy_utilisations.aggregate(s=Sum('energy_kwh'))['s'] or decimal.Decimal('0')
        monthly.total_energy_utilised = total_utilised
        monthly.loss_unit = monthly.net_energy_received - total_utilised
        if monthly.net_energy_received > 0:
            monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
        monthly.save()
        monthly.report.calculate_summary()

        return JsonResponse({
            'success': True,
            'total_utilised': float(total_utilised),
            'loss_unit': float(monthly.loss_unit),
            'loss_pct': float(monthly.monthly_loss_percent) * 100,
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_POST
def api_manage_meter_point(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    report = get_object_or_404(LossReport, pk=data.get('report_pk'))
    
    action = data.get('action')
    source_type = data.get('source_type')
    
    # Only provincial users (or system admin) can manage feeders
    # DC users can manage ENERGY_IMPORT and ENERGY_EXPORT for their own DC
    if not _can_manage_feeders(request.user, report.distribution_center, source_type):
        if source_type in ['ENERGY_IMPORT', 'ENERGY_EXPORT']:
            return JsonResponse({'error': 'Permission denied. DC users can only add ENERGY_IMPORT and ENERGY_EXPORT types. Other types require provincial approval.'}, status=403)
        else:
            return JsonResponse({'error': 'Permission denied. Only provincial users can manage this feeder type.'}, status=403)

    if action == 'create':
        name = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'error': 'Name is required'}, status=400)
        valid_types = {k for k, _ in MeterPoint.SOURCE_TYPE_CHOICES}
        if source_type not in valid_types:
            return JsonResponse({'error': 'Invalid source type'}, status=400)
        
        # Define month names for this function
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }
        
        # Handle linked DC for cross-DC energy transfer
        linked_dc_name = data.get('linked_distribution_center_name', '').strip()
        linked_dc = None
        if linked_dc_name:
            # Try to find exact match by name
            linked_dc = DistributionCenter.objects.filter(name__iexact=linked_dc_name).first()
        
        mp = MeterPoint.objects.create(
            distribution_center=report.distribution_center,
            name=name,
            code='',
            source_type=source_type,
            voltage_level=(data.get('voltage_level') or '').strip()[:20],
            multiplying_factor=decimal.Decimal(str(data.get('multiplying_factor') or 1)),
            is_active=True,
            connection_source=(data.get('connection_source') or '').strip()[:200],
            linked_distribution_center=linked_dc,
            linked_distribution_center_name=linked_dc_name,
        )
        
        # Auto-create corresponding meter point in linked DC if requested
        auto_create_linked = data.get('auto_create_linked', False)
        if auto_create_linked and linked_dc and source_type in ['ENERGY_IMPORT', 'ENERGY_EXPORT']:
            # Validate that the linked DC has a report for the same fiscal year and month
            linked_report = LossReport.objects.filter(
                distribution_center=linked_dc,
                fiscal_year=report.fiscal_year,
                month=report.month
            ).first()
            
            if not linked_report:
                month_name = month_names.get(report.month, '')
                return JsonResponse({
                    'error': f'Cannot create linked record. {linked_dc.name} does not have a report for {month_name} month in fiscal year {report.fiscal_year.year_bs}. Please ensure the linked DC has created a report for the same month.'
                }, status=400)
            
            # Determine the opposite type
            opposite_type = 'ENERGY_EXPORT' if source_type == 'ENERGY_IMPORT' else 'ENERGY_IMPORT'
            
            # Check if the opposite meter point already exists
            existing_opposite = MeterPoint.objects.filter(
                distribution_center=linked_dc,
                name=name,
                source_type=opposite_type,
                is_active=True
            ).first()
            
            if not existing_opposite:
                # Create the opposite meter point in the linked DC
                opposite_mp = MeterPoint.objects.create(
                    distribution_center=linked_dc,
                    name=name,
                    code='',
                    source_type=opposite_type,
                    voltage_level=(data.get('voltage_level') or '').strip()[:20],
                    multiplying_factor=decimal.Decimal(str(data.get('multiplying_factor') or 1)),
                    is_active=True,
                    connection_source=(data.get('connection_source') or '').strip()[:200],
                    linked_distribution_center=report.distribution_center,
                    linked_distribution_center_name=report.distribution_center.name,
                )
                
                # Mark the opposite meter point as inactive for previous months in the linked DC
                for month_num in range(1, report.month):
                    linked_monthly_data, _ = MonthlyLossData.objects.get_or_create(
                        report=linked_report,
                        month=month_num,
                        defaults={'month_name': month_names.get(month_num, '')}
                    )
                    MonthlyMeterPointStatus.objects.update_or_create(
                        monthly_data=linked_monthly_data,
                        meter_point=opposite_mp,
                        defaults={'is_active': False}
                    )
        
        # Automatically mark this meter point as inactive for previous months
        # to prevent it from appearing in historical reports
        current_month = report.month  # Current month being viewed
        for month_num in range(1, current_month):
            # Find or create monthly data for this month
            monthly_data, _ = MonthlyLossData.objects.get_or_create(
                report=report,
                month=month_num,
                defaults={'month_name': month_names.get(month_num, '')}
            )
            # Mark as inactive for previous months
            MonthlyMeterPointStatus.objects.update_or_create(
                monthly_data=monthly_data,
                meter_point=mp,
                defaults={'is_active': False}
            )
        
        return JsonResponse({
            'success': True,
            'meter_point': {
                'id': mp.pk,
                'name': mp.name,
                'source_type': mp.source_type,
                'source_type_display': mp.get_source_type_display(),
                'voltage_level': mp.voltage_level,
                'multiplying_factor': float(mp.multiplying_factor),
            },
        })

    if action == 'delete':
        mp = get_object_or_404(MeterPoint, pk=data.get('meter_point_id'))
        if mp.distribution_center_id != report.distribution_center_id:
            return JsonResponse({'error': 'Invalid meter point'}, status=400)
        
        # Check delete permissions - DC users can delete ENERGY_IMPORT and ENERGY_EXPORT for their own DC
        if not _can_manage_feeders(request.user, report.distribution_center, mp.source_type):
            if mp.source_type in ['ENERGY_IMPORT', 'ENERGY_EXPORT']:
                return JsonResponse({'error': 'Permission denied. DC users can only delete ENERGY_IMPORT and ENERGY_EXPORT types. Other types require provincial approval.'}, status=403)
            else:
                return JsonResponse({'error': 'Permission denied. Only provincial users can delete this feeder type.'}, status=403)

        monthly_id = data.get('monthly_id')
        if not monthly_id:
            return JsonResponse({'error': 'monthly_id is required for delete'}, status=400)

        monthly = get_object_or_404(MonthlyLossData, pk=monthly_id)

        # ALWAYS soft-delete: mark this feeder inactive for THIS month only.
        # The MeterPoint record is NEVER deleted — this preserves all data in
        # every previously approved report that references this feeder.
        # Only the current month's reading is removed.
        MonthlyMeterPointStatus.objects.update_or_create(
            monthly_data=monthly,
            meter_point=mp,
            defaults={'is_active': False},
        )
        # Remove the reading for this month only (approved months are untouched)
        MeterReading.objects.filter(monthly_data=monthly, meter_point=mp).delete()

        message = (
            f'"{mp.name}" removed from {monthly.month_name} only. '
            f'All other months (including approved reports) are unaffected.'
        )

        # Recalculate totals for current month
        import_types = {'SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT'}
        export_types = {'EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT'}
        total_import = decimal.Decimal('0')
        total_export = decimal.Decimal('0')
        for mr in monthly.meter_readings.select_related('meter_point').all():
            if mr.meter_point.source_type in import_types:
                total_import += mr.unit_kwh
            elif mr.meter_point.source_type in export_types:
                total_export += mr.unit_kwh
        monthly.total_energy_import = total_import
        monthly.total_energy_export = total_export
        monthly.net_energy_received = total_import - total_export
        monthly.loss_unit = monthly.net_energy_received - monthly.total_energy_utilised
        if monthly.net_energy_received > 0:
            monthly.monthly_loss_percent = monthly.loss_unit / monthly.net_energy_received
        else:
            monthly.monthly_loss_percent = decimal.Decimal('0')
        monthly.save()
        monthly.report.calculate_summary()

        return JsonResponse({
            'success': True,
            'action_taken': 'disabled_for_month',
            'message': message,
        })

    return JsonResponse({'error': 'Unknown action'}, status=400)


@login_required
@require_POST
def api_manage_consumer_category(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    report = get_object_or_404(LossReport, pk=data.get('report_pk'))
    if not _can_edit_report(request.user, report):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    action = data.get('action')
    if action == 'create':
        name = (data.get('name') or '').strip()
        if not name:
            return JsonResponse({'error': 'Name is required'}, status=400)
        dc = report.distribution_center
        for _ in range(5):
            code = f'DC{dc.pk}_{uuid.uuid4().hex[:12]}'.upper()
            code = code[:40]
            if not ConsumerCategory.objects.filter(code=code).exists():
                break
        else:
            return JsonResponse({'error': 'Could not allocate category code'}, status=500)
        cat = ConsumerCategory.objects.create(
            name=name,
            code=code,
            distribution_center=dc,
            display_order=500,
        )
        return JsonResponse({
            'success': True,
            'category': {'id': cat.pk, 'name': cat.name},
        })

    if action == 'delete':
        cat = get_object_or_404(ConsumerCategory, pk=data.get('category_id'))
        if cat.distribution_center_id != report.distribution_center_id:
            return JsonResponse({
                'error': 'Only categories added by your DC can be removed here.',
            }, status=400)
        cat.delete()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Unknown action'}, status=400)


@login_required
def api_recalculate(request, report_pk):
    report = get_object_or_404(LossReport, pk=report_pk)
    if not _can_edit_report(request.user, report):
        return JsonResponse({"error": "Permission denied"}, status=403)

    # Recalculate this report
    report.calculate_summary()

    # CASCADE: if this is an APPROVED report, recalculate all subsequent months
    # in the same DC + fiscal year so cumulative figures stay consistent.
    if report.status == "APPROVED":
        subsequent_reports = LossReport.objects.filter(
            distribution_center=report.distribution_center,
            fiscal_year=report.fiscal_year,
            month__gt=report.month,
            status="APPROVED",
        ).order_by("month")
        for subsequent in subsequent_reports:
            subsequent.calculate_summary()

    return JsonResponse({
        "success": True,
        "cumulative_loss_pct": float(report.cumulative_loss_percent) * 100,
        "total_received": float(report.total_received_kwh),
        "total_loss": float(report.total_loss_kwh),
    })


@login_required
def api_dc_feeders(request):
    """API endpoint to get feeders for a selected DC"""
    dc_id = request.GET.get('dc_id')
    report_pk = request.GET.get('report_pk')
    if not dc_id:
        return JsonResponse({'error': 'dc_id parameter required'}, status=400)
    
    try:
        dc = DistributionCenter.objects.get(pk=dc_id, is_active=True)
    except DistributionCenter.DoesNotExist:
        return JsonResponse({'error': 'Distribution Center not found'}, status=404)
    
    # If report_pk is provided, validate that the DC has a report for the same month
    if report_pk:
        try:
            report = LossReport.objects.get(pk=report_pk)
            linked_report = LossReport.objects.filter(
                distribution_center=dc,
                fiscal_year=report.fiscal_year,
                month=report.month
            ).first()
            
            if not linked_report:
                month_name = {
                    1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
                    5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
                    9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
                }.get(report.month, '')
                return JsonResponse({
                    'success': False,
                    'error': f'{dc.name} does not have a report for {month_name} month in fiscal year {report.fiscal_year.year_bs}. Please ensure the linked DC has created a report for the same month.'
                }, status=400)
        except LossReport.DoesNotExist:
            return JsonResponse({'error': 'Report not found'}, status=404)
    
    # Get all active meter points (feeders) for this DC
    feeders = MeterPoint.objects.filter(
        distribution_center=dc,
        is_active=True
    ).order_by('source_type', 'name').values('pk', 'name', 'source_type', 'connection_source')
    
    return JsonResponse({
        'success': True,
        'feeders': list(feeders),
        'dc_name': dc.name
    })


# ─────────────────────────── HELPER FUNCTIONS ───────────────────────────

def _can_create_loss_report(user):
    """Only DC staff/managers create DC loss reports.
    MD/DMD/Director/Provincial are view-only for DC reports.
    System Admin can do everything."""
    if not user.is_authenticated:
        return False
    if getattr(user, 'is_system_admin', False):
        return True
    if user.is_dc_level:
        return bool(getattr(user, 'distribution_center', None))
    # Provincial, MD, DMD, Director cannot create DC-level loss reports
    return False

def _can_admin_override_create_report(user):
    """Admin users can create reports for any DC regardless of existing reports."""
    if not user.is_authenticated:
        return False
    if getattr(user, 'is_system_admin', False):
        return True
    return False


def _can_edit_report(user, report):
    if getattr(user, 'is_system_admin', False):
        return True
    # Top management (MD/DMD/Director) can edit APPROVED reports to make corrections
    if user.is_top_management:
        return report.status == 'APPROVED'
    if user.is_provincial:
        return False
    if report.status not in ['DRAFT', 'REJECTED']:
        return False
    if user.is_dc_level:
        dc = getattr(user, 'distribution_center', None)
        if dc and dc.pk == report.distribution_center_id:
            return True
    return False


def _can_manage_feeders(user, distribution_center, source_type=None):
    """Check if user can manage feeders for a distribution center.
    Only provincial users can manage feeders for DCs under their office.
    DC users can manage ENERGY_IMPORT and ENERGY_EXPORT types for their own DC."""
    if getattr(user, 'is_system_admin', False):
        return True
    if user.is_provincial:
        po = getattr(user, 'provincial_office', None)
        if po and po.pk == distribution_center.provincial_office_id:
            return True
    # DC users can add ENERGY_IMPORT and ENERGY_EXPORT without province approval
    if user.is_dc_level and source_type in ['ENERGY_IMPORT', 'ENERGY_EXPORT']:
        dc = getattr(user, 'distribution_center', None)
        if dc and dc.pk == distribution_center.pk:
            return True
    return False


def _can_view_report(user, report):
    if getattr(user, 'is_system_admin', False):
        return True
    if user.is_top_management:
        return True
    if user.is_provincial:
        po = getattr(user, 'provincial_office', None)
        if po and po.pk == report.distribution_center.provincial_office_id:
            return True
    if user.is_dc_level:
        dc = getattr(user, 'distribution_center', None)
        if dc and dc.pk == report.distribution_center_id:
            return True
    return False


def _can_approve_report(user):
    return getattr(user, 'is_system_admin', False) or user.is_top_management or user.is_provincial



# ─────────────────────────── MESSAGING VIEWS ───────────────────────────

# Messaging
class MessageInboxView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/inbox.html'

    def get(self, request):
        inbox = Message.objects.filter(recipient=request.user).select_related('sender').order_by('-created_at')
        sent = Message.objects.filter(sender=request.user).select_related('recipient').order_by('-created_at')
        unread_count = inbox.filter(is_read=False).count()
        
        return render(request, self.template_name, {
            'inbox': inbox,
            'sent': sent,
            'unread_count': unread_count,
        })


class MessageComposeView(LoginRequiredMixin, View):
    template_name = 'nea_loss/users/message_compose.html'

    def get(self, request):
        # Get all users except current user
        users = NEAUser.objects.exclude(pk=request.user.pk).order_by('full_name')
        return render(request, self.template_name, {
            'users': users,
        })

    def post(self, request):
        recipient_id = request.POST.get('recipient')
        subject = request.POST.get('subject')
        body = request.POST.get('body')

        if not recipient_id or not subject or not body:
            messages.error(request, 'Please fill in all fields.')
            return self.get(request)

        try:
            recipient = NEAUser.objects.get(pk=recipient_id)
            message = Message.objects.create(
                sender=request.user,
                recipient=recipient,
                subject=subject,
                body=body
            )
            messages.success(request, 'Message sent successfully.')
            return redirect('message_inbox')
        except NEAUser.DoesNotExist:
            messages.error(request, 'Recipient not found.')
            return self.get(request)


class MessageDetailView(LoginRequiredMixin, View):
    template_name = 'nea_loss/reports/message_detail.html'

    def get(self, request, pk):
        message = get_object_or_404(Message, Q(pk=pk) & (Q(sender=request.user) | Q(recipient=request.user)))
        
        # Mark as read if recipient
        if message.recipient == request.user and not message.is_read:
            message.is_read = True
            message.save()
        
        # Get replies
        replies = Message.objects.filter(parent=message).order_by('created_at')
        
        return render(request, self.template_name, {
            'msg': message,
            'replies': replies,
        })


@login_required
def message_delete(request, pk):
    message = get_object_or_404(Message, Q(pk=pk) & (Q(sender=request.user) | Q(recipient=request.user)))
    message.delete()
    messages.success(request, 'Message deleted successfully.')
    return redirect('message_inbox')


@login_required
def message_reply(request, pk):
    parent_message = get_object_or_404(Message, Q(pk=pk) & (Q(sender=request.user) | Q(recipient=request.user)))
    
    if request.method == 'POST':
        body = request.POST.get('body')
        if not body:
            messages.error(request, 'Please enter a message.')
            return redirect('message_detail', pk=pk)
        
        # Determine recipient (reply to the other person)
        recipient = parent_message.sender if parent_message.recipient == request.user else parent_message.recipient
        
        reply = Message.objects.create(
            sender=request.user,
            recipient=recipient,
            subject=f"Re: {parent_message.subject}",
            body=body,
            parent=parent_message
        )
        
        messages.success(request, 'Reply sent successfully.')
        return redirect('message_detail', pk=pk)


@login_required
def api_unread_messages(request):
    count = Message.objects.filter(recipient=request.user, is_read=False).count()
    return JsonResponse({'unread': count})


# ─────────────────────────── PROVINCIAL REPORT VIEWS ───────────────────────────

class ProvincialReportCreateView(LoginRequiredMixin, View):
    """Provincial office generates monthly consolidated report from DC data."""
    template_name = 'nea_loss/reports/provincial_create.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_provincial or user.is_dc_level):
                messages.error(request, 'Only Provincial Office and DC users can create monthly reports.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            provincial_offices = ProvincialOffice.objects.all()
        elif user.is_provincial:
            provincial_offices = ProvincialOffice.objects.filter(pk=user.provincial_office_id)
        else:
            # DCS users - show all provincial offices they can create reports for
            provincial_offices = ProvincialOffice.objects.all()

        return render(request, self.template_name, {
            'fiscal_years': FiscalYear.objects.all(),
            'provincial_offices': provincial_offices,
            'months_list': [
                (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
                (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
                (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
            ],
        })

    def post(self, request):
        fy_id = request.POST.get('fiscal_year')
        month_str = request.POST.get('month', '')
        # Handle "Report Till Now" case (month=0) as show_all=True
        if month_str.strip() == '0':
            month = 0
        else:
            month = int(month_str) if month_str.strip() else None
        po_id = request.POST.get('provincial_office')
        action = request.POST.get('action', 'preview')

        try:
            fy = FiscalYear.objects.get(pk=fy_id)
            po = ProvincialOffice.objects.get(pk=po_id)
        except Exception:
            messages.error(request, 'Invalid fiscal year or provincial office.')
            return redirect('provincial_report_create')

        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }

        # Auto-determine if showing all months or specific month
        show_all = month is None or month == 0  # If no month selected, show all available months
        
        # If no month selected, find the latest available month
        if show_all:
            latest_report = LossReport.objects.filter(
                distribution_center__provincial_office=po,
                fiscal_year=fy,
                status='APPROVED'
            ).order_by('-month').first()
            
            if latest_report:
                month = latest_report.month
            else:
                messages.error(request, 'No approved reports found for this provincial office and fiscal year.')
                return redirect('provincial_report_create')

        # show_all=True -> cumulative Shrawan->month view; False -> selected month only

        # Gather all DC reports under this provincial office for this FY/month
        dcs = DistributionCenter.objects.filter(provincial_office=po)
        dc_report_data = []
        grand_total_received = 0
        grand_total_utilised = 0

        for dc in dcs:
            # ── Month filter logic ──
            # "all_months" mode: include Shrawan up to selected month (cumulative view)
            # normal mode: show ONLY the selected month
            if show_all:
                dc_reports_range = LossReport.objects.filter(
                    distribution_center=dc,
                    fiscal_year=fy,
                    month__lte=month,
                    status='APPROVED'
                ).order_by('month')
            else:
                dc_reports_range = LossReport.objects.filter(
                    distribution_center=dc,
                    fiscal_year=fy,
                    month=month,
                    status='APPROVED'
                ).order_by('month')

            # Month-specific report (always the selected month)
            month_report = LossReport.objects.filter(
                distribution_center=dc, fiscal_year=fy, month=month,
                status='APPROVED'
            ).first()

            # For cumulative calculation always use Shrawan → selected month
            dc_reports_ytd = LossReport.objects.filter(
                distribution_center=dc,
                fiscal_year=fy,
                month__lte=month,
                status='APPROVED'
            ).order_by('month')

            month_received = float(month_report.total_received_kwh) if month_report else 0
            month_utilised = float(month_report.total_utilised_kwh) if month_report else 0

            # Cumulative = Σloss_so_far / Σreceived_so_far × 100 (same formula as dashboard)
            ytd_received = sum(float(r.total_received_kwh) for r in dc_reports_ytd)
            ytd_utilised = sum(float(r.total_utilised_kwh) for r in dc_reports_ytd)
            ytd_loss = ytd_received - ytd_utilised
            ytd_cl = round(ytd_loss / ytd_received * 100, 4) if ytd_received else 0

            # Monthly loss % = loss of that month only / received of that month × 100
            monthly_il = round((month_received - month_utilised) / month_received * 100, 4) if month_received else 0

            # Monthly breakdown: only for displayed range
            monthly_breakdown = {}
            for r in dc_reports_range:
                mn = r.month
                md_loss = float(r.total_received_kwh) - float(r.total_utilised_kwh)
                md_il = round(md_loss / float(r.total_received_kwh) * 100, 4) if float(r.total_received_kwh) else 0
                monthly_breakdown[mn] = {
                    'received': float(r.total_received_kwh),
                    'utilised': float(r.total_utilised_kwh),
                    'monthly_il': md_il,
                }

            # For show_all mode, use YTD cumulative totals; for specific month, use month values
            if show_all:
                grand_total_received += ytd_received
                grand_total_utilised += ytd_utilised
            else:
                grand_total_received += month_received
                grand_total_utilised += month_utilised

            # ── DC-specific provincial yearly target ──
            dc_prov_target = DCYearlyTarget.objects.filter(
                distribution_center=dc, fiscal_year=fy
            ).first()
            dc_target_pct = float(dc_prov_target.target_loss_percent) if dc_prov_target else None

            dc_report_data.append({
                'dc': dc,
                'month_received': month_received,
                'month_utilised': month_utilised,
                'monthly_il': monthly_il,
                'ytd_received': ytd_received,
                'ytd_utilised': ytd_utilised,
                'ytd_cl': ytd_cl,
                'monthly_breakdown': monthly_breakdown,
                'report_status': month_report.status if month_report else 'NO_REPORT',
                'dc_target': dc_target_pct,    # Provincial target for this DC this month
                'nea_target': float(fy.loss_target_percent),  # NEA target (for reference only)
            })

        grand_loss = grand_total_received - grand_total_utilised
        grand_il = round(grand_loss / grand_total_received * 100, 4) if grand_total_received else 0

        if action == 'save':
            # Save/update the provincial report record
            prov_report, created = ProvincialReport.objects.get_or_create(
                provincial_office=po,
                fiscal_year=fy,
                month=month,
                defaults={'created_by': request.user, 'status': 'DRAFT'}
            )
            AuditLog.objects.create(
                user=request.user,
                action='CREATE' if created else 'UPDATE',
                model_name='ProvincialReport',
                object_id=prov_report.pk,
                description=f"{'Created' if created else 'Updated'} provincial report for {po.name} - {fy.year_bs} - {month_names.get(month,'')}",
            )
            messages.success(request, f'Provincial report for {month_names.get(month,"")} saved successfully.')
        
        elif action == 'send_to_dmd':
            # Send to DMD for approval
            try:
                prov_report = ProvincialReport.objects.get(
                    provincial_office=po,
                    fiscal_year=fy,
                    month=month
                )
                if prov_report.status != 'DRAFT':
                    messages.error(request, 'Only draft reports can be sent to DMD.')
                else:
                    prov_report.status = 'SUBMITTED_TO_DMD'
                    prov_report.submitted_to_dmd_by = request.user
                    prov_report.submitted_to_dmd_at = timezone.now()
                    prov_report.save()
                    
                    # Create notification for DMD users
                    dmd_users = NEAUser.objects.filter(role='DMD', is_active=True)
                    for dmd_user in dmd_users:
                        Notification.objects.create(
                            recipient=dmd_user,
                            notification_type='REPORT_SUBMITTED',
                            title='Provincial Report Submitted for Approval',
                            message=f'Provincial report for {po.name} - {fy.year_bs} - {month_names.get(month,"")} has been submitted for DMD approval.',
                            related_report=None
                        )
                    
                    # Create audit log
                    AuditLog.objects.create(
                        user=request.user,
                        action='SUBMIT',
                        model_name='ProvincialReport',
                        object_id=prov_report.pk,
                        description=f"Submitted provincial report for {po.name} to DMD for approval",
                        ip_address=request.META.get('REMOTE_ADDR')
                    )
                    
                    messages.success(request, f'Provincial report for {month_names.get(month,"")} sent to DMD for approval.')
            except ProvincialReport.DoesNotExist:
                messages.error(request, 'You must save the report first before sending to DMD.')

        user = request.user
        if getattr(user, 'is_system_admin', False):
            provincial_offices = ProvincialOffice.objects.all()
        else:
            provincial_offices = ProvincialOffice.objects.filter(pk=user.provincial_office_id)

        # Check if there's a saved report for this combination
        try:
            saved_report = ProvincialReport.objects.get(
                provincial_office=po,
                fiscal_year=fy,
                month=month
            )
        except ProvincialReport.DoesNotExist:
            saved_report = None

        context = {
            'fiscal_years': FiscalYear.objects.all(),
            'provincial_offices': provincial_offices,
            'months_list': [
                (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
                (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
                (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
            ],
            'report_data': dc_report_data,
            'grand_total_received': grand_total_received,
            'grand_total_utilised': grand_total_utilised,
            'grand_il': round(grand_il, 2),
            'selected_fy': fy,
            'selected_month': month,
            'selected_month_name': month_names.get(month, ''),
            'selected_po': po,
            'nea_target_pct': float(fy.loss_target_percent),
            'show_all': show_all,
            'months_range': list(range(1, month + 1)) if show_all else [month],
            'month_names': month_names,
            'saved_report': saved_report,
        }
        return render(request, self.template_name, context)



class ProvincialDCReportsView(LoginRequiredMixin, View):
    """For Provincial users: shows all DCs in their province with month selector.
    Shows which DCs have submitted a report for the selected month."""
    template_name = 'nea_loss/reports/provincial_dc_reports.html'

    def get(self, request):
        user = request.user
        if not (user.is_provincial and user.provincial_office) and not getattr(user, 'is_system_admin', False):
            return redirect('dashboard')

        active_fy = FiscalYear.objects.filter(is_active=True).first()
        fiscal_years = FiscalYear.objects.all().order_by('-year_ad_start')

        # Filters
        selected_fy_id = request.GET.get('fiscal_year')
        selected_month = request.GET.get('month', '')
        try:
            selected_month = int(selected_month)
        except (ValueError, TypeError):
            selected_month = ''

        selected_fy = active_fy
        if selected_fy_id:
            try:
                selected_fy = FiscalYear.objects.get(pk=selected_fy_id)
            except FiscalYear.DoesNotExist:
                pass

        if user.is_provincial and user.provincial_office:
            dcs = DistributionCenter.objects.filter(
                provincial_office=user.provincial_office
            ).order_by('name')
        else:
            dcs = DistributionCenter.objects.all().order_by('name')

        MONTH_CHOICES = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]

        dc_rows = []
        for dc in dcs:
            row = {'dc': dc, 'report': None, 'status': 'Not Submitted'}
            if selected_fy and selected_month:
                report = LossReport.objects.filter(
                    distribution_center=dc,
                    fiscal_year=selected_fy,
                    month=selected_month,
                ).first()
                if report:
                    row['report'] = report
                    row['status'] = report.get_status_display()
                else:
                    row['status'] = 'Not Submitted'
            dc_rows.append(row)

        submitted_count = sum(1 for r in dc_rows if r['report'])
        not_submitted_count = len(dc_rows) - submitted_count

        return render(request, self.template_name, {
            'dc_rows': dc_rows,
            'fiscal_years': fiscal_years,
            'selected_fy': selected_fy,
            'selected_month': selected_month,
            'month_choices': MONTH_CHOICES,
            'provincial_office': getattr(user, 'provincial_office', None),
            'submitted_count': submitted_count,
            'not_submitted_count': not_submitted_count,
        })

class ProvincialReportPrintView(LoginRequiredMixin, View):
    """Print provincial monthly report in formal format."""
    template_name = 'nea_loss/reports/provincial_print.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not (getattr(user, 'is_system_admin', False) or user.is_provincial or user.role in ['DMD', 'MD']):
                messages.error(request, 'Only provincial managers and DMD/MD users can print provincial reports.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            provincial_offices = ProvincialOffice.objects.all()
        else:
            provincial_offices = ProvincialOffice.objects.filter(pk=user.provincial_office_id)

        return render(request, self.template_name, {
            'provincial_offices': provincial_offices,
            'fiscal_years': FiscalYear.objects.all(),
            'months_list': [
                (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
                (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
                (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
            ],
        })

    def post(self, request):
        fy_id = request.POST.get('fiscal_year')
        month_str = request.POST.get('month', '')
        po_id = request.POST.get('provincial_office')
        
        try:
            fy = FiscalYear.objects.get(pk=fy_id)
            po = ProvincialOffice.objects.get(pk=po_id)
            # Handle "Report Till Now" case (month=0) as show_all=True
            if month_str.strip() == '0':
                month = 0
            else:
                month = int(month_str) if month_str.strip() else None
        except Exception:
            messages.error(request, 'Invalid fiscal year or provincial office.')
            return redirect('provincial_report_create')

        # Auto-determine if showing all months or specific month
        show_all = month is None or month == 0  # If no month selected, show all available months
        
        # If no month selected, find latest available month
        if show_all:
            latest_report = LossReport.objects.filter(
                distribution_center__provincial_office=po,
                fiscal_year=fy,
                status='APPROVED'
            ).order_by('-month').first()
            
            if latest_report:
                month = latest_report.month
            else:
                messages.error(request, 'No approved reports found for this provincial office and fiscal year.')
                return redirect('provincial_report_create')

        # Get all distribution centers under this provincial office
        distribution_centers = DistributionCenter.objects.filter(provincial_office=po)
        
        # Prepare monthly data for each DC
        months_range = list(range(1, month + 1)) if show_all else [month]
        months_list = [
            (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
            (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
            (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
        ]
        
        # Build monthly data structure for each DC using existing LossReport data
        dc_monthly_data = []
        for dc in distribution_centers:
            dc_data = []
            for m in months_range:
                month_name = dict(months_list).get(m, '')
                
                # Get monthly report data
                monthly_report = LossReport.objects.filter(
                    distribution_center=dc, fiscal_year=fy, month=m
                ).first()
                
                # Get monthly loss data
                monthly_loss_data = MonthlyLossData.objects.filter(
                    report__distribution_center=dc,
                    report__fiscal_year=fy,
                    month=m
                ).select_related('report')
                
                # Calculate totals from monthly loss data
                total_energy_import = sum(data.net_energy_received for data in monthly_loss_data)
                total_energy_utilised = sum(data.total_energy_utilised for data in monthly_loss_data)
                total_loss = sum(data.monthly_loss_kwh for data in monthly_loss_data)
                
                dc_data.append({
                    'month_num': m,
                    'month_name': month_name,
                    'total_energy_import': total_energy_import,
                    'total_energy_export': 0,  # Would need to calculate from export meter points
                    'net_energy_received': total_energy_import,
                    'total_energy_utilised': total_energy_utilised,
                    'loss_unit': total_loss,
                    'monthly_loss_percent': round((total_loss / total_energy_import) * 100, 4) if total_energy_import > 0 else 0,
                    'cumulative_loss_percent': round((total_loss / total_energy_import) * 100, 4) if total_energy_import > 0 else 0,
                    'dc_count': 1,
                    'submitted_count': 1 if monthly_report else 0,
                    'not_submitted_count': 0 if monthly_report else 1,
                    'meter_readings': monthly_loss_data,  # For template compatibility
                })
            
            dc_monthly_data.append({
                'dc': dc,
                'months': dc_data,
                'total_received_kwh': sum(data['net_energy_received'] for data in dc_data),
                'total_utilised_kwh': sum(data['total_energy_utilised'] for data in dc_data),
                'total_loss_kwh': sum(data['loss_unit'] for data in dc_data),
                'cumulative_loss_percent': sum(data['monthly_loss_percent'] for data in dc_data) / len(dc_data) if dc_data else 0,
                'overall_loss_percent': sum(data['monthly_loss_percent'] for data in dc_data) / len(dc_data) if dc_data else 0,
            })
        
        # Calculate provincial totals
        total_dcs = len(distribution_centers)
        total_submitted = sum(1 for dc_data in dc_monthly_data for month_data in dc_data['months'] if month_data.get('submitted_count', 0))
        total_not_submitted = total_dcs - total_submitted

        return render(request, self.template_name, {
            'provincial_office': po,
            'selected_fy': fy,
            'selected_month': month,
            'selected_month_name': dict(months_list).get(month, ''),
            'distribution_centers': dc_monthly_data,
            'months': months_range,
            'total_dcs': total_dcs,
            'total_submitted': total_submitted,
            'total_not_submitted': total_not_submitted,
        })

class ProvincialReportListView(LoginRequiredMixin, View):
    """List of saved provincial reports."""
    template_name = 'nea_loss/reports/provincial_list.html'

    def get(self, request):
        user = request.user
        status_filter = request.GET.get('status', 'all')
        
        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            reports = ProvincialReport.objects.select_related('provincial_office', 'fiscal_year', 'created_by').all()
        elif user.is_provincial and user.provincial_office:
            reports = ProvincialReport.objects.filter(provincial_office=user.provincial_office).select_related('provincial_office', 'fiscal_year')
        else:
            reports = ProvincialReport.objects.none()
        
        # Apply status filter
        if status_filter != 'all':
            reports = reports.filter(status=status_filter)
        
        # Calculate status counts
        if getattr(user, 'is_system_admin', False) or user.is_top_management:
            all_reports = ProvincialReport.objects.all()
        elif user.is_provincial and user.provincial_office:
            all_reports = ProvincialReport.objects.filter(provincial_office=user.provincial_office)
        else:
            all_reports = ProvincialReport.objects.none()
        
        status_counts = {
            'all': all_reports.count(),
            'draft': all_reports.filter(status='DRAFT').count(),
            'pending': all_reports.filter(status='SUBMITTED_TO_DMD').count(),
            'approved': all_reports.filter(status='DMD_APPROVED').count(),
            'rejected': all_reports.filter(status='DMD_REJECTED').count(),
        }
        
        return render(request, self.template_name, {
            'reports': reports,
            'fiscal_years': FiscalYear.objects.all(),
            'status_counts': status_counts,
        })


class ProvincialApprovedReportsView(LoginRequiredMixin, View):
    """Provincial users can view their DMD-approved reports."""
    template_name = 'nea_loss/reports/provincial_approved_reports.html'

    def get(self, request):
        user = request.user
        if not user.is_provincial or not user.provincial_office:
            messages.error(request, 'Only provincial users can view their approved reports.')
            return redirect('dashboard')
        
        # Get only DMD-approved reports for this provincial office
        approved_reports = ProvincialReport.objects.filter(
            provincial_office=user.provincial_office,
            status='DMD_APPROVED'
        ).select_related('fiscal_year', 'created_by', 'dmd_reviewed_by').order_by('-dmd_reviewed_at')
        
        return render(request, self.template_name, {
            'approved_reports': approved_reports,
            'provincial_office': user.provincial_office,
        })


class ProvincialReportDetailView(LoginRequiredMixin, View):
    """View saved provincial report details with full detailed data."""
    template_name = 'nea_loss/reports/provincial_detail.html'

    def get(self, request, pk):
        user = request.user
        if not (getattr(user, 'is_system_admin', False) or user.is_provincial or user.role in ['DMD', 'MD']):
            messages.error(request, 'Only provincial users and DMD/MD users can view provincial reports.')
            return redirect('dashboard')
        
        # Check if print parameter is set
        print_mode = request.GET.get('print', 'false').lower() == 'true'
        
        try:
            report = ProvincialReport.objects.select_related(
                'provincial_office', 'fiscal_year', 'created_by', 'dmd_reviewed_by', 'submitted_to_dmd_by'
            ).get(pk=pk)
        except ProvincialReport.DoesNotExist:
            messages.error(request, 'Provincial report not found.')
            return redirect('provincial_report_list')
        
        # If print mode, redirect to print template
        if print_mode:
            return self.handle_print_request(request, report)
        
        # Generate complete detailed report data (same as ProvincialReportCreateView)
        po = report.provincial_office
        fy = report.fiscal_year
        month = report.month
        
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }
        
        print(f"DEBUG: Provincial Report Detail - PO: {po.name}, FY: {fy.year_bs}, Month: {month}")
        
        show_all = False  # Detail view shows specific month
        
        # Gather all DC reports under this provincial office for this FY/month
        dcs = DistributionCenter.objects.filter(provincial_office=po)
        dc_report_data = []
        grand_total_received = 0
        grand_total_utilised = 0

        for dc in dcs:
            print(f"DEBUG: Processing DC: {dc.name}")
            # Month-specific report (always the selected month)
            month_report = LossReport.objects.filter(
                distribution_center=dc, fiscal_year=fy, month=month,
                status='APPROVED'
            ).first()
            print(f"DEBUG: Found month_report: {month_report}")

            # For cumulative calculation always use Shrawan → selected month
            dc_reports_ytd = LossReport.objects.filter(
                distribution_center=dc,
                fiscal_year=fy,
                month__lte=month,
                status='APPROVED'
            ).order_by('month')

            month_received = float(month_report.total_received_kwh) if month_report else 0
            month_utilised = float(month_report.total_utilised_kwh) if month_report else 0

            # Cumulative = Σloss_so_far / Σreceived_so_far × 100 (same formula as dashboard)
            ytd_received = sum(float(r.total_received_kwh) for r in dc_reports_ytd)
            ytd_utilised = sum(float(r.total_utilised_kwh) for r in dc_reports_ytd)
            ytd_loss = ytd_received - ytd_utilised
            ytd_cl = round(ytd_loss / ytd_received * 100, 4) if ytd_received else 0

            # Monthly loss % = loss of that month only / received of that month × 100
            monthly_il = round((month_received - month_utilised) / month_received * 100, 4) if month_received else 0

            # Get DC targets
            dc_target_pct = None
            try:
                dc_target = DCYearlyTarget.objects.get(
                    distribution_center=dc,
                    fiscal_year=fy
                )
                dc_target_pct = dc_target.target_loss_percent
            except DCYearlyTarget.DoesNotExist:
                dc_target_pct = None
            except Exception as e:
                # Handle any other errors with target fetching
                dc_target_pct = None
            
            dc_report_data.append({
                'dc': dc,
                'total_received': month_received,
                'total_utilised': month_utilised,
                'total_loss': month_received - month_utilised,
                'monthly_il': monthly_il,
                'ytd_cl': ytd_cl,
                'report_status': month_report.status if month_report else 'NO_REPORT',
                'dc_target': dc_target_pct,
                'nea_target': float(fy.loss_target_percent),
                'submitted_count': 1 if month_report else 0,
                'not_submitted_count': 0 if month_report else 1,
                'dc_count': 1,
            })
            
            grand_total_received += month_received
            grand_total_utilised += month_utilised

        # Calculate grand totals
        grand_loss = grand_total_received - grand_total_utilised
        grand_il = round(grand_loss / grand_total_received * 100, 4) if grand_total_received else 0
        
        print(f"DEBUG: Final dc_report_data count: {len(dc_report_data)}")
        print(f"DEBUG: Grand totals - Received: {grand_total_received}, Utilised: {grand_total_utilised}")
        
        return render(request, self.template_name, {
            'report': report,
            'fiscal_years': FiscalYear.objects.all(),
            # Complete detailed report data
            'report_data': dc_report_data,
            'grand_total_received': grand_total_received,
            'grand_total_utilised': grand_total_utilised,
            'grand_total_loss': grand_loss,
            'grand_il': round(grand_il, 2),
            'selected_fy': fy,
            'selected_month': month,
            'selected_month_name': month_names.get(month, ''),
            'selected_po': po,
            'nea_target_pct': float(fy.loss_target_percent),
            'show_all': show_all,
            'months_range': [month],
            'month_names': month_names,
            'total_dcs': len(dcs),
            'total_submitted': sum(1 for item in dc_report_data if item['submitted_count']),
            'total_not_submitted': sum(1 for item in dc_report_data if item['not_submitted_count']),
            'overall_loss_percent': grand_il,
        })

    def handle_print_request(self, request, report):
        """Handle print request by rendering print template."""
        # Generate the same data as the detail view
        po = report.provincial_office
        fy = report.fiscal_year
        month = report.month
        
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }
        
        show_all = False
        dcs = DistributionCenter.objects.filter(provincial_office=po)
        dc_report_data = []
        grand_total_received = 0
        grand_total_utilised = 0

        for dc in dcs:
            # Month-specific report (always the selected month)
            month_report = LossReport.objects.filter(
                distribution_center=dc, fiscal_year=fy, month=month,
                status='APPROVED'
            ).first()

            # For cumulative calculation always use Shrawan → selected month
            dc_reports_ytd = LossReport.objects.filter(
                distribution_center=dc,
                fiscal_year=fy,
                month__lte=month,
                status='APPROVED'
            ).order_by('month')

            month_received = float(month_report.total_received_kwh) if month_report else 0
            month_utilised = float(month_report.total_utilised_kwh) if month_report else 0

            # Cumulative = Σloss_so_far / Σreceived_so_far × 100 (same formula as dashboard)
            ytd_received = sum(float(r.total_received_kwh) for r in dc_reports_ytd)
            ytd_utilised = sum(float(r.total_utilised_kwh) for r in dc_reports_ytd)
            ytd_loss = ytd_received - ytd_utilised
            ytd_cl = round(ytd_loss / ytd_received * 100, 4) if ytd_received else 0

            # Monthly loss % = loss of that month only / received of that month × 100
            monthly_il = round((month_received - month_utilised) / month_received * 100, 4) if month_received else 0

            # Get DC targets
            dc_target_pct = None
            try:
                dc_target = DCYearlyTarget.objects.get(
                    distribution_center=dc,
                    fiscal_year=fy
                )
                dc_target_pct = dc_target.target_loss_percent
            except DCYearlyTarget.DoesNotExist:
                dc_target_pct = None
            except Exception as e:
                # Handle any other errors with target fetching
                dc_target_pct = None
            
            dc_report_data.append({
                'dc': dc,
                'total_received': month_received,
                'total_utilised': month_utilised,
                'total_loss': month_received - month_utilised,
                'monthly_il': monthly_il,
                'ytd_cl': ytd_cl,
                'report_status': month_report.status if month_report else 'NO_REPORT',
                'dc_target': dc_target_pct,
                'nea_target': float(fy.loss_target_percent),
                'submitted_count': 1 if month_report else 0,
                'not_submitted_count': 0 if month_report else 1,
                'dc_count': 1,
            })
            
            grand_total_received += month_received
            grand_total_utilised += month_utilised

        # Calculate grand totals
        grand_loss = grand_total_received - grand_total_utilised
        grand_il = round(grand_loss / grand_total_received * 100, 4) if grand_total_received else 0
        
        return render(request, 'nea_loss/reports/provincial_print_fixed.html', {
            'report': report,
            'provincial_office': po,
            'selected_fy': fy,
            'selected_month': month,
            'selected_month_name': month_names.get(month, ''),
            'report_data': dc_report_data,
            'grand_total_received': grand_total_received,
            'grand_total_utilised': grand_total_utilised,
            'grand_total_loss': grand_loss,
            'grand_il': grand_il,
            'nea_target_pct': float(fy.loss_target_percent),
            'show_all': show_all,
            'months_range': [month],
            'month_names': month_names,
            'total_dcs': len(dcs),
            'total_submitted': sum(1 for item in dc_report_data if item['submitted_count']),
            'total_not_submitted': sum(1 for item in dc_report_data if item['not_submitted_count']),
            'overall_loss_percent': grand_il,
        })


class ProvincialReportReviewView(LoginRequiredMixin, View):
    """Minimal data-only view for reviewing provincial reports without extra UI elements"""
    template_name = 'nea_loss/reports/provincial_review.html'

    def get(self, request, pk):
        user = request.user
        if not (getattr(user, 'is_system_admin', False) or user.is_provincial or user.role in ['DMD', 'MD']):
            messages.error(request, 'Only provincial users and DMD/MD users can view provincial reports.')
            return redirect('dashboard')
        
        try:
            report = ProvincialReport.objects.select_related(
                'provincial_office', 'fiscal_year', 'created_by', 'dmd_reviewed_by', 'submitted_to_dmd_by'
            ).get(pk=pk)
        except ProvincialReport.DoesNotExist:
            messages.error(request, 'Provincial report not found.')
            return redirect('provincial_report_list')
        
        # Generate simplified report data for review
        po = report.provincial_office
        fy = report.fiscal_year
        month = report.month
        
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }
        
        dcs = DistributionCenter.objects.filter(provincial_office=po)
        dc_report_data = []
        grand_total_received = 0
        grand_total_utilised = 0

        for dc in dcs:
            # Get monthly data for this DC and month
            monthly_data_list = MonthlyLossData.objects.filter(
                report__distribution_center=dc,
                report__fiscal_year=fy,
                report__month=month
            ).select_related('report')
            
            dc_total_received = 0
            dc_total_utilised = 0
            
            for monthly_data in monthly_data_list:
                dc_total_received += monthly_data.net_energy_received
                dc_total_utilised += monthly_data.total_energy_utilised
            
            # Get DC targets (use yearly targets as monthly targets are deprecated)
            dc_target_pct = None
            try:
                dc_target = DCYearlyTarget.objects.get(
                    distribution_center=dc,
                    fiscal_year=fy
                )
                dc_target_pct = float(dc_target.target_loss_percent)
            except DCYearlyTarget.DoesNotExist:
                dc_target_pct = None
            except Exception:
                dc_target_pct = None
            
            # Calculate DC loss percentages
            dc_loss = dc_total_received - dc_total_utilised
            dc_loss_pct = round(dc_loss / dc_total_received * 100, 4) if dc_total_received > 0 else 0
            
            # Get DC monthly report status
            month_report = None
            try:
                month_report = LossReport.objects.get(
                    distribution_center=dc,
                    fiscal_year=fy,
                    month=month
                )
            except LossReport.DoesNotExist:
                pass
            
            dc_report_data.append({
                'dc': dc,
                'total_received': dc_total_received,
                'total_utilised': dc_total_utilised,
                'total_loss': dc_loss,
                'monthly_il': dc_loss_pct,
                'report_status': month_report.status if month_report else 'NO_REPORT',
                'dc_target': dc_target_pct,
            })
            
            grand_total_received += dc_total_received
            grand_total_utilised += dc_total_utilised

        # Calculate grand totals
        grand_loss = grand_total_received - grand_total_utilised
        grand_il = round(grand_loss / grand_total_received * 100, 4) if grand_total_received else 0
        
        return render(request, self.template_name, {
            'report': report,
            'report_data': dc_report_data,
            'grand_total_received': grand_total_received,
            'grand_total_utilised': grand_total_utilised,
            'grand_total_loss': grand_loss,
            'grand_il': round(grand_il, 2),
            'nea_target_pct': float(fy.loss_target_percent),
            'total_dcs': len(dcs),
            'total_submitted': sum(1 for item in dc_report_data if item['report_status'] != 'NO_REPORT'),
            'total_not_submitted': sum(1 for item in dc_report_data if item['report_status'] == 'NO_REPORT'),
            'month_name': month_names.get(month, ''),
            'can_approve': user.role in ['DMD', 'MD'],
        })


def _generate_excel_report(report):
    """Generate the formatted Excel loss report matching NEA format"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loss Analysis Report"

    # Styles
    header_font = Font(name='Arial', bold=True, size=14, color='1B4F72')
    title_font = Font(name='Arial', bold=True, size=11)
    label_font = Font(name='Arial', size=10)
    number_font = Font(name='Arial', size=10)
    blue_fill = PatternFill('solid', start_color='D6EAF8')
    green_fill = PatternFill('solid', start_color='D5F5E3')
    red_fill = PatternFill('solid', start_color='FADBD8')
    gray_fill = PatternFill('solid', start_color='F2F3F4')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    def thin_border():
        thin = Side(style='thin')
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    # Set column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20
    for col in ['E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 12

    row = 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = 'Nepal Electricity Authority'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=16, color='1B4F72')
    ws[f'A{row}'].alignment = center

    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f'{report.distribution_center.name}'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=13)
    ws[f'A{row}'].alignment = center

    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = 'Loss Analysis Report'
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].alignment = center

    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f'Fiscal Year: {report.fiscal_year.year_bs}'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=11)
    ws[f'A{row}'].alignment = center

    row += 1
    ws.merge_cells(f'A{row}:J{row}')
    ws[f'A{row}'] = f'Cumulative Loss (%) NEA Target: {report.fiscal_year.loss_target_percent}%'
    # Provincial monthly targets
    row += 1
    ws[f'A{row}'] = 'Provincial Monthly Targets (Loss %)'
    ws[f'A{row}'].font = Font(name='Arial', bold=True, size=10)
    ws[f'A{row}'].fill = PatternFill('solid', start_color='D6EAF8')
    row += 1
    # Get months data for provincial targets section
    months = list(report.monthly_data.order_by('month'))
    prov_target_headers = ['Month'] + [m.month_name for m in months]
    for ci, h in enumerate(prov_target_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = PatternFill('solid', start_color='EBF5FB')
        cell.alignment = Alignment(horizontal='center')
    row += 1
    prov_targets = DCYearlyTarget.objects.filter(
        distribution_center=report.distribution_center, fiscal_year=report.fiscal_year
    ).first()
    prov_target_pct = float(prov_targets.target_loss_percent) if prov_targets else None
    ws.cell(row=row, column=1, value='Target (%)').font = Font(name='Arial', bold=True, size=9)
    for mi, md in enumerate(months):
        val = prov_target_pct if prov_target_pct else '—'
        cell = ws.cell(row=row, column=2 + mi, value=val)
        cell.font = Font(name='Arial', size=9)
        cell.alignment = Alignment(horizontal='center')
        if isinstance(val, float):
            cell.number_format = '0.000%'
            cell.value = val / 100
    ws[f'A{row}'].font = Font(name='Arial', size=10, color='C0392B')
    ws[f'A{row}'].alignment = center

    row += 2
    month_names = [m.month_name for m in months]

    # Summary header
    headers = ['', '', 'Particular', '', ] + month_names + ['Total', 'Remarks']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = title_font
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border()

    row += 1
    # Summary rows
    summary_data = [
        ('Total Received Unit (kWh)', [float(m.net_energy_received) for m in months], float(report.total_received_kwh), green_fill),
        ('Total Utilised Unit (kWh)', [float(m.total_energy_utilised) for m in months], float(report.total_utilised_kwh), None),
        ('Loss Unit (kWh)', [float(m.loss_unit) for m in months], float(report.total_loss_kwh), red_fill),
        ('Monthly Loss Percentage', [round(float(m.monthly_loss_percent) * 100, 4) for m in months], None, None),
        ('Cumulative Loss Percentage', [round(float(m.cumulative_loss_percent) * 100, 4) for m in months], round(float(report.cumulative_loss_percent) * 100, 4), None),
    ]

    for label, values, total, fill in summary_data:
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].alignment = left
        if fill:
            ws[f'A{row}'].fill = fill
        for ci, v in enumerate(values, 5):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = number_font
            cell.alignment = right
            cell.number_format = '#,##0.00'
            cell.border = thin_border()
            if fill:
                cell.fill = fill
        total_cell = ws.cell(row=row, column=5 + len(values), value=total if total is not None else '')
        total_cell.font = Font(name='Arial', bold=True, size=10)
        total_cell.alignment = right
        total_cell.border = thin_border()
        if fill:
            total_cell.fill = fill
        row += 1

    row += 1

    # Preload data for section tables
    import_types = ['SUBSTATION', 'FEEDER_11KV', 'FEEDER_33KV', 'INTERBRANCH', 'IPP', 'ENERGY_IMPORT']
    export_types = ['EXPORT_DC', 'EXPORT_IPP', 'ENERGY_EXPORT']
    all_points = MeterPoint.objects.filter(
        distribution_center=report.distribution_center,
        source_type__in=import_types + export_types,
    )

    month_ids = [m.pk for m in months]
    meter_readings = MeterReading.objects.filter(
        monthly_data_id__in=month_ids,
        meter_point__in=all_points,
    ).select_related('meter_point')
    
    # Create detailed reading map for all meter readings
    reading_map = {(r.monthly_data_id, r.meter_point_id): r for r in meter_readings}
    
    # Also create unit_kwh map for summary calculations
    unit_kwh_map = {(r.monthly_data_id, r.meter_point_id): float(r.unit_kwh) for r in meter_readings}

    # Build set of (monthly_data_id, meter_point_id) pairs deleted for that specific month.
    # These cells must show as blank (not 0) — feeder was not present in that month.
    deleted_for_month = set(
        MonthlyMeterPointStatus.objects.filter(
            monthly_data_id__in=month_ids,
            is_active=False,
        ).values_list('monthly_data_id', 'meter_point_id')
    )

    cats = ConsumerCategory.objects.filter(is_active=True).filter(
        Q(distribution_center__isnull=True) | Q(distribution_center_id=report.distribution_center_id)
    ).order_by('display_order', 'name')

    energy_utilisations = EnergyUtilisation.objects.filter(
        monthly_data_id__in=month_ids,
        consumer_category__in=cats,
    )
    eu_map = {(e.monthly_data_id, e.consumer_category_id): e for e in energy_utilisations}

    consumer_counts = ConsumerCount.objects.filter(
        monthly_data_id__in=month_ids,
        consumer_category__in=cats,
    )
    cc_map = {(c.monthly_data_id, c.consumer_category_id): c for c in consumer_counts}

    def write_energy_section(title, points, source_types, start_row):
        points = list(points)
        # Columns: A=S.No, B=Name, then months, then Total
        last_col = 2 + len(months) + 1
        last_col_letter = get_column_letter(last_col)
        ws.merge_cells(f'A{start_row}:{last_col_letter}{start_row}')
        ws[f'A{start_row}'] = title
        ws[f'A{start_row}'].font = title_font
        ws[f'A{start_row}'].fill = gray_fill
        ws[f'A{start_row}'].alignment = left

        header_row = start_row + 1
        headers = ['S.No.', 'Meter / Consumer Source'] + month_names + ['Total (kWh)']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=ci, value=h)
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = thin_border()

        table_row = header_row + 1
        for idx, mp in enumerate(points, 1):
            ws.cell(row=table_row, column=1, value=idx).alignment = center
            ws.cell(row=table_row, column=2, value=mp.name).alignment = left
            row_sum = 0.0
            for mi, md in enumerate(months):
                # If feeder was deleted for this specific month, leave cell blank
                if (md.pk, mp.pk) in deleted_for_month:
                    cell = ws.cell(row=table_row, column=3 + mi, value='—')
                    cell.alignment = center
                    cell.font = Font(name='Arial', size=9, color='AAAAAA')
                elif mp.source_type in source_types:
                    val = unit_kwh_map.get((md.pk, mp.pk), 0.0)
                    row_sum += val
                    cell = ws.cell(row=table_row, column=3 + mi, value=val)
                    cell.number_format = '#,##0.00'
                    cell.alignment = right
                else:
                    cell = ws.cell(row=table_row, column=3 + mi, value=0.0)
                    cell.number_format = '#,##0.00'
                    cell.alignment = right
            total_cell = ws.cell(row=table_row, column=3 + len(months), value=row_sum)
            total_cell.number_format = '#,##0.00'
            total_cell.font = Font(name='Arial', bold=True, size=9)
            total_cell.alignment = right
            table_row += 1

        # Grand total
        ws.cell(row=table_row, column=2, value='Grand Total:').font = title_font
        grand = 0.0
        for mi, md in enumerate(months):
            col_sum = 0.0
            for mp in points:
                if (md.pk, mp.pk) not in deleted_for_month and mp.source_type in source_types:
                    col_sum += unit_kwh_map.get((md.pk, mp.pk), 0.0)
            grand += col_sum
            cell = ws.cell(row=table_row, column=3 + mi, value=col_sum)
            cell.number_format = '#,##0.00'
            cell.alignment = right
            cell.fill = green_fill
        total_cell = ws.cell(row=table_row, column=3 + len(months), value=grand)
        total_cell.number_format = '#,##0.00'
        total_cell.font = Font(name='Arial', bold=True, size=9)
        total_cell.alignment = right
        return table_row + 1

    # Section A - Energy Import (Detailed Format)
    dc_import_points = all_points.filter(source_type__in=import_types).order_by('source_type', 'name')
    
    if not dc_import_points.exists():
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'No Import Meter Points Configured for this Distribution Center'
        ws[f'A{row}'].font = Font(name='Arial', size=11, color='C0392B')
        ws[f'A{row}'].alignment = center
        row += 2
    else:
        # Section A Header
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'A. Energy Import - Detailed Meter Readings'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = gray_fill
        ws[f'A{row}'].alignment = left
        row += 1
        
        # Headers for detailed readings
        detailed_headers = ['S.No.', 'Meter Point', 'Month', 'Previous Reading', 'Present Reading', 'Difference', 'Multiplying Factor', 'Unit (kWh)']
        for ci, h in enumerate(detailed_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = thin_border()
        row += 1
        
        # Detailed meter readings data for Import points
        for idx, mp in enumerate(dc_import_points, 1):
            for mi, md in enumerate(months):
                reading = reading_map.get((md.pk, mp.pk))
                if reading and float(reading.present_reading) > 0:  # Only show rows with actual data
                    ws.cell(row=row, column=1, value=idx).alignment = center
                    ws.cell(row=row, column=2, value=mp.name).alignment = left
                    ws.cell(row=row, column=3, value=md.month_name).alignment = center
                    ws.cell(row=row, column=4, value=float(reading.previous_reading)).number_format = '#,##0.000'
                    ws.cell(row=row, column=5, value=float(reading.present_reading)).number_format = '#,##0.000'
                    ws.cell(row=row, column=6, value=float(reading.difference)).number_format = '#,##0.000'
                    ws.cell(row=row, column=7, value=float(reading.multiplying_factor)).number_format = '#,##0.000'
                    ws.cell(row=row, column=8, value=float(reading.unit_kwh)).number_format = '#,##0.00'
                    
                    # Add borders to all cells
                    for col in range(1, 9):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border()
                        cell.font = Font(name='Arial', size=8)
                    
                    row += 1
            idx += 1

    row += 2

    # Section B - Energy Export (Detailed Format)
    dc_export_points = all_points.filter(source_type__in=export_types).order_by('source_type', 'name')
    
    if not dc_export_points.exists():
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'No Export Meter Points Configured for this Distribution Center'
        ws[f'A{row}'].font = Font(name='Arial', size=11, color='C0392B')
        ws[f'A{row}'].alignment = center
        row += 2
    else:
        # Section B Header
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'] = 'B. Energy Export - Detailed Meter Readings'
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = gray_fill
        ws[f'A{row}'].alignment = left
        row += 1
        
        # Headers for detailed readings
        detailed_headers = ['S.No.', 'Meter Point', 'Month', 'Previous Reading', 'Present Reading', 'Difference', 'Multiplying Factor', 'Unit (kWh)']
        for ci, h in enumerate(detailed_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = blue_fill
            cell.alignment = center
            cell.border = thin_border()
        row += 1
        
        # Detailed meter readings data for Export points
        for idx, mp in enumerate(dc_export_points, 1):
            for mi, md in enumerate(months):
                reading = reading_map.get((md.pk, mp.pk))
                if reading and float(reading.present_reading) > 0:  # Only show rows with actual data
                    ws.cell(row=row, column=1, value=idx).alignment = center
                    ws.cell(row=row, column=2, value=mp.name).alignment = left
                    ws.cell(row=row, column=3, value=md.month_name).alignment = center
                    ws.cell(row=row, column=4, value=float(reading.previous_reading)).number_format = '#,##0.000'
                    ws.cell(row=row, column=5, value=float(reading.present_reading)).number_format = '#,##0.000'
                    ws.cell(row=row, column=6, value=float(reading.difference)).number_format = '#,##0.000'
                    ws.cell(row=row, column=7, value=float(reading.multiplying_factor)).number_format = '#,##0.000'
                    ws.cell(row=row, column=8, value=float(reading.unit_kwh)).number_format = '#,##0.00'
                    
                    # Add borders to all cells
                    for col in range(1, 9):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border()
                        cell.font = Font(name='Arial', size=8)
                    
                    row += 1
            idx += 1

    # Section C - Net Received
    last_col = 2 + len(months) + 1
    last_col_letter = get_column_letter(last_col)
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'] = 'C. Net Energy Received (kWh)'
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = gray_fill
    ws[f'A{row}'].alignment = left
    header_row = row + 1
    headers = ['Particular'] + month_names + ['Total (kWh)']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border()
    row += 2
    ws.cell(row=row, column=1, value='Net Received').font = title_font
    grand_net = 0.0
    for mi, md in enumerate(months):
        val = float(md.net_energy_received)
        grand_net += val
        cell = ws.cell(row=row, column=2 + mi, value=val)
        cell.number_format = '#,##0.00'
        cell.alignment = right
    total_cell = ws.cell(row=row, column=2 + len(months), value=float(report.total_received_kwh))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(name='Arial', bold=True, size=9)
    total_cell.alignment = right
    row += 2

    # Section D - Energy Utilised (consumer categories)
    last_col = 2 + len(months) + 1
    last_col_letter = get_column_letter(last_col)
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'] = 'D. Energy Utilised (kWh) by Consumer Category'
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = gray_fill
    ws[f'A{row}'].alignment = left
    header_row = row + 1
    headers = ['S.No.', 'Consumer Category'] + month_names + ['Total (kWh)']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border()
    table_row = header_row + 1
    for idx, cat in enumerate(cats, 1):
        ws.cell(row=table_row, column=1, value=idx).alignment = center
        ws.cell(row=table_row, column=2, value=cat.name).alignment = left
        row_sum = 0.0
        for mi, md in enumerate(months):
            eu = eu_map.get((md.pk, cat.pk))
            val = float(eu.energy_kwh) if eu else 0.0
            row_sum += val
            cell = ws.cell(row=table_row, column=3 + mi, value=val)
            cell.number_format = '#,##0.00'
            cell.alignment = right
        total_cell = ws.cell(row=table_row, column=3 + len(months), value=row_sum)
        total_cell.number_format = '#,##0.00'
        total_cell.font = Font(name='Arial', bold=True, size=9)
        total_cell.alignment = right
        table_row += 1

    # Grand total row
    ws.cell(row=table_row, column=2, value='Grand Total:').font = title_font
    for mi, md in enumerate(months):
        val = float(md.total_energy_utilised)
        cell = ws.cell(row=table_row, column=3 + mi, value=val)
        cell.number_format = '#,##0.00'
        cell.alignment = right
        cell.fill = green_fill
    total_cell = ws.cell(row=table_row, column=3 + len(months), value=float(report.total_utilised_kwh))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(name='Arial', bold=True, size=9)
    total_cell.alignment = right
    row = table_row + 2

    # Section E - Consumer Count (consumer categories)
    ws.merge_cells(f'A{row}:{last_col_letter}{row}')
    ws[f'A{row}'] = 'E. Consumer Count by Consumer Category'
    ws[f'A{row}'].font = title_font
    ws[f'A{row}'].fill = gray_fill
    ws[f'A{row}'].alignment = left
    header_row = row + 1
    headers = ['S.No.', 'Consumer Category'] + month_names + ['Total Count']
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=ci, value=h)
        cell.font = Font(name='Arial', bold=True, size=9)
        cell.fill = blue_fill
        cell.alignment = center
        cell.border = thin_border()

    table_row = header_row + 1
    for idx, cat in enumerate(cats, 1):
        ws.cell(row=table_row, column=1, value=idx).alignment = center
        ws.cell(row=table_row, column=2, value=cat.name).alignment = left
        row_sum = 0
        for mi, md in enumerate(months):
            cc = cc_map.get((md.pk, cat.pk))
            val = int(cc.count) if cc else 0
            row_sum += val
            cell = ws.cell(row=table_row, column=3 + mi, value=val)
            cell.alignment = right
        total_cell = ws.cell(row=table_row, column=3 + len(months), value=row_sum)
        total_cell.font = Font(name='Arial', bold=True, size=9)
        total_cell.alignment = right
        table_row += 1

    row = table_row + 1
    return wb


# ==================== DC REPORT OVERRIDE VIEWS ====================

class OverrideRequestView(LoginRequiredMixin, View):
    """DC users can request overrides for missing months"""
    template_name = 'nea_loss/overrides/request.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_dc_level:
            messages.error(request, 'Only DC users can request overrides.')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        dc = user.distribution_center
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        
        if not dc or not active_fy:
            messages.error(request, 'No distribution center or active fiscal year found.')
            return redirect('dashboard')

        # Find months that need overrides (missing previous months)
        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }

        # Get existing reports for this DC and fiscal year
        existing_reports = LossReport.objects.filter(
            distribution_center=dc,
            fiscal_year=active_fy
        ).order_by('month')

        # Get existing overrides
        existing_overrides = DCReportOverride.objects.filter(
            distribution_center=dc,
            fiscal_year=active_fy
        ).order_by('-created_at')

        # Show all months from start month onwards for override requests
        available_months = []
        approved_months = [r.month for r in existing_reports.filter(status='APPROVED')]
        existing_report_months = [r.month for r in existing_reports]
        
        for month_num in range(dc.report_start_month, 13):  # From start month to end
            # Only show months that don't have existing reports
            if month_num not in existing_report_months:
                available_months.append((month_num, month_names[month_num]))

        return render(request, self.template_name, {
            'distribution_center': dc,
            'fiscal_year': active_fy,
            'missing_months': available_months,
            'existing_overrides': existing_overrides,
            'month_names': month_names,
        })

    def post(self, request):
        user = request.user
        dc = user.distribution_center
        active_fy = FiscalYear.objects.filter(is_active=True).first()

        resume_month = int(request.POST.get('resume_month'))
        reason = request.POST.get('reason', '').strip()
        skip_from = request.POST.get('skip_from_month')
        skip_to = request.POST.get('skip_to_month')

        if not reason:
            messages.error(request, 'Please provide a reason for the override request.')
            return self.get(request)

        # Check if override already exists
        existing = DCReportOverride.objects.filter(
            distribution_center=dc,
            fiscal_year=active_fy,
            resume_month=resume_month
        ).first()

        if existing:
            if existing.status == 'PENDING':
                messages.error(request, 'You already have a pending request for this month.')
                return self.get(request)
            elif existing.status == 'APPROVED':
                messages.info(request, 'An override for this month has already been approved.')
                return self.get(request)

        # Create new override request
        override = DCReportOverride.objects.create(
            distribution_center=dc,
            fiscal_year=active_fy,
            resume_month=resume_month,
            reason=reason,
            requested_by=user
        )

        if skip_from and skip_to:
            override.skip_from_month = int(skip_from)
            override.skip_to_month = int(skip_to)
            override.save()

        # Notify system admins
        admin_users = NEAUser.objects.filter(role='SYS_ADMIN', is_active=True)
        for admin in admin_users:
            Notification.objects.create(
                recipient=admin,
                notification_type='OVERRIDE_REQUESTED',
                title=f'Override Request: {dc.name}',
                message=f'{user.full_name} requests override for {override.get_resume_month_display()}. Reason: {reason[:100]}...',
                related_report=None
            )

        # Create audit log
        AuditLog.objects.create(
            user=user,
            action='CREATE',
            model_name='DCReportOverride',
            object_id=override.pk,
            description=f"Requested override for {dc.name} - {active_fy.year_bs} - {override.get_resume_month_display()}"
        )

        messages.success(request, 'Override request submitted. System administrators will review your request.')
        return redirect('override_request')


class OverrideManagementView(LoginRequiredMixin, View):
    """System admin view to manage override requests"""
    template_name = 'nea_loss/overrides/manage.html'

    def dispatch(self, request, *args, **kwargs):
        if not getattr(request.user, 'is_system_admin', False):
            messages.error(request, 'Only system administrators can manage overrides.')
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        pending_overrides = DCReportOverride.objects.filter(status='PENDING').order_by('-created_at')
        all_overrides = DCReportOverride.objects.all().order_by('-created_at')

        return render(request, self.template_name, {
            'pending_overrides': pending_overrides,
            'all_overrides': all_overrides,
        })

    def post(self, request):
        override_id = request.POST.get('override_id')
        action = request.POST.get('action')  # 'approve', 'reject', 'activate', or 'deactivate'
        admin_notes = request.POST.get('admin_notes', '').strip()

        if not override_id or action not in ['approve', 'reject', 'activate', 'deactivate']:
            messages.error(request, 'Invalid request.')
            return self.get(request)

        try:
            override = DCReportOverride.objects.get(pk=override_id)
            
            if action == 'approve':
                override.approve(approved_by=request.user, admin_notes=admin_notes)
                messages.success(request, f'Override approved for {override.distribution_center.name}.')
                
                # Notify the requesting user
                Notification.objects.create(
                    recipient=override.requested_by,
                    notification_type='OVERRIDE_APPROVED',
                    title=f'Override Approved for {override.distribution_center.name}',
                    message=f'Your request to resume reporting from {override.get_resume_month_display()} has been approved.',
                    related_report=None
                )
                
                audit_action = 'APPROVE'
                
            elif action == 'reject':
                override.reject(approved_by=request.user, admin_notes=admin_notes)
                messages.success(request, f'Override rejected for {override.distribution_center.name}.')
                
                # Notify the requesting user
                Notification.objects.create(
                    recipient=override.requested_by,
                    notification_type='OVERRIDE_REJECTED',
                    title=f'Override Rejected for {override.distribution_center.name}',
                    message=f'Your request to resume reporting from {override.get_resume_month_display()} has been rejected.',
                    related_report=None
                )
                
                audit_action = 'REJECT'
                
            elif action == 'activate':
                override.activate()
                messages.success(request, f'Override activated for {override.distribution_center.name}.')
                audit_action = 'ACTIVATE'
                
            elif action == 'deactivate':
                override.deactivate()
                messages.success(request, f'Override deactivated for {override.distribution_center.name}. DC users can no longer create reports for {override.get_resume_month_display()}.')
                audit_action = 'DEACTIVATE'

            # Create audit log
            AuditLog.objects.create(
                user=request.user,
                action=audit_action,
                model_name='DCReportOverride',
                object_id=override.pk,
                description=f"{audit_action} override for {override.distribution_center.name} - {override.fiscal_year.year_bs}"
            )

        except DCReportOverride.DoesNotExist:
            messages.error(request, 'Override not found.')


# ─────────────────────────── DCS MONTHLY REPORT VIEWS ───────────────────────────

class DCSMonthlyReportCreateView(LoginRequiredMixin, View):
    """DCS users generate monthly consolidated reports from DC data."""
    template_name = 'nea_loss/reports/dcs_monthly_create.html'

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if not user.is_dc_level:
                messages.error(request, 'Only DCS users can create monthly reports.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        # Get the user's distribution center
        dc = user.distribution_center
        if not dc:
            messages.error(request, 'No distribution center assigned to your account.')
            return redirect('dashboard')

        return render(request, self.template_name, {
            'fiscal_years': FiscalYear.objects.all(),
            'distribution_center': dc,
            'months_list': [
                (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
                (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
                (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
            ],
        })

    def post(self, request):
        user = request.user
        dc = user.distribution_center
        
        if not dc:
            messages.error(request, 'No distribution center assigned to your account.')
            return redirect('dcs_monthly_report_create')

        fy_id = request.POST.get('fiscal_year')
        month_str = request.POST.get('month', '')
        # Handle "Report Till Now" case (month=0) as show_all=True
        if month_str.strip() == '0':
            original_month = 0
        else:
            original_month = int(month_str) if month_str.strip() else None
        action = request.POST.get('action', 'preview')

        try:
            fy = FiscalYear.objects.get(pk=fy_id)
        except Exception:
            messages.error(request, 'Invalid fiscal year.')
            return redirect('dcs_monthly_report_create')

        month_names = {
            1: 'Shrawan', 2: 'Bhadra', 3: 'Ashwin', 4: 'Kartik',
            5: 'Mangsir', 6: 'Poush', 7: 'Magh', 8: 'Falgun',
            9: 'Chaitra', 10: 'Baisakh', 11: 'Jestha', 12: 'Ashadh'
        }

        # Auto-determine if showing all months or specific month
        show_all = original_month is None or original_month == 0  # If no month selected, show all available months
        
        # If no month selected, find the latest available month
        if show_all:
            latest_report = LossReport.objects.filter(
                distribution_center=dc,
                fiscal_year=fy,
                status='APPROVED'
            ).order_by('-month').first()
            
            if latest_report:
                month = latest_report.month
            else:
                messages.error(request, 'No approved reports found for this distribution center and fiscal year.')
                return redirect('dcs_monthly_report_create')
        else:
            month = original_month

        # Get the report data for the selected month(s)
        if show_all:
            # Get all approved reports from Shrawan to selected month
            reports = LossReport.objects.filter(
                distribution_center=dc,
                fiscal_year=fy,
                month__lte=month,
                status='APPROVED'
            ).order_by('month')
        else:
            # Get only the specific month
            reports = LossReport.objects.filter(
                distribution_center=dc,
                fiscal_year=fy,
                month=month,
                status='APPROVED'
            )

        if not reports.exists():
            messages.error(request, 'No approved reports found for the selected period.')
            return redirect('dcs_monthly_report_create')

        # Build comprehensive report data structure
        dc_report_data = []
        grand_total_received = 0
        grand_total_utilised = 0
        cumulative_received = 0
        cumulative_loss = 0

        for report in reports:
            dc_target_pct = getattr(report, 'loss_target_percent', None)
            
            monthly_received = float(report.total_received_kwh or 0)
            monthly_utilised = float(report.total_utilised_kwh or 0)
            monthly_loss = float(report.total_loss_kwh or 0)
            monthly_il = round(monthly_loss / monthly_received * 100, 4) if monthly_received > 0 else 0
            
            # Calculate cumulative totals
            cumulative_received += monthly_received
            cumulative_loss += monthly_loss
            cumulative_il = round(cumulative_loss / cumulative_received * 100, 4) if cumulative_received > 0 else 0
            
            # Get detailed monthly data
            monthly_data = report.monthly_data.all()
            
            # Import meter readings
            import_readings = []
            import_total = 0
            for md in monthly_data:
                for reading in md.meter_readings.filter(meter_point__source_type__in=["SUBSTATION","FEEDER_11KV","FEEDER_33KV","INTERBRANCH","IPP","ENERGY_IMPORT"]):
                    unit_kwh = float(reading.unit_kwh or 0)
                    import_total += unit_kwh
                    import_readings.append({
                        'meter_point': reading.meter_point,
                        'present_reading': reading.present_reading,
                        'previous_reading': reading.previous_reading,
                        'difference': reading.difference,
                        'mf': reading.multiplying_factor,
                        'unit_kwh': unit_kwh
                    })
            
            # Export meter readings
            export_readings = []
            export_total = 0
            for md in monthly_data:
                for reading in md.meter_readings.filter(meter_point__source_type__in=["EXPORT_DC","EXPORT_IPP","ENERGY_EXPORT"]):
                    unit_kwh = float(reading.unit_kwh or 0)
                    export_total += unit_kwh
                    export_readings.append({
                        'meter_point': reading.meter_point,
                        'present_reading': reading.present_reading,
                        'previous_reading': reading.previous_reading,
                        'difference': reading.difference,
                        'mf': reading.multiplying_factor,
                        'unit_kwh': unit_kwh
                    })
            
            # Consumer utilisation data
            consumer_utilisations = []
            for md in monthly_data:
                for utilisation in md.energy_utilisations.all():
                    consumer_utilisations.append({
                        'category': utilisation.consumer_category,
                        'energy_kwh': float(utilisation.energy_kwh or 0),
                        'remarks': utilisation.remarks
                    })
            
            # Consumer count data
            consumer_counts = []
            for md in monthly_data:
                for count in md.consumer_counts.all():
                    consumer_counts.append({
                        'category': count.consumer_category,
                        'count': int(count.count or 0),
                        'remarks': count.remarks
                    })
            
            dc_report_data.append({
                'dc': dc,
                'month': report.month,
                'month_name': month_names.get(report.month, ''),
                'total_received': monthly_received,
                'total_utilised': monthly_utilised,
                'total_loss': monthly_loss,
                'monthly_il': monthly_il,
                'cumulative_il': cumulative_il,
                'dc_target': dc_target_pct,
                'nea_target': float(fy.loss_target_percent),
                'import_readings': import_readings,
                'export_readings': export_readings,
                'import_total': import_total,
                'export_total': export_total,
                'consumer_utilisations': consumer_utilisations,
                'consumer_counts': consumer_counts
            })
            
            grand_total_received += monthly_received
            grand_total_utilised += monthly_utilised

        grand_loss = grand_total_received - grand_total_utilised
        grand_il = round(grand_loss / grand_total_received * 100, 4) if grand_total_received else 0

        # Prepare consolidated data for "Report Till Now" with monthwise columns
        consolidated_imports = {}
        consolidated_exports = {}
        consolidated_utilisation = {}
        consolidated_counts = {}
        
        if show_all:
            # Get all months in the report for column headers
            report_months = [row['month_name'] for row in dc_report_data]
            
            for row in dc_report_data:
                month_name = row['month_name']
                
                # Consolidate import readings with monthwise columns
                for reading in row['import_readings']:
                    meter_name = reading['meter_point'].name
                    source_type = reading['meter_point'].source_type
                    # Create unique key combining meter name and source type
                    unique_key = f"{meter_name}_{source_type}"
                    
                    if unique_key not in consolidated_imports:
                        consolidated_imports[unique_key] = {
                            'meter_point': reading['meter_point'],
                            'monthly_data': {month: 0 for month in report_months},
                            'total_units': 0
                        }
                    consolidated_imports[unique_key]['monthly_data'][month_name] += reading['unit_kwh']
                    consolidated_imports[unique_key]['total_units'] += reading['unit_kwh']
                
                # Consolidate export readings with monthwise columns
                for reading in row['export_readings']:
                    meter_name = reading['meter_point'].name
                    source_type = reading['meter_point'].source_type
                    # Create unique key combining meter name and source type
                    unique_key = f"{meter_name}_{source_type}"
                    
                    if unique_key not in consolidated_exports:
                        consolidated_exports[unique_key] = {
                            'meter_point': reading['meter_point'],
                            'monthly_data': {month: 0 for month in report_months},
                            'total_units': 0
                        }
                    consolidated_exports[unique_key]['monthly_data'][month_name] += reading['unit_kwh']
                    consolidated_exports[unique_key]['total_units'] += reading['unit_kwh']
                
                # Consolidate consumer utilisation with monthwise columns
                for utilisation in row['consumer_utilisations']:
                    cat_name = utilisation['category'].name
                    if cat_name not in consolidated_utilisation:
                        consolidated_utilisation[cat_name] = {
                            'category': utilisation['category'],
                            'monthly_data': {month: 0 for month in report_months},
                            'total_energy': 0
                        }
                    consolidated_utilisation[cat_name]['monthly_data'][month_name] += utilisation['energy_kwh']
                    consolidated_utilisation[cat_name]['total_energy'] += utilisation['energy_kwh']
                
                # Consolidate consumer counts with monthwise columns
                for count in row['consumer_counts']:
                    cat_name = count['category'].name
                    if cat_name not in consolidated_counts:
                        consolidated_counts[cat_name] = {
                            'category': count['category'],
                            'monthly_data': {month: 0 for month in report_months},
                            'total_count': 0
                        }
                    consolidated_counts[cat_name]['monthly_data'][month_name] += count['count']
                    consolidated_counts[cat_name]['total_count'] += count['count']
            
            # Store months for template use
            consolidated_months = report_months
            
            # Ensure consolidated data is available even if empty
            if not consolidated_imports and not consolidated_exports and not consolidated_utilisation and not consolidated_counts:
                # Create empty structure to avoid template errors
                consolidated_imports = {}
                consolidated_exports = {}
                consolidated_utilisation = {}
                consolidated_counts = {}
            
            # Calculate grand totals for each consolidated section
            import_monthly_totals = {month: 0 for month in report_months}
            import_grand_total = 0
            
            for meter_name, data in consolidated_imports.items():
                for month in report_months:
                    import_monthly_totals[month] += data['monthly_data'][month]
                import_grand_total += data['total_units']
            
            export_monthly_totals = {month: 0 for month in report_months}
            export_grand_total = 0
            
            for meter_name, data in consolidated_exports.items():
                for month in report_months:
                    export_monthly_totals[month] += data['monthly_data'][month]
                export_grand_total += data['total_units']
            
            utilisation_monthly_totals = {month: 0 for month in report_months}
            utilisation_grand_total = 0
            
            for cat_name, data in consolidated_utilisation.items():
                for month in report_months:
                    utilisation_monthly_totals[month] += data['monthly_data'][month]
                utilisation_grand_total += data['total_energy']
            
            count_monthly_totals = {month: 0 for month in report_months}
            count_grand_total = 0
            
            for cat_name, data in consolidated_counts.items():
                for month in report_months:
                    count_monthly_totals[month] += data['monthly_data'][month]
                count_grand_total += data['total_count']

        context = {
            'fiscal_years': FiscalYear.objects.all(),
            'distribution_center': dc,
            'months_list': [
                (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
                (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
                (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
            ],
            'report_data': dc_report_data,
            'grand_total_received': grand_total_received,
            'grand_total_utilised': grand_total_utilised,
            'grand_il': round(grand_il, 2),
            'selected_fy': fy,
            'selected_month': month,
            'selected_month_name': month_names.get(month, ''),
            'nea_target_pct': float(fy.loss_target_percent),
            'show_all': show_all,
            'months_range': list(range(1, 13)) if show_all else [original_month if original_month is not None else 0],
            'month_names': month_names,
            'consolidated_imports': consolidated_imports,
            'consolidated_exports': consolidated_exports,
            'consolidated_utilisation': consolidated_utilisation,
            'consolidated_counts': consolidated_counts,
            'consolidated_months': consolidated_months if show_all else [],
            'import_monthly_totals': import_monthly_totals if show_all else {},
            'import_grand_total': import_grand_total if show_all else 0,
            'export_monthly_totals': export_monthly_totals if show_all else {},
            'export_grand_total': export_grand_total if show_all else 0,
            'utilisation_monthly_totals': utilisation_monthly_totals if show_all else {},
            'utilisation_grand_total': utilisation_grand_total if show_all else 0,
            'count_monthly_totals': count_monthly_totals if show_all else {},
            'count_grand_total': count_grand_total if show_all else 0,
        }
        return render(request, self.template_name, context)

        return self.get(request)


# ─────────────────────────── PROVINCE REPORTS (APPROVED) ───────────────────────────

class ProvinceReportsApprovedView(LoginRequiredMixin, View):
    """List of approved provincial reports for top management and system admins."""
    template_name = 'nea_loss/reports/province_reports_approved.html'

    def get(self, request):
        user = request.user
        if not (getattr(user, 'is_system_admin', False) or user.is_top_management):
            messages.error(request, 'Only top management and system administrators can view approved provincial reports.')
            return redirect('dashboard')

        # Get all approved provincial reports
        reports = ProvincialReport.objects.filter(status='DMD_APPROVED').select_related(
            'provincial_office', 'fiscal_year', 'created_by', 'dmd_reviewed_by'
        ).order_by('-created_at')

        return render(request, self.template_name, {
            'reports': reports,
        })


class ProvinceReportsReviewView(LoginRequiredMixin, View):
    """Minimal data-only view for approved provincial reports."""
    template_name = 'nea_loss/reports/province_reports_review.html'

    def get(self, request):
        user = request.user
        if not (getattr(user, 'is_system_admin', False) or user.is_top_management):
            messages.error(request, 'Only top management and system administrators can view approved provincial reports.')
            return redirect('dashboard')

        # Get all approved provincial reports
        reports = ProvincialReport.objects.filter(status='DMD_APPROVED').select_related(
            'provincial_office', 'fiscal_year', 'created_by', 'dmd_reviewed_by'
        ).order_by('-created_at')

        return render(request, self.template_name, {
            'reports': reports,
        })


# ─────────────────────────── DMD APPROVAL SYSTEM ───────────────────────────

class DMDApprovalDashboardView(LoginRequiredMixin, View):
    """DMD/MD dashboard for approving/rejecting provincial reports. System admin can also access."""
    template_name = 'nea_loss/approvals/dmd_approval_dashboard.html'

    def get(self, request):
        user = request.user
        if user.role not in ['DMD', 'MD'] and not getattr(user, 'is_system_admin', False):
            messages.error(request, 'Only DMD, MD users and system admin can access this page.')
            return redirect('dashboard')

        # Get reports submitted to DMD
        pending_reports = ProvincialReport.objects.filter(
            status='SUBMITTED_TO_DMD'
        ).select_related(
            'provincial_office', 'fiscal_year', 'submitted_to_dmd_by'
        ).order_by('-submitted_to_dmd_at')

        # Get recently processed reports
        processed_reports = ProvincialReport.objects.filter(
            status__in=['DMD_APPROVED', 'DMD_REJECTED']
        ).select_related(
            'provincial_office', 'fiscal_year', 'dmd_reviewed_by'
        ).order_by('-dmd_reviewed_at')[:10]

        return render(request, self.template_name, {
            'pending_reports': pending_reports,
            'processed_reports': processed_reports,
            'pending_count': pending_reports.count(),
        })


@login_required
def dmd_approve_provincial_report(request, pk):
    """Approve a provincial report"""
    user = request.user
    if user.role not in ['DMD', 'MD'] and not getattr(user, 'is_system_admin', False):
        messages.error(request, 'Only DMD, MD users and system admin can approve reports.')
        return redirect('dashboard')

    try:
        report = ProvincialReport.objects.get(pk=pk)
        if report.status != 'SUBMITTED_TO_DMD':
            messages.error(request, 'This report cannot be approved.')
            return redirect('dmd_approval_dashboard')

        report.status = 'DMD_APPROVED'
        report.dmd_reviewed_by = user
        report.dmd_reviewed_at = timezone.now()
        report.save()

        # Create notification
        Notification.objects.create(
            recipient=report.created_by,
            notification_type='REPORT_APPROVED',
            title='Provincial Report Approved',
            message=f'Your provincial report for {report.provincial_office.name} - {report.fiscal_year.year_bs} - {report.get_month_display()} has been approved by DMD.',
            related_report=None
        )

        # Create audit log
        AuditLog.objects.create(
            user=user,
            action='APPROVE',
            model_name='ProvincialReport',
            object_id=report.pk,
            description=f"DMD approved provincial report for {report.provincial_office.name}",
            ip_address=request.META.get('REMOTE_ADDR')
        )

        messages.success(request, 'Provincial report approved successfully.')
    except ProvincialReport.DoesNotExist:
        messages.error(request, 'Provincial report not found.')

    return redirect('dmd_approval_dashboard')


@login_required
def provincial_report_excel_export(request, pk):
    """Export provincial report to Excel"""
    user = request.user
    if not (getattr(user, 'is_system_admin', False) or user.is_provincial):
        messages.error(request, 'Only provincial managers can export provincial reports.')
        return redirect('dashboard')
    
    try:
        report = ProvincialReport.objects.select_related(
            'provincial_office', 'fiscal_year', 'created_by'
        ).get(pk=pk)
    except ProvincialReport.DoesNotExist:
        messages.error(request, 'Provincial report not found.')
        return redirect('provincial_report_list')
    
    # Generate Excel similar to LossReport export but with provincial data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provincial Report"
    
    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Write headers
    headers = [
        'Provincial Office', 'Fiscal Year', 'Month', 'Status',
        'Created By', 'Created At', 'DMD Reviewed By', 'DMD Reviewed At'
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Write report data
    row = 2
    ws.cell(row=row, column=1, value=report.provincial_office.name)
    ws.cell(row=row, column=2, value=report.fiscal_year.year_bs)
    ws.cell(row=row, column=3, value=report.get_month_display())
    ws.cell(row=row, column=4, value=report.get_status_display())
    ws.cell(row=row, column=5, value=report.created_by.full_name if report.created_by else '')
    ws.cell(row=row, column=6, value=report.created_at.strftime('%Y-%m-%d %H:%M'))
    ws.cell(row=row, column=7, value=report.dmd_reviewed_by.full_name if report.dmd_reviewed_by else '')
    ws.cell(row=row, column=8, value=report.dmd_reviewed_at.strftime('%Y-%m-%d %H:%M') if report.dmd_reviewed_at else '')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=provincial_report_{report.provincial_office.code}_{report.fiscal_year.year_bs}_{report.get_month_display()}.xlsx'
    wb.save(response)
    
    return response


@login_required
def dmd_reject_provincial_report(request, pk):
    """Reject a provincial report"""
    user = request.user
    if user.role not in ['DMD', 'MD'] and not getattr(user, 'is_system_admin', False):
        messages.error(request, 'Only DMD, MD users and system admin can reject reports.')
        return redirect('dashboard')

    if request.method == 'POST':
        try:
            report = ProvincialReport.objects.get(pk=pk)
            if report.status != 'SUBMITTED_TO_DMD':
                messages.error(request, 'This report cannot be rejected.')
                return redirect('dmd_approval_dashboard')

            remarks = request.POST.get('remarks', '').strip()
            if not remarks:
                messages.error(request, 'Remarks are required when rejecting a report.')
                return redirect('dmd_approval_dashboard')

            report.status = 'DMD_REJECTED'
            report.dmd_reviewed_by = user
            report.dmd_reviewed_at = timezone.now()
            report.dmd_remarks = remarks
            report.save()

            # Create notification
            Notification.objects.create(
                recipient=report.created_by,
                notification_type='REPORT_REJECTED',
                title='Provincial Report Rejected',
                message=f'Your provincial report for {report.provincial_office.name} - {report.fiscal_year.year_bs} - {report.get_month_display()} has been rejected by DMD. Remarks: {remarks}',
                related_report=None
            )

            # Create audit log
            AuditLog.objects.create(
                user=user,
                action='REJECT',
                model_name='ProvincialReport',
                object_id=report.pk,
                description=f"DMD rejected provincial report for {report.provincial_office.name}. Remarks: {remarks}",
                ip_address=request.META.get('REMOTE_ADDR')
            )

            messages.success(request, 'Provincial report rejected successfully.')
        except ProvincialReport.DoesNotExist:
            messages.error(request, 'Provincial report not found.')

    return redirect('dmd_approval_dashboard')


# ─────────────────────────── DMD CREATE REPORT ───────────────────────────

class DMDCreateReportView(LoginRequiredMixin, View):
    """DMD/MD creates comprehensive report with all provinces and DCS data to calculate overall NEA loss. System admin can also access."""
    template_name = 'nea_loss/reports/dmd_create_report.html'

    MONTH_CHOICES = [
        (0, 'All Months (Year to Date)'),
        (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
        (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
        (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
    ]

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if user.role not in ['DMD', 'MD'] and not getattr(user, 'is_system_admin', False):
                messages.error(request, 'Only DMD, MD users and system admin can access this page.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        selected_month = request.GET.get('month', '0')
        selected_month = int(selected_month) if selected_month.isdigit() else 0
        
        # Get all provinces with their distribution centers
        provinces = Province.objects.prefetch_related('offices__distribution_centers').all()
        
        province_data = []
        grand_total_received = 0
        grand_total_utilised = 0
        grand_total_loss = 0
        
        for province in provinces:
            provincial_offices = province.offices.all()
            province_received = 0
            province_utilised = 0
            province_loss = 0
            dc_count = 0
            
            office_data = []
            for office in provincial_offices:
                dcs = office.distribution_centers.filter(is_active=True)
                office_received = 0
                office_utilised = 0
                office_loss = 0
                
                dc_data = []
                for dc in dcs:
                    # Get approved reports for this DC in the active fiscal year
                    if selected_month == 0:
                        # All months (year to date)
                        reports = LossReport.objects.filter(
                            distribution_center=dc,
                            fiscal_year=active_fy,
                            status='APPROVED'
                        ) if active_fy else []
                    else:
                        # Specific month only
                        reports = LossReport.objects.filter(
                            distribution_center=dc,
                            fiscal_year=active_fy,
                            month=selected_month,
                            status='APPROVED'
                        ) if active_fy else []
                    
                    dc_received = sum(float(r.total_received_kwh) for r in reports)
                    dc_utilised = sum(float(r.total_utilised_kwh) for r in reports)
                    dc_loss = dc_received - dc_utilised
                    dc_loss_pct = round(dc_loss / dc_received * 100, 4) if dc_received else 0
                    
                    dc_data.append({
                        'dc': dc,
                        'received': dc_received,
                        'utilised': dc_utilised,
                        'loss': dc_loss,
                        'loss_pct': dc_loss_pct,
                        'report_count': reports.count(),
                    })
                    
                    office_received += dc_received
                    office_utilised += dc_utilised
                    office_loss += dc_loss
                    dc_count += 1
                
                office_loss_pct = round(office_loss / office_received * 100, 4) if office_received else 0
                
                office_data.append({
                    'office': office,
                    'dcs': dc_data,
                    'received': office_received,
                    'utilised': office_utilised,
                    'loss': office_loss,
                    'loss_pct': office_loss_pct,
                    'dc_count': len(dc_data),
                })
                
                province_received += office_received
                province_utilised += office_utilised
                province_loss += office_loss
            
            province_loss_pct = round(province_loss / province_received * 100, 4) if province_received else 0
            
            province_data.append({
                'province': province,
                'offices': office_data,
                'received': province_received,
                'utilised': province_utilised,
                'loss': province_loss,
                'loss_pct': province_loss_pct,
                'office_count': len(office_data),
                'dc_count': dc_count,
            })
            
            grand_total_received += province_received
            grand_total_utilised += province_utilised
            grand_total_loss += province_loss
        
        overall_loss_pct = round(grand_total_loss / grand_total_received * 100, 4) if grand_total_received else 0
        nea_target_pct = float(active_fy.loss_target_percent) if active_fy else 3.35
        
        # Get month name for display
        month_name = 'All Months (Year to Date)'
        if selected_month != 0:
            month_name = dict(self.MONTH_CHOICES).get(selected_month, '')
        
        return render(request, self.template_name, {
            'active_fy': active_fy,
            'province_data': province_data,
            'grand_total_received': grand_total_received,
            'grand_total_utilised': grand_total_utilised,
            'grand_total_loss': grand_total_loss,
            'overall_loss_pct': overall_loss_pct,
            'nea_target_pct': nea_target_pct,
            'total_provinces': len(province_data),
            'total_offices': sum(p['office_count'] for p in province_data),
            'total_dcs': sum(p['dc_count'] for p in province_data),
            'month_choices': self.MONTH_CHOICES,
            'selected_month': selected_month,
            'month_name': month_name,
        })


class DMDCreateReportPrintView(LoginRequiredMixin, View):
    """Print-friendly version of DMD/MD comprehensive report. System admin can also access."""
    template_name = 'nea_loss/reports/dmd_create_report_print.html'

    MONTH_CHOICES = [
        (0, 'All Months (Year to Date)'),
        (1,'Shrawan'),(2,'Bhadra'),(3,'Ashwin'),(4,'Kartik'),
        (5,'Mangsir'),(6,'Poush'),(7,'Magh'),(8,'Falgun'),
        (9,'Chaitra'),(10,'Baisakh'),(11,'Jestha'),(12,'Ashadh'),
    ]

    def dispatch(self, request, *args, **kwargs):
        user = request.user
        if user.is_authenticated:
            if user.role not in ['DMD', 'MD'] and not getattr(user, 'is_system_admin', False):
                messages.error(request, 'Only DMD, MD users and system admin can access this page.')
                return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        active_fy = FiscalYear.objects.filter(is_active=True).first()
        selected_month = request.GET.get('month', '0')
        selected_month = int(selected_month) if selected_month.isdigit() else 0
        
        # Get all provinces with their distribution centers
        provinces = Province.objects.prefetch_related('offices__distribution_centers').all()
        
        province_data = []
        grand_total_received = 0
        grand_total_utilised = 0
        grand_total_loss = 0
        
        for province in provinces:
            provincial_offices = province.offices.all()
            province_received = 0
            province_utilised = 0
            province_loss = 0
            dc_count = 0
            
            office_data = []
            for office in provincial_offices:
                dcs = office.distribution_centers.filter(is_active=True)
                office_received = 0
                office_utilised = 0
                office_loss = 0
                
                dc_data = []
                for dc in dcs:
                    # Get approved reports for this DC in the active fiscal year
                    if selected_month == 0:
                        # All months (year to date)
                        reports = LossReport.objects.filter(
                            distribution_center=dc,
                            fiscal_year=active_fy,
                            status='APPROVED'
                        ) if active_fy else []
                    else:
                        # Specific month only
                        reports = LossReport.objects.filter(
                            distribution_center=dc,
                            fiscal_year=active_fy,
                            month=selected_month,
                            status='APPROVED'
                        ) if active_fy else []
                    
                    dc_received = sum(float(r.total_received_kwh) for r in reports)
                    dc_utilised = sum(float(r.total_utilised_kwh) for r in reports)
                    dc_loss = dc_received - dc_utilised
                    dc_loss_pct = round(dc_loss / dc_received * 100, 4) if dc_received else 0
                    
                    dc_data.append({
                        'dc': dc,
                        'received': dc_received,
                        'utilised': dc_utilised,
                        'loss': dc_loss,
                        'loss_pct': dc_loss_pct,
                        'report_count': reports.count(),
                    })
                    
                    office_received += dc_received
                    office_utilised += dc_utilised
                    office_loss += dc_loss
                    dc_count += 1
                
                office_loss_pct = round(office_loss / office_received * 100, 4) if office_received else 0
                
                office_data.append({
                    'office': office,
                    'dcs': dc_data,
                    'received': office_received,
                    'utilised': office_utilised,
                    'loss': office_loss,
                    'loss_pct': office_loss_pct,
                    'dc_count': len(dc_data),
                })
                
                province_received += office_received
                province_utilised += office_utilised
                province_loss += office_loss
            
            province_loss_pct = round(province_loss / province_received * 100, 4) if province_received else 0
            
            province_data.append({
                'province': province,
                'offices': office_data,
                'received': province_received,
                'utilised': province_utilised,
                'loss': province_loss,
                'loss_pct': province_loss_pct,
                'office_count': len(office_data),
                'dc_count': dc_count,
            })
            
            grand_total_received += province_received
            grand_total_utilised += province_utilised
            grand_total_loss += province_loss
        
        overall_loss_pct = round(grand_total_loss / grand_total_received * 100, 4) if grand_total_received else 0
        nea_target_pct = float(active_fy.loss_target_percent) if active_fy else 3.35
        
        # Get month name for display
        month_name = 'All Months (Year to Date)'
        if selected_month != 0:
            month_name = dict(self.MONTH_CHOICES).get(selected_month, '')
        
        return render(request, self.template_name, {
            'active_fy': active_fy,
            'province_data': province_data,
            'grand_total_received': grand_total_received,
            'grand_total_utilised': grand_total_utilised,
            'grand_total_loss': grand_total_loss,
            'overall_loss_pct': overall_loss_pct,
            'nea_target_pct': nea_target_pct,
            'total_provinces': len(province_data),
            'total_offices': sum(p['office_count'] for p in province_data),
            'total_dcs': sum(p['dc_count'] for p in province_data),
            'month_name': month_name,
            'generated_date': timezone.now(),
        })


# ─────────────────────────── FEEDER MANAGEMENT VIEWS ───────────────────────────

class FeederListView(LoginRequiredMixin, View):
    """DC users can view their feeders (read-only). System admin can view all feeders."""
    template_name = 'nea_loss/feeders/list.html'

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            # System admin can view all feeders
            feeders = MeterPoint.objects.filter(is_active=True).order_by('distribution_center__name', 'source_type', 'name')
            pending_requests = FeederRequest.objects.filter(status='PENDING').order_by('-created_at')
            return render(request, self.template_name, {
                'distribution_center': None,
                'feeders': feeders,
                'pending_requests': pending_requests,
                'is_system_admin': True,
            })
        
        if not user.is_dc_level or not user.distribution_center:
            messages.error(request, 'Access denied. DC users only.')
            return redirect('dashboard')

        dc = user.distribution_center
        feeders = MeterPoint.objects.filter(
            distribution_center=dc,
            is_active=True
        ).order_by('source_type', 'name')

        # Get pending requests for this DC
        pending_requests = FeederRequest.objects.filter(
            distribution_center=dc,
            status='PENDING'
        ).order_by('-created_at')

        return render(request, self.template_name, {
            'distribution_center': dc,
            'feeders': feeders,
            'pending_requests': pending_requests,
        })


class FeederRequestView(LoginRequiredMixin, View):
    """DC users can request feeder additions/deletions. System admin can also request."""
    template_name = 'nea_loss/feeders/request.html'

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            # System admin can request feeder changes for any DC
            dcs = DistributionCenter.objects.filter(is_active=True).order_by('name')
            feeders = MeterPoint.objects.filter(is_active=True).order_by('distribution_center__name', 'source_type', 'name')
            return render(request, self.template_name, {
                'distribution_center': None,
                'feeders': feeders,
                'dcs': dcs,
                'source_type_choices': MeterPoint.SOURCE_TYPE_CHOICES,
                'is_system_admin': True,
            })
        
        if not user.is_dc_level or not user.distribution_center:
            messages.error(request, 'Access denied. DC users only.')
            return redirect('dashboard')

        dc = user.distribution_center
        feeders = MeterPoint.objects.filter(
            distribution_center=dc,
            is_active=True
        ).order_by('source_type', 'name')

        return render(request, self.template_name, {
            'distribution_center': dc,
            'feeders': feeders,
            'source_type_choices': MeterPoint.SOURCE_TYPE_CHOICES,
        })

    def post(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            # System admin can request feeder changes for any DC
            dc_id = request.POST.get('distribution_center_id')
            if not dc_id:
                messages.error(request, 'Please select a distribution center.')
                return redirect('feeder_request')
            
            dc = get_object_or_404(DistributionCenter, pk=dc_id)
        elif not user.is_dc_level or not user.distribution_center:
            messages.error(request, 'Access denied. DC users only.')
            return redirect('dashboard')
        else:
            dc = user.distribution_center

        request_type = request.POST.get('request_type')
        feeder_name = request.POST.get('feeder_name', '').strip()
        connection_source = request.POST.get('connection_source', '').strip()
        source_type = request.POST.get('source_type')
        voltage_level = request.POST.get('voltage_level', '').strip()
        multiplying_factor = request.POST.get('multiplying_factor', 1)
        meter_point_id = request.POST.get('meter_point_id')
        reason = request.POST.get('reason', '').strip()

        if not request_type or not feeder_name or not reason:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('feeder_request')

        if request_type == 'DELETE' and not meter_point_id:
            messages.error(request, 'Please select a feeder to delete.')
            return redirect('feeder_request')

        # Create the request
        feeder_request = FeederRequest.objects.create(
            distribution_center=dc,
            requested_by=user,
            request_type=request_type,
            feeder_name=feeder_name,
            connection_source=connection_source,
            source_type=source_type or '',
            voltage_level=voltage_level,
            multiplying_factor=multiplying_factor or 1,
            meter_point_id=meter_point_id if request_type == 'DELETE' else None,
            reason=reason,
        )

        # Notify provincial users
        provincial_users = NEAUser.objects.filter(
            provincial_office=dc.provincial_office,
            is_active=True
        )
        for prov_user in provincial_users:
            Notification.objects.create(
                recipient=prov_user,
                notification_type='FEEDER_REQUESTED',
                title=f'Feeder Change Request from {dc.name}',
                message=f'{user.full_name} has requested to {request_type.lower()} feeder "{feeder_name}". Reason: {reason}',
            )

        messages.success(request, 'Your feeder change request has been submitted to the provincial office.')
        return redirect('feeder_list')


class FeederManagementView(LoginRequiredMixin, View):
    """Provincial users can manage feeders for their DCs. System admin can manage all feeders."""
    template_name = 'nea_loss/feeders/management.html'

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            # System admin can manage all feeders
            dcs = DistributionCenter.objects.filter(is_active=True).order_by('name')
            all_feeders = MeterPoint.objects.filter(
                is_active=True
            ).select_related('distribution_center').order_by(
                'distribution_center__name',
                'source_type',
                'name'
            )
            feeders_by_dc = {}
            for feeder in all_feeders:
                dc_name = feeder.distribution_center.name
                if dc_name not in feeders_by_dc:
                    feeders_by_dc[dc_name] = []
                feeders_by_dc[dc_name].append(feeder)
            return render(request, self.template_name, {
                'provincial_office': None,
                'dcs': dcs,
                'feeders_by_dc': feeders_by_dc,
                'source_type_choices': MeterPoint.SOURCE_TYPE_CHOICES,
                'is_system_admin': True,
            })
        
        if not user.is_provincial or not user.provincial_office:
            messages.error(request, 'Access denied. Provincial users only.')
            return redirect('dashboard')

        po = user.provincial_office
        dcs = DistributionCenter.objects.filter(
            provincial_office=po,
            is_active=True
        ).order_by('name')

        # Get all feeders for all DCs under this province
        all_feeders = MeterPoint.objects.filter(
            distribution_center__provincial_office=po,
            is_active=True
        ).select_related('distribution_center').order_by(
            'distribution_center__name',
            'source_type',
            'name'
        )

        # Group feeders by DC
        feeders_by_dc = {}
        for feeder in all_feeders:
            dc_name = feeder.distribution_center.name
            if dc_name not in feeders_by_dc:
                feeders_by_dc[dc_name] = []
            feeders_by_dc[dc_name].append(feeder)

        return render(request, self.template_name, {
            'provincial_office': po,
            'dcs': dcs,
            'feeders_by_dc': feeders_by_dc,
            'source_type_choices': MeterPoint.SOURCE_TYPE_CHOICES,
        })


class FeederRequestsView(LoginRequiredMixin, View):
    """Provincial users can view and approve/reject feeder requests. System admin can view and approve/reject all requests."""
    template_name = 'nea_loss/feeders/requests.html'

    def get(self, request):
        user = request.user
        if getattr(user, 'is_system_admin', False):
            # System admin can view and approve/reject all feeder requests
            requests = FeederRequest.objects.all().select_related(
                'distribution_center', 'requested_by', 'meter_point'
            ).order_by('-created_at')
            return render(request, self.template_name, {
                'provincial_office': None,
                'requests': requests,
                'is_system_admin': True,
            })
        
        if not user.is_provincial or not user.provincial_office:
            messages.error(request, 'Access denied. Provincial users only.')
            return redirect('dashboard')

        po = user.provincial_office
        requests = FeederRequest.objects.filter(
            distribution_center__provincial_office=po
        ).select_related('distribution_center', 'requested_by', 'meter_point').order_by('-created_at')

        return render(request, self.template_name, {
            'provincial_office': po,
            'requests': requests,
        })


@login_required
@require_POST
def feeder_request_approve(request, pk):
    """Approve a feeder request"""
    user = request.user
    if not user.is_provincial and not getattr(user, 'is_system_admin', False):
        messages.error(request, 'Access denied. Provincial users and system admin only.')
        return redirect('dashboard')

    feeder_request = get_object_or_404(FeederRequest, pk=pk)
    
    # Check if this request belongs to a DC under this provincial office (skip for system admin)
    if not getattr(user, 'is_system_admin', False) and feeder_request.distribution_center.provincial_office != user.provincial_office:
        messages.error(request, 'Access denied. This request is not under your jurisdiction.')
        return redirect('feeder_requests')

    notes = request.POST.get('notes', '')
    feeder_request.approve(user, notes)

    # Notify the requester
    Notification.objects.create(
        recipient=feeder_request.requested_by,
        notification_type='FEEDER_APPROVED',
        title=f'Feeder Change Approved',
        message=f'Your request to {feeder_request.get_request_type_display()} feeder "{feeder_request.feeder_name}" has been approved.',
    )

    messages.success(request, f'Feeder request for "{feeder_request.feeder_name}" has been approved.')
    return redirect('feeder_requests')


@login_required
@require_POST
def feeder_request_reject(request, pk):
    """Reject a feeder request"""
    user = request.user
    if not user.is_provincial and not getattr(user, 'is_system_admin', False):
        messages.error(request, 'Access denied. Provincial users and system admin only.')
        return redirect('dashboard')

    feeder_request = get_object_or_404(FeederRequest, pk=pk)
    
    # Check if this request belongs to a DC under this provincial office (skip for system admin)
    if not getattr(user, 'is_system_admin', False) and feeder_request.distribution_center.provincial_office != user.provincial_office:
        messages.error(request, 'Access denied. This request is not under your jurisdiction.')
        return redirect('feeder_requests')

    notes = request.POST.get('notes', '')
    feeder_request.reject(user, notes)

    # Notify the requester
    Notification.objects.create(
        recipient=feeder_request.requested_by,
        notification_type='FEEDER_REJECTED',
        title=f'Feeder Change Rejected',
        message=f'Your request to {feeder_request.get_request_type_display()} feeder "{feeder_request.feeder_name}" has been rejected. Notes: {notes}',
    )

    messages.success(request, f'Feeder request for "{feeder_request.feeder_name}" has been rejected.')
    return redirect('feeder_requests')


# ─────────────────────────── DCS DETAIL VIEWS ───────────────────────────

def _summarize_dcs_edit_changes(edit_request):
    """Human-readable summary lines for an edit request."""
    changes = edit_request.proposed_changes or {}
    lines = []
    if changes.get('introduction'):
        lines.append('Introduction updated')
    if changes.get('established_date'):
        lines.append(f"Established date: {changes['established_date']}")
    if changes.get('coverage_area'):
        lines.append(f"Coverage: {changes['coverage_area']}")
    if changes.get('total_capacity'):
        lines.append(f"Capacity: {changes['total_capacity']} kVA")
    if edit_request.pending_image:
        lines.append('New DCS image uploaded')
    officials = changes.get('officials') or []
    if officials:
        lines.append(f"{len(officials)} official(s)")
    feeders = changes.get('feeders') or []
    if feeders:
        lines.append(f"{len(feeders)} feeder(s)")
    consumers = changes.get('consumer_types') or []
    if consumers:
        lines.append(f"{len(consumers)} consumer type(s)")
    return lines


def _get_latest_monthly_data_for_dc(dc, fiscal_year=None):
    """Return the chronologically latest MonthlyLossData row for a DC."""
    qs = MonthlyLossData.objects.filter(
        report__distribution_center=dc,
    ).select_related('report', 'report__fiscal_year')
    if fiscal_year:
        qs = qs.filter(report__fiscal_year=fiscal_year)
    return qs.order_by(
        '-report__fiscal_year__year_ad_start',
        '-report__month',
    ).first()


def record_dcs_history_snapshot(report):
    """Persist an approved monthly report into DCHistoryEntry for the History page."""
    if report.status != 'APPROVED':
        return
    monthly = report.monthly_data.filter(month=report.month).first()
    if not monthly:
        monthly = report.monthly_data.order_by('month').first()
    consumer_breakdown = {}
    total_consumers = 0
    if monthly:
        for cc in ConsumerCount.objects.filter(monthly_data=monthly).select_related('consumer_category'):
            consumer_breakdown[cc.consumer_category.name] = cc.count
            total_consumers += cc.count
    loss_pct = (
        float(report.cumulative_loss_percent) * 100
        if report.total_received_kwh and float(report.total_received_kwh) > 0
        else 0
    )
    prov_manager = None
    if report.distribution_center.provincial_office_id:
        prov_manager = NEAUser.objects.filter(
            provincial_office=report.distribution_center.provincial_office,
            role='PROVINCIAL_MANAGER',
        ).first()
    DCHistoryEntry.objects.update_or_create(
        distribution_center=report.distribution_center,
        fiscal_year=report.fiscal_year,
        month=report.month,
        defaults={
            'dc_manager': report.submitted_by or report.created_by,
            'provincial_manager': report.approved_by or prov_manager,
            'total_received_kwh': report.total_received_kwh,
            'total_utilised_kwh': report.total_utilised_kwh,
            'total_loss_kwh': report.total_loss_kwh,
            'loss_percent': round(loss_pct, 4),
            'total_consumers': total_consumers,
            'consumer_breakdown': consumer_breakdown or None,
            'report_created_at': report.created_at,
            'report_submitted_at': report.submission_date,
            'report_approved_at': report.approval_date,
        },
    )


def _backfill_history_entries(queryset_filter=None):
    """Create missing history rows from approved loss reports."""
    reports = LossReport.objects.filter(status='APPROVED').select_related(
        'distribution_center', 'fiscal_year', 'created_by', 'submitted_by', 'approved_by',
    )
    if queryset_filter is not None:
        reports = reports.filter(**queryset_filter)
    for report in reports.iterator():
        record_dcs_history_snapshot(report)


def _get_consumer_rows_from_monthly(monthly_data):
    """Build consumer category rows from the latest monthly report data."""
    if not monthly_data:
        return []
    counts = ConsumerCount.objects.filter(
        monthly_data=monthly_data,
    ).select_related('consumer_category').order_by('consumer_category__display_order', 'consumer_category__name')
    energy_map = {
        eu.consumer_category_id: eu
        for eu in EnergyUtilisation.objects.filter(monthly_data=monthly_data)
    }
    rows = []
    for cc in counts:
        eu = energy_map.get(cc.consumer_category_id)
        rows.append({
            'category_name': cc.consumer_category.name,
            'consumer_count': cc.count,
            'energy_kwh': float(eu.energy_kwh) if eu else None,
        })
    return rows


class DCSDetailView(LoginRequiredMixin, View):
    """View DCS details - picture, introduction, officials, feeders, consumer types"""
    template_name = 'nea_loss/dcs_detail/detail.html'

    def get(self, request, dc_id=None):
        user = request.user
        
        # Determine which DC to show
        if dc_id:
            dc = get_object_or_404(DistributionCenter, pk=dc_id)
        elif user.is_dc_level and user.distribution_center:
            dc = user.distribution_center
        else:
            messages.error(request, 'No distribution center specified.')
            return redirect('dashboard')
        
        # Check permissions
        if not getattr(user, 'is_system_admin', False) and not user.is_provincial and not user.is_top_management and not (user.is_dc_level and user.distribution_center == dc):
            messages.error(request, 'Access denied.')
            return redirect('dashboard')
        
        # Get or create DCS detail
        dcs_detail, created = DCSDetail.objects.get_or_create(
            distribution_center=dc
        )
        
        # Get related data
        officials = dcs_detail.officials.filter(is_active=True).order_by('designation', 'name')
        feeders = dcs_detail.feeders.filter(is_active=True).order_by('name')
        consumer_types = dcs_detail.consumer_types.all().order_by('category_name')
        
        # Get actual feeders from database
        actual_feeders = MeterPoint.objects.filter(
            distribution_center=dc,
            is_active=True
        ).order_by('name')
        
        # Get fiscal year
        current_fiscal_year = FiscalYear.objects.filter(
            is_active=True
        ).first()

        # Latest loss data for active FY (falls back to any FY if none in active year)
        latest_loss_data = _get_latest_monthly_data_for_dc(dc, current_fiscal_year)
        if not latest_loss_data and current_fiscal_year:
            latest_loss_data = _get_latest_monthly_data_for_dc(dc)

        consumer_counts = _get_consumer_rows_from_monthly(latest_loss_data)
        
        # Edit requests for this DCS
        pending_edit_request = DCSDetailEditRequest.objects.filter(
            dcs_detail=dcs_detail, status='PENDING',
        ).select_related('requested_by').order_by('-created_at').first()

        display_image = None
        image_is_pending = False
        if dcs_detail.image:
            display_image = dcs_detail.image
        elif pending_edit_request and pending_edit_request.pending_image:
            display_image = pending_edit_request.pending_image
            image_is_pending = True
        user_pending_request = None
        if user.is_dc_level and pending_edit_request and pending_edit_request.requested_by_id == user.pk:
            user_pending_request = pending_edit_request
        last_edit_request = DCSDetailEditRequest.objects.filter(
            dcs_detail=dcs_detail,
        ).exclude(status='PENDING').select_related('approved_by').order_by('-created_at').first()
        
        can_edit = (
            user.is_dc_level
            and user.distribution_center_id == dc.pk
            and pending_edit_request is None
        )

        loss_pct_display = None
        if latest_loss_data and latest_loss_data.net_energy_received:
            loss_pct_display = round(float(latest_loss_data.monthly_loss_percent) * 100, 4)

        total_profile_consumers = sum(c.consumer_count for c in consumer_types)
        total_report_consumers = sum(row.get('consumer_count', 0) for row in consumer_counts)
        feeder_record_count = feeders.count() + actual_feeders.count()

        return render(request, self.template_name, {
            'dc': dc,
            'dcs_detail': dcs_detail,
            'officials': officials,
            'feeders': feeders,
            'consumer_types': consumer_types,
            'actual_feeders': actual_feeders,
            'consumer_counts': consumer_counts,
            'latest_loss_data': latest_loss_data,
            'loss_pct_display': loss_pct_display,
            'current_fiscal_year': current_fiscal_year,
            'pending_edit_request': pending_edit_request,
            'user_pending_request': user_pending_request,
            'last_edit_request': last_edit_request,
            'can_edit': can_edit,
            'total_profile_consumers': total_profile_consumers,
            'total_report_consumers': total_report_consumers,
            'feeder_record_count': feeder_record_count,
            'has_intro': bool((dcs_detail.introduction or '').strip()),
            'has_image': bool(dcs_detail.image),
            'display_image': display_image,
            'image_is_pending': image_is_pending,
        })


class DCSDetailEditView(LoginRequiredMixin, View):
    """Edit DCS details - requires province password"""
    template_name = 'nea_loss/dcs_detail/edit.html'

    def get(self, request, dc_id):
        user = request.user
        
        if not user.is_dc_level:
            messages.error(request, 'Access denied. DC users only.')
            return redirect('dashboard')
        
        dc = get_object_or_404(DistributionCenter, pk=dc_id)
        if user.distribution_center_id != dc.pk:
            messages.error(request, 'Access denied. You can only edit your own DC details.')
            return redirect('dashboard')
        
        dcs_detail, created = DCSDetail.objects.get_or_create(distribution_center=dc)
        pending_edit = DCSDetailEditRequest.objects.filter(
            dcs_detail=dcs_detail, status='PENDING',
        ).first()
        if pending_edit:
            messages.warning(
                request,
                'You already have a pending edit request awaiting provincial approval.',
            )
            return redirect('dcs_detail', dc_id=dc.pk)
        
        return render(request, self.template_name, {
            'dc': dc,
            'dcs_detail': dcs_detail,
        })

    def post(self, request, dc_id):
        from django.db import transaction

        user = request.user
        
        if not user.is_dc_level:
            messages.error(request, 'Access denied. DC users only.')
            return redirect('dashboard')
        
        dc = get_object_or_404(DistributionCenter, pk=dc_id)
        if user.distribution_center_id != dc.pk:
            messages.error(request, 'Access denied. You can only edit your own DC details.')
            return redirect('dashboard')
        
        dcs_detail, created = DCSDetail.objects.get_or_create(distribution_center=dc)

        if DCSDetailEditRequest.objects.filter(dcs_detail=dcs_detail, status='PENDING').exists():
            messages.error(request, 'A pending edit request already exists for this DCS.')
            return redirect('dcs_detail', dc_id=dc.pk)

        try:
            proposed_changes = {
                'introduction': request.POST.get('introduction', ''),
                'established_date': request.POST.get('established_date') or None,
                'coverage_area': request.POST.get('coverage_area', ''),
                'total_capacity': request.POST.get('total_capacity') or None,
            }
            image_file = request.FILES.get('image')

            officials_data = []
            official_names = request.POST.getlist('official_name[]')
            official_designations = request.POST.getlist('official_designation[]')
            official_phones = request.POST.getlist('official_phone[]')
            official_emails = request.POST.getlist('official_email[]')
            official_joining_dates = request.POST.getlist('official_joining_date[]')

            for i, name in enumerate(official_names):
                name = (name or '').strip()
                if name:
                    officials_data.append({
                        'name': name,
                        'designation': official_designations[i] if i < len(official_designations) else '',
                        'phone': official_phones[i] if i < len(official_phones) else '',
                        'email': official_emails[i] if i < len(official_emails) else '',
                        'joining_date': official_joining_dates[i] if i < len(official_joining_dates) and official_joining_dates[i] else None,
                        'is_active': True,
                    })

            proposed_changes['officials'] = officials_data

            feeders_data = []
            feeder_names = request.POST.getlist('feeder_name[]')
            feeder_codes = request.POST.getlist('feeder_code[]')
            feeder_voltages = request.POST.getlist('feeder_voltage[]')
            feeder_lengths = request.POST.getlist('feeder_length[]')
            feeder_loads = request.POST.getlist('feeder_load[]')
            feeder_transformers = request.POST.getlist('feeder_transformers[]')

            for i, name in enumerate(feeder_names):
                name = (name or '').strip()
                if name:
                    feeders_data.append({
                        'name': name,
                        'feeder_code': feeder_codes[i] if i < len(feeder_codes) else '',
                        'voltage_level': feeder_voltages[i] if i < len(feeder_voltages) else '',
                        'length_km': feeder_lengths[i] if i < len(feeder_lengths) and feeder_lengths[i] else None,
                        'connected_load': feeder_loads[i] if i < len(feeder_loads) and feeder_loads[i] else None,
                        'transformer_count': feeder_transformers[i] if i < len(feeder_transformers) and feeder_transformers[i] else None,
                        'is_active': True,
                    })

            proposed_changes['feeders'] = feeders_data

            consumer_types_data = []
            consumer_names = request.POST.getlist('consumer_name[]')
            consumer_counts = request.POST.getlist('consumer_count[]')
            consumer_loads = request.POST.getlist('consumer_load[]')

            for i, name in enumerate(consumer_names):
                name = (name or '').strip()
                if name:
                    consumer_types_data.append({
                        'category_name': name,
                        'consumer_count': int(consumer_counts[i]) if i < len(consumer_counts) and consumer_counts[i] else 0,
                        'connected_load': consumer_loads[i] if i < len(consumer_loads) and consumer_loads[i] else 0,
                    })

            proposed_changes['consumer_types'] = consumer_types_data

            with transaction.atomic():
                edit_request = DCSDetailEditRequest.objects.create(
                    dcs_detail=dcs_detail,
                    requested_by=user,
                    proposed_changes=proposed_changes,
                    status='PENDING',
                )
                if image_file:
                    edit_request.pending_image = image_file
                    edit_request.save(update_fields=['pending_image'])

                provincial_users = NEAUser.objects.filter(
                    provincial_office=dc.provincial_office,
                    role='PROVINCIAL_MANAGER',
                    is_active=True,
                )
                for prov_user in provincial_users:
                    Notification.objects.create(
                        recipient=prov_user,
                        notification_type='FEEDER_REQUESTED',
                        title='DCS Detail Edit Request',
                        message=f'{user.full_name} has requested changes to {dc.name} details.',
                    )

            messages.success(request, 'Your edit request has been submitted for provincial approval.')
            return redirect('dcs_detail', dc_id=dc.pk)

        except Exception as exc:
            messages.error(request, f'Could not submit edit request: {exc}')
            return render(request, self.template_name, {
                'dc': dc,
                'dcs_detail': dcs_detail,
            })


class DCSDetailApprovalView(LoginRequiredMixin, View):
    """Provincial users can approve/reject DCS detail edit requests"""
    template_name = 'nea_loss/dcs_detail/approval.html'

    def get(self, request):
        user = request.user
        
        if not user.is_provincial and not getattr(user, 'is_system_admin', False):
            messages.error(request, 'Access denied. Provincial users only.')
            return redirect('dashboard')
        
        # Get pending requests
        if getattr(user, 'is_system_admin', False):
            pending_requests = DCSDetailEditRequest.objects.filter(
                status='PENDING'
            ).select_related('dcs_detail__distribution_center', 'requested_by').order_by('-created_at')
        else:
            pending_requests = DCSDetailEditRequest.objects.filter(
                status='PENDING',
                dcs_detail__distribution_center__provincial_office=user.provincial_office
            ).select_related('dcs_detail__distribution_center', 'requested_by').order_by('-created_at')
        
        # Get recent approved/rejected requests
        if getattr(user, 'is_system_admin', False):
            recent_requests = DCSDetailEditRequest.objects.filter(
                status__in=['APPROVED', 'REJECTED']
            ).select_related('dcs_detail__distribution_center', 'requested_by', 'approved_by').order_by('-created_at')[:20]
        else:
            recent_requests = DCSDetailEditRequest.objects.filter(
                status__in=['APPROVED', 'REJECTED'],
                dcs_detail__distribution_center__provincial_office=user.provincial_office
            ).select_related('dcs_detail__distribution_center', 'requested_by', 'approved_by').order_by('-created_at')[:20]
        
        for req in pending_requests:
            req.change_summary = _summarize_dcs_edit_changes(req)

        return render(request, self.template_name, {
            'pending_requests': pending_requests,
            'recent_requests': recent_requests,
        })


@login_required
@require_POST
def dcs_detail_approve(request, pk):
    """Approve a DCS detail edit request"""
    user = request.user
    
    if not user.is_provincial and not getattr(user, 'is_system_admin', False):
        messages.error(request, 'Access denied. Provincial users only.')
        return redirect('dcs_detail_approval')
    
    edit_request = get_object_or_404(DCSDetailEditRequest, pk=pk)
    
    # Check jurisdiction
    if not getattr(user, 'is_system_admin', False) and edit_request.dcs_detail.distribution_center.provincial_office != user.provincial_office:
        messages.error(request, 'Access denied. This request is not under your jurisdiction.')
        return redirect('dcs_detail_approval')
    
    # Verify password
    password = request.POST.get('approval_password')
    office = edit_request.dcs_detail.distribution_center.provincial_office
    expected = office.edit_approval_password or ''
    if not expected:
        messages.error(request, 'Provincial approval password is not set. Set it under Set Approval Password first.')
        return redirect('dcs_detail_approval')
    if password != expected:
        messages.error(request, 'Invalid approval password.')
        return redirect('dcs_detail_approval')
    
    edit_request.approve(user)
    
    # Notify the requester
    Notification.objects.create(
        recipient=edit_request.requested_by,
        notification_type='FEEDER_APPROVED',
        title='DCS Detail Edit Approved',
        message=f'Your edit request for {edit_request.dcs_detail.distribution_center.name} has been approved.',
    )
    
    messages.success(request, 'DCS detail edit has been approved.')
    return redirect('dcs_detail_approval')


@login_required
@require_POST
def dcs_detail_reject(request, pk):
    """Reject a DCS detail edit request"""
    user = request.user
    
    if not user.is_provincial and not getattr(user, 'is_system_admin', False):
        messages.error(request, 'Access denied. Provincial users only.')
        return redirect('dcs_detail_approval')
    
    edit_request = get_object_or_404(DCSDetailEditRequest, pk=pk)
    
    # Check jurisdiction
    if not getattr(user, 'is_system_admin', False) and edit_request.dcs_detail.distribution_center.provincial_office != user.provincial_office:
        messages.error(request, 'Access denied. This request is not under your jurisdiction.')
        return redirect('dcs_detail_approval')
    
    reason = request.POST.get('rejection_reason', '')
    edit_request.reject(user, reason)
    
    # Notify the requester
    Notification.objects.create(
        recipient=edit_request.requested_by,
        notification_type='FEEDER_REJECTED',
        title='DCS Detail Edit Rejected',
        message=f'Your edit request for {edit_request.dcs_detail.distribution_center.name} has been rejected. Reason: {reason}',
    )
    
    messages.success(request, 'DCS detail edit has been rejected.')
    return redirect('dcs_detail_approval')


class ProvincePasswordView(LoginRequiredMixin, View):
    """Provincial users can set their approval password"""
    template_name = 'nea_loss/dcs_detail/set_password.html'

    def get(self, request):
        user = request.user
        
        if not user.is_provincial and not getattr(user, 'is_system_admin', False):
            messages.error(request, 'Access denied. Provincial users only.')
            return redirect('dashboard')
        
        if getattr(user, 'is_system_admin', False):
            provincial_offices = ProvincialOffice.objects.all()
        else:
            provincial_offices = [user.provincial_office] if user.provincial_office else []
        
        return render(request, self.template_name, {
            'provincial_offices': provincial_offices,
        })

    def post(self, request):
        user = request.user
        
        if not user.is_provincial and not getattr(user, 'is_system_admin', False):
            messages.error(request, 'Access denied. Provincial users only.')
            return redirect('dashboard')
        
        office_id = request.POST.get('provincial_office')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        if not new_password or new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return redirect('province_password')
        
        if getattr(user, 'is_system_admin', False):
            office = get_object_or_404(ProvincialOffice, pk=office_id)
        else:
            office = user.provincial_office
        
        office.edit_approval_password = new_password
        office.save()
        
        messages.success(request, 'Approval password has been set successfully.')
        return redirect('province_password')


# ─────────────────────────── HISTORY VIEWS ───────────────────────────

class DCHistoryView(LoginRequiredMixin, View):
    """View DC history with date filters and charts"""
    template_name = 'nea_loss/dcs_history/history.html'

    def get(self, request):
        user = request.user
        active_fy = FiscalYear.objects.filter(is_active=True).first()

        fy_id = request.GET.get('fy')
        month = request.GET.get('month')
        dc_id = request.GET.get('dc')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        view_mode = request.GET.get('view', 'grouped')

        scope_filter = {}
        if user.is_dc_level and user.distribution_center_id:
            scope_filter['distribution_center_id'] = user.distribution_center_id
            dc_list = DistributionCenter.objects.filter(pk=user.distribution_center_id)
        elif user.is_provincial and user.provincial_office_id:
            scope_filter['distribution_center__provincial_office_id'] = user.provincial_office_id
            dc_list = DistributionCenter.objects.filter(
                provincial_office=user.provincial_office, is_active=True,
            ).order_by('name')
        elif getattr(user, 'is_system_admin', False) or user.is_top_management:
            dc_list = DistributionCenter.objects.filter(is_active=True).order_by('name')
        else:
            dc_list = DistributionCenter.objects.none()

        if dc_id:
            try:
                dc_pk = int(dc_id)
                if (
                    getattr(user, 'is_system_admin', False)
                    or getattr(user, 'is_top_management', False)
                    or dc_list.filter(pk=dc_pk).exists()
                ):
                    scope_filter = {'distribution_center_id': dc_pk}
            except (TypeError, ValueError):
                pass

        _backfill_history_entries(scope_filter if scope_filter else None)

        history_entries = DCHistoryEntry.objects.select_related(
            'distribution_center', 'distribution_center__provincial_office', 'fiscal_year',
            'dc_manager', 'provincial_manager',
        )
        if scope_filter:
            history_entries = history_entries.filter(**scope_filter)
        else:
            history_entries = history_entries.none()
        
        # For DC users, also fetch province-level data for comparison charts
        comparison_entries = history_entries
        if user.is_dc_level and user.distribution_center_id:
            user_dc = DistributionCenter.objects.filter(pk=user.distribution_center_id).first()
            if user_dc and user_dc.provincial_office:
                comparison_entries = DCHistoryEntry.objects.select_related(
                    'distribution_center', 'distribution_center__provincial_office', 'fiscal_year',
                ).filter(
                    distribution_center__provincial_office=user_dc.provincial_office
                )
                # Apply same filters to comparison entries
                if fy_id:
                    comparison_entries = comparison_entries.filter(fiscal_year_id=fy_id)
                elif not start_date and not end_date and active_fy:
                    comparison_entries = comparison_entries.filter(fiscal_year=active_fy)
                if month:
                    try:
                        comparison_entries = comparison_entries.filter(month=int(month))
                    except (TypeError, ValueError):
                        pass
                if start_date:
                    comparison_entries = comparison_entries.filter(report_created_at__date__gte=start_date)
                if end_date:
                    comparison_entries = comparison_entries.filter(report_created_at__date__lte=end_date)

        if fy_id:
            history_entries = history_entries.filter(fiscal_year_id=fy_id)
        elif not start_date and not end_date and active_fy:
            history_entries = history_entries.filter(fiscal_year=active_fy)

        if month:
            try:
                history_entries = history_entries.filter(month=int(month))
            except (TypeError, ValueError):
                pass

        if start_date:
            history_entries = history_entries.filter(report_created_at__date__gte=start_date)
        if end_date:
            history_entries = history_entries.filter(report_created_at__date__lte=end_date)

        history_entries = list(history_entries.order_by(
            '-fiscal_year__year_ad_start', '-month', 'distribution_center__name',
        ))

        # For DC comparison, use province-level data for DC users
        comparison_entries_list = list(comparison_entries.order_by(
            '-fiscal_year__year_ad_start', '-month', 'distribution_center__name',
        )) if user.is_dc_level else history_entries

        chart_data = self._prepare_chart_data(history_entries, comparison_entries_list)

        total_received = sum(float(e.total_received_kwh) for e in history_entries)
        total_loss = sum(float(e.total_loss_kwh) for e in history_entries)
        avg_loss_pct = round(total_loss / total_received * 100, 2) if total_received else 0

        history_by_fy = {}
        if view_mode == 'grouped':
            for entry in history_entries:
                fy_key = entry.fiscal_year.year_bs
                history_by_fy.setdefault(fy_key, []).append(entry)

        fiscal_years = FiscalYear.objects.order_by('-year_ad_start')
        months_list = [
            (1, 'Shrawan'), (2, 'Bhadra'), (3, 'Ashwin'), (4, 'Kartik'),
            (5, 'Mangsir'), (6, 'Poush'), (7, 'Magh'), (8, 'Falgun'),
            (9, 'Chaitra'), (10, 'Baisakh'), (11, 'Jestha'), (12, 'Ashadh'),
        ]

        show_dc_column = dc_list.count() > 1

        return render(request, self.template_name, {
            'history_entries': history_entries,
            'history_by_fy': history_by_fy,
            'chart_data': chart_data,
            'chart_data_json': json.dumps(chart_data),
            'fiscal_years': fiscal_years,
            'months_list': months_list,
            'dc_list': dc_list,
            'show_dc_column': show_dc_column,
            'summary': {
                'count': len(history_entries),
                'total_received': total_received,
                'total_loss': total_loss,
                'avg_loss_pct': avg_loss_pct,
            },
            'filters': {
                'start_date': start_date or '',
                'end_date': end_date or '',
                'fy_id': fy_id or '',
                'month': month or '',
                'dc_id': dc_id or '',
                'view': view_mode,
            },
            'active_fy': active_fy,
        })
    
    def _prepare_chart_data(self, history_entries, comparison_entries=None):
        """Prepare data for charts"""
        # Use comparison_entries for DC comparison if provided (for DC users)
        dc_comparison_entries = comparison_entries if comparison_entries else history_entries
        
        # Group by month for trend analysis
        monthly_data = {}
        for entry in history_entries:
            month_key = f"{entry.fiscal_year.year_bs}-{entry.month}"
            if month_key not in monthly_data:
                monthly_data[month_key] = {
                    'month': entry.get_month_display(),
                    'fiscal_year': entry.fiscal_year.year_bs,
                    'total_received': 0,
                    'total_loss': 0,
                    'total_utilised': 0,
                    'loss_percent': 0,
                    'total_consumers': 0,
                }
            monthly_data[month_key]['total_received'] += float(entry.total_received_kwh)
            monthly_data[month_key]['total_loss'] += float(entry.total_loss_kwh)
            monthly_data[month_key]['total_utilised'] += float(entry.total_received_kwh) - float(entry.total_loss_kwh)
            monthly_data[month_key]['total_consumers'] += entry.total_consumers
        
        # Calculate percentages
        for key, data in monthly_data.items():
            if data['total_received'] > 0:
                data['loss_percent'] = round((data['total_loss'] / data['total_received']) * 100, 2)
        
        # Group by DC for comparison (using comparison_entries for province-level comparison)
        dc_data = {}
        for entry in dc_comparison_entries:
            dc_name = entry.distribution_center.name
            if dc_name not in dc_data:
                dc_data[dc_name] = {
                    'name': dc_name,
                    'total_received': 0,
                    'total_loss': 0,
                    'total_utilised': 0,
                    'loss_percent': 0,
                }
            dc_data[dc_name]['total_received'] += float(entry.total_received_kwh)
            dc_data[dc_name]['total_loss'] += float(entry.total_loss_kwh)
            dc_data[dc_name]['total_utilised'] += float(entry.total_received_kwh) - float(entry.total_loss_kwh)
        
        for key, data in dc_data.items():
            if data['total_received'] > 0:
                data['loss_percent'] = round((data['total_loss'] / data['total_received']) * 100, 2)
        
        # Group by fiscal year for year-over-year analysis
        yearly_data = {}
        for entry in history_entries:
            fy_key = entry.fiscal_year.year_bs
            if fy_key not in yearly_data:
                yearly_data[fy_key] = {
                    'fiscal_year': fy_key,
                    'total_received': 0,
                    'total_loss': 0,
                    'loss_percent': 0,
                }
            yearly_data[fy_key]['total_received'] += float(entry.total_received_kwh)
            yearly_data[fy_key]['total_loss'] += float(entry.total_loss_kwh)
        
        for key, data in yearly_data.items():
            if data['total_received'] > 0:
                data['loss_percent'] = round((data['total_loss'] / data['total_received']) * 100, 2)
        
        # Energy distribution for pie chart
        total_received = sum(float(e.total_received_kwh) for e in history_entries)
        total_loss = sum(float(e.total_loss_kwh) for e in history_entries)
        total_utilised = total_received - total_loss
        
        energy_distribution = {
            'received': total_received,
            'utilised': total_utilised,
            'loss': total_loss,
        }
        
        return {
            'monthly_trend': list(monthly_data.values()),
            'dc_comparison': list(dc_data.values()),
            'yearly_trend': list(yearly_data.values()),
            'energy_distribution': energy_distribution,
        }


@login_required
def api_history_data(request):
    """API endpoint for history chart data"""
    user = request.user
    
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    dc_filter = request.GET.get('dc')
    
    history_entries = DCHistoryEntry.objects.select_related(
        'distribution_center', 'fiscal_year'
    )
    
    if dc_filter:
        history_entries = history_entries.filter(distribution_center_id=dc_filter)
    elif user.is_dc_level and user.distribution_center:
        history_entries = history_entries.filter(distribution_center=user.distribution_center)
    elif user.is_provincial and user.provincial_office:
        history_entries = history_entries.filter(
            distribution_center__provincial_office=user.provincial_office
        )
    
    if start_date:
        history_entries = history_entries.filter(report_created_at__gte=start_date)
    if end_date:
        history_entries = history_entries.filter(report_created_at__lte=end_date)
    
    history_entries = history_entries.order_by('report_created_at')
    
    # Prepare data for charts
    labels = []
    loss_data = []
    received_data = []
    
    for entry in history_entries:
        labels.append(f"{entry.get_month_display()} {entry.fiscal_year.year_bs}")
        loss_data.append(float(entry.loss_percent))
        received_data.append(float(entry.total_received_kwh))
    
    return JsonResponse({
        'labels': labels,
        'loss_data': loss_data,
        'received_data': received_data,
    })


# ─────────────────────────── DCS LIST VIEW FOR PROVINCE/DMD/MD ───────────────────────────

class DCsListView(LoginRequiredMixin, View):
    """List of DCs for Province, DMD, and MD users to view DCS details"""
    template_name = 'nea_loss/dcs/dcs_list.html'

    def get(self, request):
        user = request.user
        
        # Filter DCs based on user role
        if user.is_provincial and user.provincial_office:
            # Province users see DCs under their provincial office
            dcs = DistributionCenter.objects.filter(
                provincial_office=user.provincial_office,
                is_active=True
            ).select_related('provincial_office', 'provincial_office__province').order_by('name')
        elif user.is_top_management:
            # MD/DMD/DIRECTOR see all DCs
            dcs = DistributionCenter.objects.filter(
                is_active=True
            ).select_related('provincial_office', 'provincial_office__province').order_by('provincial_office__province', 'name')
        else:
            messages.error(request, 'Access denied.')
            return redirect('dashboard')
        
        # Get DCS details for each DC
        dc_details = []
        for dc in dcs:
            dcs_detail = DCSDetail.objects.filter(distribution_center=dc).first()
            dc_details.append({
                'dc': dc,
                'dcs_detail': dcs_detail,
                'has_detail': dcs_detail is not None,
                'province': dc.provincial_office.province.name if dc.provincial_office else 'N/A',
                'provincial_office': dc.provincial_office.name if dc.provincial_office else 'N/A',
            })
        
        return render(request, self.template_name, {
            'dc_details': dc_details,
            'user_role': user.role,
        })
